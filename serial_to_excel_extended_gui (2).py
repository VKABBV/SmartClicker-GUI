#!/usr/bin/env python3
"""Launch the UWB capture GUI with added measurement controls.

This file extends the source GUI in serial_to_excel(New).py. The original
layout is kept; the extensions add:

- a "Constellation Changed" button next to Start/Stop Logging
- an "Export Measurement List" button in Storage
- per-anchor true-distance controls in Measurement Session
- per-anchor Excel export metadata
"""

from __future__ import annotations

import importlib.machinery
import importlib.util
import math
import re
import sqlite3
import statistics
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any
from tkinter import simpledialog

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

sys.dont_write_bytecode = True


APP_DIR = Path(__file__).resolve().parent
BASE_DIR = APP_DIR.parent if APP_DIR.name == "__pycache__" else APP_DIR
ORIGINAL_SOURCE = BASE_DIR / "serial_to_excel(New).py"


def load_original_gui_module():
    if not ORIGINAL_SOURCE.exists():
        raise FileNotFoundError(
            "Base GUI source not found. Expected serial_to_excel(New).py at "
            f"{ORIGINAL_SOURCE}"
        )

    loader = importlib.machinery.SourceFileLoader(
        "_original_uwb_capture_gui",
        str(ORIGINAL_SOURCE),
    )
    spec = importlib.util.spec_from_loader(loader.name, loader)
    if spec is None:
        raise RuntimeError("Could not create loader spec for the original GUI.")

    module = importlib.util.module_from_spec(spec)
    sys.modules[loader.name] = module
    loader.exec_module(module)
    return module


original = load_original_gui_module()
tk = original.tk
ttk = original.ttk
filedialog = original.filedialog
messagebox = original.messagebox

SOURCE_DIRECT = "direct laser"
SOURCE_PYTHAGOREAN = "pythagorean calculation"
SOURCE_LEGACY = "legacy session laser"
ANCHOR_TRUTH_TABLE = "anchor_true_distances"
ILLEGAL_EXCEL_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def safe_float(value: Any) -> float | None:
    try:
        if value in (None, ""):
            return None
        out = float(value)
        if math.isnan(out) or math.isinf(out):
            return None
        return out
    except (TypeError, ValueError):
        return None


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned.strip("._") or "measurement_list"


def format_meter(value: Any, unknown: str = "unknown") -> str:
    number = safe_float(value)
    if number is None:
        return unknown
    return f"{number:.3f}".rstrip("0").rstrip(".")


def excel_cell(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_EXCEL_CHARS_RE.sub("", value)
    return value


def excel_row(values: list[Any]) -> list[Any]:
    return [excel_cell(value) for value in values]


def condition_label_from_session(session: sqlite3.Row) -> str:
    label = condition_text(session)
    return label.replace("/", "_") if label else "UNLABELED"


def walk_widgets(widget):
    for child in widget.winfo_children():
        yield child
        yield from walk_widgets(child)


def find_labelframe(root, text: str):
    for widget in walk_widgets(root):
        try:
            if widget.cget("text") == text:
                return widget
        except tk.TclError:
            continue
    return None


def next_constellation_label(current: str) -> str:
    match = re.search(r"^(.*?)(\d+)$", current.strip())
    if not match:
        return f"{current.strip() or 'constellation'}_2"
    prefix, number = match.groups()
    return f"{prefix}{int(number) + 1}"


class ExtendedUwbCaptureApp(original.UwbCaptureApp):
    """Original GUI plus targeted measurement workflow extensions."""

    def __init__(self) -> None:
        super().__init__()
        self.anchor_true_distances: dict[str, dict[str, Any]] = {}
        self.detected_anchor_ids: set[str] = set()
        self._pyth_updating = False
        self._install_extensions()

    def _build_session_panel(self, parent: Any) -> None:
        frame = ttk.LabelFrame(parent, text="Measurement Session", padding=8)
        frame.pack(fill=tk.X, pady=(0, 8))
        frame.columnconfigure(1, weight=1)

        rows = [
            ("LOS alert threshold m", self.threshold_var),
            ("Stability alert std m", self.instability_threshold_var),
        ]
        for row_index, (label, variable) in enumerate(rows):
            ttk.Label(frame, text=label).grid(row=row_index, column=0, sticky="w", pady=2)
            ttk.Entry(frame, textvariable=variable).grid(
                row=row_index,
                column=1,
                columnspan=2,
                sticky="ew",
                pady=2,
            )

        self.constellation_status_var = tk.StringVar(
            value=f"Constellation: {self.constellation_var.get().strip() or 'not set'}"
        )

        condition_row = ttk.Frame(frame)
        condition_row.grid(row=2, column=0, columnspan=3, sticky="w", pady=(8, 0))
        ttk.Checkbutton(condition_row, text="LOS", variable=self.los_var).pack(side=tk.LEFT)
        ttk.Checkbutton(condition_row, text="NLOS", variable=self.nlos_var).pack(side=tk.LEFT, padx=(14, 0))

        action_row = ttk.Frame(frame)
        action_row.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(10, 0))
        action_row.columnconfigure(0, weight=1)
        action_row.columnconfigure(1, weight=1)
        self.start_button = ttk.Button(
            action_row,
            text="Start Logging",
            style="Accent.TButton",
            command=self.start_capture,
        )
        self.start_button.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        self.stop_button = ttk.Button(action_row, text="Stop Logging", command=self.stop_capture)
        self.stop_button.grid(row=0, column=1, sticky="ew", padx=(5, 0))

        ttk.Label(frame, textvariable=self.constellation_status_var).grid(
            row=4,
            column=0,
            columnspan=3,
            sticky="w",
            pady=(8, 0),
        )
        ttk.Label(frame, textvariable=self.session_status_var, style="Status.TLabel").grid(
            row=5,
            column=0,
            columnspan=3,
            sticky="w",
            pady=(4, 0),
        )

    def _install_extensions(self) -> None:
        self._add_constellation_button()
        self._add_anchor_true_distance_controls()
        self._add_measurement_list_button()

    def _add_constellation_button(self) -> None:
        action_row = getattr(self, "start_button", None)
        if action_row is None:
            return

        parent = self.start_button.master
        parent.columnconfigure(2, weight=1)
        self.constellation_changed_button = ttk.Button(
            parent,
            text="Constellation",
            command=self.open_constellation_dialog,
        )
        self.constellation_changed_button.grid(
            row=0,
            column=2,
            sticky="ew",
            padx=(5, 0),
        )

    def _add_anchor_true_distance_controls(self) -> None:
        session_frame = find_labelframe(self, "Measurement Session")
        if session_frame is None:
            return

        self.anchor_id_var = tk.StringVar()
        self.anchor_true_distance_var = tk.StringVar()
        self.pyth_side_a_var = tk.StringVar()
        self.pyth_side_b_var = tk.StringVar()
        self.pyth_result_var = tk.StringVar(value="")

        rows = [
            int(child.grid_info().get("row", 0))
            for child in session_frame.grid_slaves()
            if child.grid_info()
        ]
        row_index = (max(rows) + 1) if rows else 0
        session_frame.columnconfigure(2, weight=1)

        container = ttk.LabelFrame(
            session_frame,
            text="Per-anchor true distance",
            padding=8,
        )
        container.grid(
            row=row_index,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=(8, 0),
        )
        container.columnconfigure(1, weight=1)
        container.columnconfigure(3, weight=1)

        ttk.Label(container, text="Anchor ID").grid(row=0, column=0, sticky="w")
        self.anchor_id_combo = ttk.Combobox(
            container,
            textvariable=self.anchor_id_var,
            values=[],
            width=16,
            state="normal",
        )
        self.anchor_id_combo.grid(row=0, column=1, sticky="ew", padx=(6, 12))

        ttk.Label(container, text="True distance m").grid(row=0, column=2, sticky="w")
        ttk.Entry(container, textvariable=self.anchor_true_distance_var, width=14).grid(
            row=0,
            column=3,
            sticky="ew",
            padx=(6, 12),
        )
        ttk.Button(
            container,
            text="Save direct distance",
            command=self.save_direct_true_distance,
        ).grid(row=0, column=4, sticky="ew")

        helper = ttk.Frame(container)
        helper.grid(row=1, column=0, columnspan=5, sticky="ew", pady=(8, 0))
        for col in (1, 3, 5):
            helper.columnconfigure(col, weight=1)

        ttk.Label(helper, text="Side A m").grid(row=0, column=0, sticky="w")
        ttk.Entry(helper, textvariable=self.pyth_side_a_var, width=10).grid(
            row=0,
            column=1,
            sticky="ew",
            padx=(6, 12),
        )
        ttk.Label(helper, text="Side B m").grid(row=0, column=2, sticky="w")
        ttk.Entry(helper, textvariable=self.pyth_side_b_var, width=10).grid(
            row=0,
            column=3,
            sticky="ew",
            padx=(6, 12),
        )
        ttk.Label(helper, text="Result m").grid(row=0, column=4, sticky="w")
        ttk.Entry(helper, textvariable=self.pyth_result_var, width=10, state="readonly").grid(
            row=0,
            column=5,
            sticky="ew",
            padx=(6, 0),
        )

        self.pyth_canvas = tk.Canvas(
            helper,
            height=150,
            background="#ffffff",
            highlightthickness=1,
            highlightbackground="#c8c8c8",
        )
        self.pyth_canvas.grid(row=1, column=0, columnspan=6, sticky="ew", pady=(8, 0))
        self.pyth_canvas.bind("<Configure>", lambda _event: self._draw_pythagorean_triangle())
        self.pyth_canvas.bind("<Button-1>", self._set_pythagorean_from_canvas)
        self.pyth_canvas.bind("<B1-Motion>", self._set_pythagorean_from_canvas)

        ttk.Button(helper, text="Calculate", command=self.calculate_pythagorean_distance).grid(
            row=2,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(8, 0),
        )
        ttk.Button(
            helper,
            text="Use calculated distance for selected anchor",
            command=self.use_pythagorean_distance_for_anchor,
        ).grid(row=2, column=2, columnspan=4, sticky="ew", padx=(8, 0), pady=(8, 0))

        self.anchor_truth_table = ttk.Treeview(
            container,
            columns=("anchor_id", "true_distance_m", "distance_source"),
            show="headings",
            height=4,
        )
        self.anchor_truth_table.heading("anchor_id", text="Anchor ID")
        self.anchor_truth_table.heading("true_distance_m", text="True distance m")
        self.anchor_truth_table.heading("distance_source", text="True distance source")
        self.anchor_truth_table.column("anchor_id", width=90, anchor="center")
        self.anchor_truth_table.column("true_distance_m", width=130, anchor="center")
        self.anchor_truth_table.column("distance_source", width=190, anchor="w")
        self.anchor_truth_table.grid(row=2, column=0, columnspan=5, sticky="ew", pady=(8, 0))
        self.anchor_truth_table.bind("<<TreeviewSelect>>", self.on_anchor_truth_selected)

        self.pyth_side_a_var.trace_add("write", self._on_pythagorean_sides_changed)
        self.pyth_side_b_var.trace_add("write", self._on_pythagorean_sides_changed)
        self._draw_pythagorean_triangle()

    def _add_measurement_list_button(self) -> None:
        storage = find_labelframe(self, "Storage")
        if storage is None:
            return

        self.export_measurement_list_button = ttk.Button(
            storage,
            text="Export Measurement List",
            command=self.export_measurement_list,
        )
        self.export_measurement_list_button.grid(
            row=3,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=(8, 0),
        )

    def _selected_anchor_id(self) -> str | None:
        anchor_id = self.anchor_id_var.get().strip()
        if not anchor_id:
            messagebox.showerror(
                "Missing anchor ID",
                "Select or enter an Anchor ID before saving a true distance.",
                parent=self,
            )
            return None
        return anchor_id

    def _positive_float_from_var(self, variable: Any, label: str) -> float | None:
        value = safe_float(variable.get())
        if value is None or value <= 0:
            messagebox.showerror(
                "Invalid value",
                f"{label} must be numeric and greater than 0.",
                parent=self,
            )
            return None
        return value

    def _on_pythagorean_sides_changed(self, *_args: Any) -> None:
        if self._pyth_updating:
            return
        self._update_pythagorean_from_sides()

    def _update_pythagorean_from_sides(self) -> float | None:
        side_a = safe_float(self.pyth_side_a_var.get())
        side_b = safe_float(self.pyth_side_b_var.get())
        if side_a is None or side_a <= 0 or side_b is None or side_b <= 0:
            self._pyth_updating = True
            try:
                self.pyth_result_var.set("")
            finally:
                self._pyth_updating = False
            self._draw_pythagorean_triangle()
            return None

        result = math.sqrt(side_a**2 + side_b**2)
        self._pyth_updating = True
        try:
            self.pyth_result_var.set(format_meter(result))
        finally:
            self._pyth_updating = False
        self._draw_pythagorean_triangle()
        return result

    def _triangle_plot_area(self) -> tuple[int, int, int, int]:
        canvas = self.pyth_canvas
        width = max(canvas.winfo_width(), 280)
        height = max(canvas.winfo_height(), 150)
        left = 28
        top = 18
        right = width - 24
        bottom = height - 32
        return left, top, right, bottom

    def _draw_pythagorean_triangle(self) -> None:
        if not hasattr(self, "pyth_canvas"):
            return

        canvas = self.pyth_canvas
        canvas.delete("all")
        side_a = safe_float(self.pyth_side_a_var.get())
        side_b = safe_float(self.pyth_side_b_var.get())
        if side_a is None or side_a <= 0:
            side_a = 3.0
        if side_b is None or side_b <= 0:
            side_b = 2.0

        left, top, right, bottom = self._triangle_plot_area()
        plot_w = max(right - left, 1)
        plot_h = max(bottom - top, 1)
        scale = min(plot_w / side_a, plot_h / side_b)
        x0 = left
        y0 = bottom
        x1 = left + side_a * scale
        y1 = bottom
        x2 = x1
        y2 = bottom - side_b * scale
        hypotenuse = math.sqrt(side_a**2 + side_b**2)

        canvas.create_line(x0, y0, x1, y1, fill="#1f77b4", width=3)
        canvas.create_line(x1, y1, x2, y2, fill="#2ca02c", width=3)
        canvas.create_line(x0, y0, x2, y2, fill="#d62728", width=3)
        canvas.create_rectangle(x1 - 14, y1 - 14, x1, y1, outline="#666666")
        canvas.create_oval(x2 - 6, y2 - 6, x2 + 6, y2 + 6, fill="#d97706", outline="")

        canvas.create_text((x0 + x1) / 2, y1 + 14, text=f"A {format_meter(side_a)} m", fill="#1f77b4")
        canvas.create_text(x1 + 36, (y1 + y2) / 2, text=f"B {format_meter(side_b)} m", fill="#2ca02c")
        canvas.create_text((x0 + x2) / 2 - 8, (y0 + y2) / 2 - 10, text=f"true {format_meter(hypotenuse)} m", fill="#d62728")
        canvas.create_text(right - 64, top + 10, text="drag point", fill="#666666")

    def _set_pythagorean_from_canvas(self, event: Any) -> None:
        left, top, right, bottom = self._triangle_plot_area()
        plot_w = max(right - left, 1)
        plot_h = max(bottom - top, 1)
        x = min(max(float(event.x), left + 4), right)
        y = min(max(float(event.y), top), bottom - 4)

        current_a = safe_float(self.pyth_side_a_var.get()) or 0.0
        current_b = safe_float(self.pyth_side_b_var.get()) or 0.0
        current_result = safe_float(self.pyth_result_var.get()) or 0.0
        max_meter = max(current_a, current_b, current_result, 5.0)
        side_a = max(0.01, ((x - left) / plot_w) * max_meter)
        side_b = max(0.01, ((bottom - y) / plot_h) * max_meter)
        result = math.sqrt(side_a**2 + side_b**2)

        self._pyth_updating = True
        try:
            self.pyth_side_a_var.set(format_meter(side_a))
            self.pyth_side_b_var.set(format_meter(side_b))
            self.pyth_result_var.set(format_meter(result))
        finally:
            self._pyth_updating = False
        self._draw_pythagorean_triangle()

    def save_direct_true_distance(self) -> None:
        anchor_id = self._selected_anchor_id()
        if anchor_id is None:
            return
        distance = self._positive_float_from_var(
            self.anchor_true_distance_var,
            "Direct true distance",
        )
        if distance is None:
            return
        self._save_anchor_true_distance(
            anchor_id,
            distance,
            SOURCE_DIRECT,
            None,
            None,
        )

    def calculate_pythagorean_distance(self) -> float | None:
        side_a = self._positive_float_from_var(self.pyth_side_a_var, "Pythagorean Side A")
        side_b = self._positive_float_from_var(self.pyth_side_b_var, "Pythagorean Side B")
        if side_a is None or side_b is None:
            return None
        result = math.sqrt(side_a**2 + side_b**2)
        self._pyth_updating = True
        try:
            self.pyth_result_var.set(format_meter(result))
        finally:
            self._pyth_updating = False
        self._draw_pythagorean_triangle()
        return result

    def use_pythagorean_distance_for_anchor(self) -> None:
        anchor_id = self._selected_anchor_id()
        if anchor_id is None:
            return
        side_a = self._positive_float_from_var(self.pyth_side_a_var, "Pythagorean Side A")
        side_b = self._positive_float_from_var(self.pyth_side_b_var, "Pythagorean Side B")
        if side_a is None or side_b is None:
            return
        result = math.sqrt(side_a**2 + side_b**2)
        self.pyth_result_var.set(format_meter(result))
        self.anchor_true_distance_var.set(format_meter(result))
        self._save_anchor_true_distance(
            anchor_id,
            result,
            SOURCE_PYTHAGOREAN,
            side_a,
            side_b,
        )

    def _save_anchor_true_distance(
        self,
        anchor_id: str,
        distance: float,
        source: str,
        side_a: float | None,
        side_b: float | None,
    ) -> None:
        self.anchor_true_distances[anchor_id] = {
            "anchor_id": anchor_id,
            "true_distance_m": distance,
            "distance_source": source,
            "pythagorean_side_a_m": side_a,
            "pythagorean_side_b_m": side_b,
        }
        self._register_detected_anchor(anchor_id)
        self._refresh_anchor_truth_table()
        self._persist_anchor_true_distance(anchor_id)
        self.log_raw(
            f"# True distance saved for anchor {anchor_id}: {format_meter(distance)} m ({source})"
        )

    def on_anchor_truth_selected(self, _event: Any = None) -> None:
        selected = self.anchor_truth_table.selection()
        if not selected:
            return
        values = self.anchor_truth_table.item(selected[0], "values")
        if not values:
            return
        anchor_id = str(values[0])
        data = self.anchor_true_distances.get(anchor_id)
        if not data:
            return
        self.anchor_id_var.set(anchor_id)
        self.anchor_true_distance_var.set(format_meter(data["true_distance_m"]))
        self.pyth_side_a_var.set(
            "" if data.get("pythagorean_side_a_m") is None else format_meter(data["pythagorean_side_a_m"])
        )
        self.pyth_side_b_var.set(
            "" if data.get("pythagorean_side_b_m") is None else format_meter(data["pythagorean_side_b_m"])
        )
        if data.get("distance_source") == SOURCE_PYTHAGOREAN:
            self.pyth_result_var.set(format_meter(data["true_distance_m"]))
        else:
            self.pyth_result_var.set("")
        self._draw_pythagorean_triangle()

    def _refresh_anchor_truth_table(self) -> None:
        if not hasattr(self, "anchor_truth_table"):
            return
        for item in self.anchor_truth_table.get_children():
            self.anchor_truth_table.delete(item)
        for anchor_id in sorted(self.anchor_true_distances, key=str):
            data = self.anchor_true_distances[anchor_id]
            self.anchor_truth_table.insert(
                "",
                tk.END,
                values=(
                    anchor_id,
                    format_meter(data["true_distance_m"]),
                    data["distance_source"],
                ),
            )

    def _register_detected_anchor(self, anchor_id: Any) -> None:
        if anchor_id in (None, ""):
            return
        anchor_text = str(anchor_id).strip()
        if not anchor_text:
            return
        self.detected_anchor_ids.add(anchor_text)
        if hasattr(self, "anchor_id_combo"):
            self.anchor_id_combo.configure(values=sorted(self.detected_anchor_ids, key=str))

    def mark_constellation_changed(self) -> None:
        self.open_constellation_dialog()

    def open_constellation_dialog(self) -> None:
        current = self.constellation_var.get().strip()
        window = tk.Toplevel(self)
        window.title("Constellation")
        window.transient(self)
        window.grab_set()
        window.columnconfigure(0, weight=1)

        notebook = ttk.Notebook(window)
        notebook.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        tab = ttk.Frame(notebook, padding=10)
        tab.columnconfigure(1, weight=1)
        notebook.add(tab, text="Constellation label")

        mode_var = tk.StringVar(value="same" if current else "new")
        label_var = tk.StringVar(value=current or "constellation_1")

        ttk.Label(tab, text="Is this still the same constellation?").grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(0, 8),
        )
        ttk.Radiobutton(tab, text="Yes, same constellation", variable=mode_var, value="same").grid(
            row=1,
            column=0,
            columnspan=2,
            sticky="w",
        )
        ttk.Radiobutton(tab, text="No, new constellation", variable=mode_var, value="new").grid(
            row=2,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(0, 8),
        )
        ttk.Label(tab, text="Constellation label").grid(row=3, column=0, sticky="w", pady=2)
        label_entry = ttk.Entry(tab, textvariable=label_var, width=34)
        label_entry.grid(row=3, column=1, sticky="ew", pady=2)
        label_entry.focus_set()

        button_row = ttk.Frame(tab)
        button_row.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        button_row.columnconfigure(0, weight=1)
        ttk.Button(
            button_row,
            text="Save",
            command=lambda: self._apply_constellation_choice(mode_var.get(), label_var.get(), window),
        ).grid(row=0, column=0, sticky="e", padx=(0, 6))
        ttk.Button(button_row, text="Cancel", command=window.destroy).grid(row=0, column=1, sticky="e")

    def _apply_constellation_choice(self, mode: str, label: str, window: Any) -> None:
        current = self.constellation_var.get().strip()
        label = label.strip()
        if mode == "same" and not label:
            label = current
        if not label:
            messagebox.showerror(
                "Missing constellation label",
                "Enter a constellation label before saving.",
                parent=window,
            )
            return

        if mode == "same":
            self.constellation_var.set(label)
            self._set_constellation_status(label)
            self.log_raw(f"# Constellation confirmed unchanged: {label}")
            window.destroy()
            return

        was_capturing = bool(getattr(self, "capture_active", False))
        if was_capturing:
            proceed = messagebox.askyesno(
                "New constellation",
                "Stop the current logging session and start a new one with this constellation label?",
                parent=window,
            )
            if not proceed:
                return
            self.stop_capture()

        self.constellation_var.set(label)
        self._set_constellation_status(label)
        self.log_raw(f"# New constellation label: {label}")
        window.destroy()

        if was_capturing:
            self.start_capture()

    def _set_constellation_status(self, label: str) -> None:
        if hasattr(self, "constellation_status_var"):
            self.constellation_status_var.set(f"Constellation: {label or 'not set'}")

    def _ensure_anchor_truth_schema(self) -> None:
        store = getattr(self, "store", None)
        if store is None:
            return
        store.conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {ANCHOR_TRUTH_TABLE} (
                session_id TEXT NOT NULL,
                anchor_id TEXT NOT NULL,
                true_distance_m REAL NOT NULL,
                distance_source TEXT NOT NULL,
                pythagorean_side_a_m REAL,
                pythagorean_side_b_m REAL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (session_id, anchor_id),
                FOREIGN KEY(session_id) REFERENCES sessions(id)
            )
            """
        )
        store.conn.commit()

    def _persist_anchor_true_distance(self, anchor_id: str) -> None:
        store = getattr(self, "store", None)
        session_id = getattr(self, "current_session_id", None)
        if store is None or not session_id:
            return
        data = self.anchor_true_distances.get(anchor_id)
        if not data:
            return
        self._ensure_anchor_truth_schema()
        store.conn.execute(
            f"""
            INSERT OR REPLACE INTO {ANCHOR_TRUTH_TABLE} (
                session_id, anchor_id, true_distance_m, distance_source,
                pythagorean_side_a_m, pythagorean_side_b_m, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session_id,
                anchor_id,
                data["true_distance_m"],
                data["distance_source"],
                data.get("pythagorean_side_a_m"),
                data.get("pythagorean_side_b_m"),
                original.now_local_iso(),
            ),
        )
        store.conn.commit()

    def _persist_all_anchor_true_distances(self) -> None:
        for anchor_id in list(self.anchor_true_distances):
            self._persist_anchor_true_distance(anchor_id)

    def _anchor_true_distance_for(self, anchor_id: Any) -> dict[str, Any] | None:
        if anchor_id not in (None, ""):
            data = self.anchor_true_distances.get(str(anchor_id).strip())
            if data:
                return data
        return None

    def start_capture(self) -> None:
        super().start_capture()
        if getattr(self, "store", None) is None or not getattr(self, "current_session_id", None):
            return
        self._ensure_anchor_truth_schema()
        self._persist_all_anchor_true_distances()

    def handle_serial_line(self, line: str) -> None:
        for record in original.parse_serial_line(line):
            if record.anchor_id:
                self._register_detected_anchor(record.anchor_id)
        super().handle_serial_line(line)

    def check_los_alert(self, record: Any) -> float | None:
        distance = record.distance_m if record.distance_m is not None else record.mean_distance_m
        if distance is None:
            return None

        truth = self._anchor_true_distance_for(record.anchor_id)
        if not truth:
            return None

        true_distance = truth["true_distance_m"]
        absolute_error = abs(float(distance) - float(true_distance))
        threshold = safe_float(self.threshold_var.get()) or 0.0

        if (
            getattr(self, "store", None) is not None
            and getattr(self, "current_session_id", None)
            and self.los_var.get()
            and not self.nlos_var.get()
            and absolute_error > threshold
        ):
            anchor = record.anchor_id or "unknown"
            message = (
                f"LOS measurement from anchor {anchor} is {absolute_error:.3f} m away "
                f"from true distance ({float(distance):.3f} m measured vs "
                f"{float(true_distance):.3f} m true)."
            )
            self.store.insert_alert(
                self.current_session_id,
                anchor,
                float(distance),
                float(true_distance),
                absolute_error,
                message,
            )
            self.log_alert(message)

            current_time = time.monotonic()
            last_prompt = self.last_alert_prompt.get(anchor, 0.0)
            if current_time - last_prompt > original.ALERT_PROMPT_COOLDOWN_SECONDS:
                self.last_alert_prompt[anchor] = current_time
                messagebox.showwarning("LOS measurement warning", message, parent=self)
        return absolute_error

    def add_record_to_table(self, record: Any, alert_error: float | None = None) -> None:
        display_time = datetime.now().strftime("%H:%M:%S")
        distance = record.distance_m if record.distance_m is not None else record.mean_distance_m
        truth = self._anchor_true_distance_for(record.anchor_id)
        error = alert_error
        if error is None and truth and distance is not None:
            error = abs(float(distance) - float(truth["true_distance_m"]))

        row = (
            display_time,
            record.anchor_id or "",
            "" if record.sample_index is None else str(record.sample_index),
            "" if distance is None else f"{float(distance):.3f}",
            "" if error is None else f"{float(error):.3f}",
            "" if record.rx_power_dbm is None else f"{float(record.rx_power_dbm):.2f}",
            "" if record.fp_power_dbm is None else f"{float(record.fp_power_dbm):.2f}",
            "" if record.cir_power is None else f"{float(record.cir_power):.2f}",
            record.status or "",
            record.source or "",
        )

        tags: tuple[str, ...] = ()
        threshold = safe_float(self.threshold_var.get())
        if error is not None and threshold is not None and error > threshold:
            tags = ("alert",)
        if record.kind == "failure":
            tags = ("failure",)
        elif record.kind == "summary":
            tags = ("summary",)

        self.sample_table.insert("", tk.END, values=row, tags=tags)
        self.table_rows += 1
        if self.table_rows > original.MAX_TABLE_ROWS:
            first = self.sample_table.get_children()[0]
            self.sample_table.delete(first)
            self.table_rows -= 1
        self.sample_table.yview_moveto(1.0)

    def export_current_session(self) -> None:
        if getattr(self, "store", None) is None or not getattr(self, "current_session_id", None):
            messagebox.showwarning(
                "No session",
                "Start a capture session before exporting.",
                parent=self,
            )
            return

        if getattr(self, "capture_active", False):
            proceed = messagebox.askyesno(
                "Session still running",
                "The session is still capturing. Export a snapshot of the data collected so far?",
                parent=self,
            )
            if not proceed:
                return

        self._ensure_anchor_truth_schema()
        self._persist_all_anchor_true_distances()

        output_dir = filedialog.askdirectory(
            initialdir=self.output_dir_var.get(),
            title="Choose folder for per-anchor Excel exports",
            parent=self,
        )
        if not output_dir:
            return

        try:
            paths = export_session_per_anchor(
                self.store.conn,
                self.current_session_id,
                Path(output_dir),
            )
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc), parent=self)
            return

        if not paths:
            messagebox.showwarning(
                "No measurements",
                "The current session does not contain anchor measurements yet.",
                parent=self,
            )
            return

        messagebox.showinfo(
            "Export complete",
            "Wrote:\n" + "\n".join(str(path) for path in paths),
            parent=self,
        )

    def export_measurement_list(self) -> None:
        db_path = Path(self.db_path_var.get()).expanduser()
        if not db_path.exists():
            messagebox.showwarning(
                "No measurements",
                "No SQLite capture database exists yet. Start logging first.",
                parent=self,
            )
            return

        rows = build_measurement_rows(db_path)
        if not rows:
            messagebox.showwarning(
                "No measurements",
                "The current capture database does not contain measurement rows yet.",
                parent=self,
            )
            return

        default_name = f"{safe_filename(db_path.stem)}_ground_truth_measurement_list.xlsx"
        output_path = filedialog.asksaveasfilename(
            initialdir=self.output_dir_var.get(),
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            parent=self,
        )
        if not output_path:
            return

        try:
            write_measurement_workbook(rows, Path(output_path))
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc), parent=self)
            return

        messagebox.showinfo("Export complete", f"Wrote {output_path}", parent=self)


def table_exists(con: sqlite3.Connection, table_name: str) -> bool:
    row = con.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def fetch_anchor_truths(con: sqlite3.Connection, session_id: str) -> dict[str, dict[str, Any]]:
    if not table_exists(con, ANCHOR_TRUTH_TABLE):
        return {}
    rows = con.execute(
        f"""
        SELECT anchor_id, true_distance_m, distance_source,
               pythagorean_side_a_m, pythagorean_side_b_m
        FROM {ANCHOR_TRUTH_TABLE}
        WHERE session_id = ?
        """,
        (session_id,),
    ).fetchall()
    return {
        str(row["anchor_id"]): {
            "anchor_id": str(row["anchor_id"]),
            "true_distance_m": row["true_distance_m"],
            "distance_source": row["distance_source"],
            "pythagorean_side_a_m": row["pythagorean_side_a_m"],
            "pythagorean_side_b_m": row["pythagorean_side_b_m"],
        }
        for row in rows
    }


def session_anchor_ids(con: sqlite3.Connection, session_id: str) -> list[str]:
    rows = con.execute(
        """
        SELECT anchor_id FROM samples WHERE session_id = ? AND anchor_id IS NOT NULL
        UNION
        SELECT anchor_id FROM summaries WHERE session_id = ? AND anchor_id IS NOT NULL
        ORDER BY anchor_id
        """,
        (session_id, session_id),
    ).fetchall()
    return [str(row["anchor_id"]) for row in rows if row["anchor_id"] not in (None, "")]


def measured_mean_for_anchor(
    samples: list[sqlite3.Row],
    summaries: list[sqlite3.Row],
) -> float | None:
    values = [float(row["distance_m"]) for row in samples if row["distance_m"] is not None]
    if values:
        return statistics.fmean(values)
    summary_values = [
        float(row["mean_distance_m"])
        for row in summaries
        if row["mean_distance_m"] is not None
    ]
    if summary_values:
        return summary_values[-1]
    return None


def export_session_per_anchor(
    con: sqlite3.Connection,
    session_id: str,
    output_dir: Path,
) -> list[Path]:
    session = con.execute("SELECT * FROM sessions WHERE id = ?", (session_id,)).fetchone()
    if session is None:
        raise ValueError(f"Unknown session id: {session_id}")

    truths = fetch_anchor_truths(con, session_id)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    condition = condition_label_from_session(session)
    constellation = safe_filename(session["constellation_label"] or "constellation")

    paths: list[Path] = []
    for anchor_id in session_anchor_ids(con, session_id):
        samples = con.execute(
            "SELECT * FROM samples WHERE session_id = ? AND anchor_id = ? ORDER BY id",
            (session_id, anchor_id),
        ).fetchall()
        summaries = con.execute(
            "SELECT * FROM summaries WHERE session_id = ? AND anchor_id = ? ORDER BY id",
            (session_id, anchor_id),
        ).fetchall()
        alerts = con.execute(
            "SELECT * FROM alerts WHERE session_id = ? AND anchor_id = ? ORDER BY id",
            (session_id, anchor_id),
        ).fetchall()
        raw_lines = con.execute(
            "SELECT id, timestamp, parsed, raw_line FROM raw_lines WHERE session_id = ? ORDER BY id",
            (session_id,),
        ).fetchall()

        truth = truths.get(anchor_id)
        mean_measured = measured_mean_for_anchor(samples, summaries)

        filename = (
            f"CONSTELLATION_{constellation}_"
            f"ANCHOR_{safe_filename(anchor_id)}_"
            f"TRUE_{format_meter(truth['true_distance_m'] if truth else None)}m_"
            f"MEASURED_{format_meter(mean_measured)}m_"
            f"{safe_filename(condition)}_{timestamp}.xlsx"
        )
        path = output_dir / filename
        write_anchor_session_workbook(
            session,
            anchor_id,
            truth,
            samples,
            summaries,
            alerts,
            raw_lines,
            mean_measured,
            path,
        )
        paths.append(path)
    return paths


def write_anchor_session_workbook(
    session: sqlite3.Row,
    anchor_id: str,
    truth: dict[str, Any] | None,
    samples: list[sqlite3.Row],
    summaries: list[sqlite3.Row],
    alerts: list[sqlite3.Row],
    raw_lines: list[sqlite3.Row],
    mean_measured: float | None,
    path: Path,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Samples"

    true_distance = truth["true_distance_m"] if truth else None
    true_source = truth["distance_source"] if truth else ""
    side_a = truth.get("pythagorean_side_a_m") if truth else None
    side_b = truth.get("pythagorean_side_b_m") if truth else None
    condition = condition_label_from_session(session)

    columns = [
        "timestamp",
        "anchor_id",
        "sample_index",
        "measured_uwb_distance_m",
        "true_ground_truth_distance_m",
        "true_distance_source",
        "pythagorean_side_a_m",
        "pythagorean_side_b_m",
        "los_nlos",
        "rx_dbm",
        "fp_dbm",
        "cir_power",
        "alert_metric_m",
        "status",
        "source",
        "cir_raw",
        "raw_line",
    ]
    ws.append(columns)

    for row in samples:
        measured = row["distance_m"]
        alert_metric = (
            abs(float(measured) - float(true_distance))
            if measured is not None and true_distance is not None
            else None
        )
        ws.append(
            excel_row([
                row["timestamp"],
                row["anchor_id"],
                row["sample_index"],
                measured,
                true_distance,
                true_source,
                side_a,
                side_b,
                condition,
                row["rx_power_dbm"],
                row["fp_power_dbm"],
                row["cir_power"],
                alert_metric,
                row["status"],
                row["source"],
                row["cir_raw"],
                row["raw_line"],
            ])
        )

    add_rows_sheet(
        wb,
        "Summaries",
        [
            "timestamp",
            "anchor_id",
            "mean_measured_uwb_distance_m",
            "true_ground_truth_distance_m",
            "true_distance_source",
            "pythagorean_side_a_m",
            "pythagorean_side_b_m",
            "los_nlos",
            "alert_metric_m",
            "good_count",
            "status",
            "source",
            "raw_line",
        ],
        [
            [
                row["timestamp"],
                row["anchor_id"],
                row["mean_distance_m"],
                true_distance,
                true_source,
                side_a,
                side_b,
                condition,
                abs(float(row["mean_distance_m"]) - float(true_distance))
                if row["mean_distance_m"] is not None and true_distance is not None
                else None,
                row["good_count"],
                row["status"],
                row["source"],
                row["raw_line"],
            ]
            for row in summaries
        ],
    )

    add_rows_sheet(
        wb,
        "Anchor True Distance",
        ["anchor_id", "true_distance_m", "distance_source", "pythagorean_side_a_m", "pythagorean_side_b_m"],
        [[anchor_id, true_distance, true_source, side_a, side_b]],
    )
    add_rows_sheet(
        wb,
        "Session",
        ["field", "value"],
        [
            ["session_id", session["id"]],
            ["started_at", session["started_at"]],
            ["stopped_at", session["stopped_at"]],
            ["constellation_label", session["constellation_label"]],
            ["los_nlos", condition],
            ["anchor_id", anchor_id],
            ["mean_measured_uwb_distance_m", mean_measured],
            ["true_ground_truth_distance_m", true_distance],
            ["true_distance_source", true_source],
        ],
    )
    add_rows_sheet(wb, "Alerts", list(alerts[0].keys()) if alerts else ["No rows"], alerts)
    add_rows_sheet(wb, "Raw Lines", list(raw_lines[0].keys()) if raw_lines else ["No rows"], raw_lines)

    style_workbook(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def add_rows_sheet(
    wb: openpyxl.Workbook,
    title: str,
    columns: list[str],
    rows: list[Any],
) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    if not rows:
        return
    for row in rows:
        if isinstance(row, sqlite3.Row):
            ws.append(excel_row([row[col] for col in columns]))
        else:
            ws.append(excel_row(list(row)))


def style_workbook(wb: openpyxl.Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            header = column_cells[0].value
            if not header:
                continue
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 42)


def build_measurement_rows(db_path: Path) -> list[dict[str, Any]]:
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    try:
        sessions = con.execute(
            """
            SELECT *
            FROM sessions
            ORDER BY started_at, id
            """
        ).fetchall()

        sample_rows = con.execute(
            """
            SELECT session_id, anchor_id, distance_m
            FROM samples
            WHERE distance_m IS NOT NULL
            """
        ).fetchall()

        summary_rows = con.execute(
            """
            SELECT session_id, anchor_id, mean_distance_m, good_count, status, source
            FROM summaries
            WHERE mean_distance_m IS NOT NULL
            ORDER BY timestamp
            """
        ).fetchall()

        truth_rows = (
            con.execute(
                f"""
                SELECT session_id, anchor_id, true_distance_m, distance_source,
                       pythagorean_side_a_m, pythagorean_side_b_m
                FROM {ANCHOR_TRUTH_TABLE}
                """
            ).fetchall()
            if table_exists(con, ANCHOR_TRUTH_TABLE)
            else []
        )
    finally:
        con.close()

    samples_by_key: dict[tuple[str, str], list[float]] = {}
    for row in sample_rows:
        anchor_id = row["anchor_id"] or ""
        key = (row["session_id"], anchor_id)
        samples_by_key.setdefault(key, []).append(float(row["distance_m"]))

    latest_summary_by_key: dict[tuple[str, str], sqlite3.Row] = {}
    for row in summary_rows:
        anchor_id = row["anchor_id"] or ""
        latest_summary_by_key[(row["session_id"], anchor_id)] = row

    truth_by_key: dict[tuple[str, str], sqlite3.Row] = {}
    for row in truth_rows:
        anchor_id = row["anchor_id"] or ""
        truth_by_key[(row["session_id"], anchor_id)] = row

    session_by_id = {row["id"]: row for row in sessions}
    keys = sorted(
        set(samples_by_key) | set(latest_summary_by_key),
        key=lambda item: (session_by_id[item[0]]["started_at"], item[1]),
    )

    rows: list[dict[str, Any]] = []
    for session_id, anchor_id in keys:
        session = session_by_id[session_id]
        values = samples_by_key.get((session_id, anchor_id), [])
        latest_summary = latest_summary_by_key.get((session_id, anchor_id))
        truth = truth_by_key.get((session_id, anchor_id))
        mean_distance = (
            float(latest_summary["mean_distance_m"])
            if latest_summary is not None
            else statistics.fmean(values)
            if values
            else None
        )
        rows.append(
            {
                "session_id": session_id,
                "started_at": session["started_at"],
                "stopped_at": session["stopped_at"],
                "constellation_label": session["constellation_label"],
                "condition_hint": condition_text(session),
                "tag_id": session["tag_id"] or "",
                "anchor_id": anchor_id,
                "uwb_mean_distance_m": mean_distance,
                "sample_count": len(values),
                "sample_std_m": statistics.stdev(values) if len(values) > 1 else None,
                "true_ground_truth_distance_m": truth["true_distance_m"] if truth is not None else None,
                "true_distance_source": truth["distance_source"] if truth is not None else "",
                "pythagorean_side_a_m": truth["pythagorean_side_a_m"] if truth is not None else None,
                "pythagorean_side_b_m": truth["pythagorean_side_b_m"] if truth is not None else None,
                "session_ground_truth_m": session["ground_truth_m"],
                "outlier_threshold_m": session["outlier_threshold_m"],
                "status": latest_summary["status"] if latest_summary is not None else "",
                "source": latest_summary["source"] if latest_summary is not None else "",
            }
        )
    return rows


def condition_text(session: sqlite3.Row) -> str:
    bits = []
    if session["condition_los"]:
        bits.append("LOS")
    if session["condition_nlos"]:
        bits.append("NLOS")
    return "/".join(bits)


def write_measurement_workbook(rows: list[dict[str, Any]], path: Path) -> None:
    columns = [
        "session_id",
        "started_at",
        "stopped_at",
        "constellation_label",
        "condition_hint",
        "tag_id",
        "anchor_id",
        "true_ground_truth_distance_m",
        "true_distance_source",
        "pythagorean_side_a_m",
        "pythagorean_side_b_m",
        "uwb_mean_distance_m",
        "range_minus_true_m",
        "ground_truth_los_nlos",
        "session_ground_truth_m",
        "outlier_threshold_m",
        "sample_count",
        "sample_std_m",
        "status",
        "source",
        "notes",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Measurements"
    ws.append(columns)

    for row in rows:
        ws.append([row.get(col, "") if row.get(col) is not None else "" for col in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    computed_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    col_index = {name: idx + 1 for idx, name in enumerate(columns)}
    editable_cols = [
        "tag_id",
        "ground_truth_los_nlos",
        "notes",
    ]
    computed_cols = [
        "range_minus_true_m",
    ]

    for row_idx in range(2, len(rows) + 2):
        true_distance = ws.cell(row_idx, col_index["true_ground_truth_distance_m"]).coordinate
        uwb = ws.cell(row_idx, col_index["uwb_mean_distance_m"]).coordinate
        ws.cell(row_idx, col_index["range_minus_true_m"]).value = (
            f'=IF(OR({true_distance}="",ISBLANK({uwb})),"",{uwb}-{true_distance})'
        )

    for col_name in editable_cols:
        col = col_index[col_name]
        for row_idx in range(2, max(len(rows) + 2, 200)):
            ws.cell(row_idx, col).fill = editable_fill

    for col_name in computed_cols:
        col = col_index[col_name]
        for row_idx in range(2, max(len(rows) + 2, 200)):
            ws.cell(row_idx, col).fill = computed_fill

    validation = DataValidation(type="list", formula1='"LOS,NLOS,Unknown"', allow_blank=True)
    ws.add_data_validation(validation)
    gt_col_letter = openpyxl.utils.get_column_letter(col_index["ground_truth_los_nlos"])
    validation.add(f"{gt_col_letter}2:{gt_col_letter}1000")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col_name in enumerate(columns, start=1):
        width = min(max(len(col_name) + 2, 12), 34)
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    info = wb.create_sheet("Instructions")
    info.append(["Field", "How to use it"])
    info.append(["Yellow cells", "Fill these after the capture session: tag ID, LOS/NLOS ground truth, and notes."])
    info.append(["true_ground_truth_distance_m", "Per-anchor true tag-to-anchor distance saved in the Measurement Session controls."])
    info.append(["true_distance_source", "direct laser or pythagorean calculation."])
    info.append(["range_minus_true_m", "UWB mean distance minus the saved per-anchor true distance."])
    info.append(["Constellation", "Use the GUI's Constellation Changed button whenever the anchor layout changes."])
    info.column_dimensions["A"].width = 28
    info.column_dimensions["B"].width = 120
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    app = ExtendedUwbCaptureApp()
    app.mainloop()


if __name__ == "__main__":
    main()

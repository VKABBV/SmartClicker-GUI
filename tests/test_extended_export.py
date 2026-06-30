import tempfile
import sqlite3
import unittest
from pathlib import Path

from openpyxl import load_workbook

from uwb_capture.common import ParsedRecord, now_local_iso
from uwb_capture.extended_gui import (
    RESPONDER_LABEL_TABLE,
    ensure_responder_label_schema,
    export_session_per_anchor,
    build_measurement_rows,
    measurement_list_default_filename,
)
from uwb_capture.store import MeasurementStore


class ExtendedExportTests(unittest.TestCase):
    def test_measurement_rows_use_saved_per_responder_los_nlos(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "measurements.sqlite"
            store = MeasurementStore(db_path)
            session_id = store.create_session(
                {
                    "condition_los": True,
                    "condition_nlos": False,
                    "constellation_label": "bench",
                    "send_start_stop": False,
                }
            )
            self._insert_sample(store, session_id, "7", 1.25)
            self._insert_sample(store, session_id, "8", 2.50)
            ensure_responder_label_schema(store.conn)
            store.conn.executemany(
                f"""
                INSERT INTO {RESPONDER_LABEL_TABLE} (
                    session_id, anchor_id, ground_truth_los_nlos, updated_at
                )
                VALUES (?, ?, ?, ?)
                """,
                [
                    (session_id, "7", "LOS", now_local_iso()),
                    (session_id, "8", "NLOS", now_local_iso()),
                ],
            )
            store.conn.commit()
            store.close()

            rows = build_measurement_rows(db_path)

        labels = {row["anchor_id"]: row["ground_truth_los_nlos"] for row in rows}
        hints = {row["anchor_id"]: row["condition_hint"] for row in rows}
        self.assertEqual(labels, {"7": "LOS", "8": "NLOS"})
        self.assertEqual(hints, {"7": "LOS", "8": "LOS"})

    def test_per_anchor_workbooks_use_saved_per_responder_los_nlos(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = MeasurementStore(Path(tmp) / "measurements.sqlite")
            session_id = store.create_session(
                {
                    "condition_los": True,
                    "condition_nlos": False,
                    "constellation_label": "bench",
                    "send_start_stop": False,
                }
            )
            self._insert_sample(store, session_id, "7", 1.25)
            self._insert_sample(store, session_id, "8", 2.50)
            ensure_responder_label_schema(store.conn)
            store.conn.executemany(
                f"""
                INSERT INTO {RESPONDER_LABEL_TABLE} (
                    session_id, anchor_id, ground_truth_los_nlos, updated_at
                )
                VALUES (?, ?, ?, ?)
                """,
                [
                    (session_id, "7", "LOS", now_local_iso()),
                    (session_id, "8", "NLOS", now_local_iso()),
                ],
            )
            store.conn.commit()
            paths = export_session_per_anchor(store.conn, session_id, Path(tmp) / "exports")
            store.close()

            by_anchor = {self._anchor_from_filename(path): path for path in paths}
            self.assertEqual(set(by_anchor), {"7", "8"})
            self.assertIn("_LOS_", by_anchor["7"].name)
            self.assertIn("_NLOS_", by_anchor["8"].name)
            self.assertEqual(self._first_sample_los_nlos(by_anchor["7"]), "LOS")
            self.assertEqual(self._first_sample_los_nlos(by_anchor["8"]), "NLOS")

    def test_per_anchor_export_filenames_include_export_timestamp(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = MeasurementStore(Path(tmp) / "measurements.sqlite")
            session_id = store.create_session(
                {
                    "condition_los": True,
                    "condition_nlos": False,
                    "constellation_label": "bench",
                    "send_start_stop": False,
                }
            )
            self._insert_sample(store, session_id, "7", 1.25)
            paths = export_session_per_anchor(
                store.conn,
                session_id,
                Path(tmp) / "exports",
                timestamp="20260619_143015",
            )
            store.close()

        self.assertEqual(len(paths), 1)
        self.assertTrue(paths[0].name.endswith("_20260619_143015.xlsx"))

    def test_measurement_rows_split_same_anchor_by_sample_los_nlos(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "measurements.sqlite"
            store = MeasurementStore(db_path)
            session_id = store.create_session({"constellation_label": "bench"})
            self._insert_sample(store, session_id, "7", 1.00, los_nlos="LOS")
            self._insert_sample(store, session_id, "7", 2.00, los_nlos="NLOS")
            store.close()

            rows = build_measurement_rows(db_path)

        by_label = {row["ground_truth_los_nlos"]: row for row in rows}
        self.assertEqual(set(by_label), {"LOS", "NLOS"})
        self.assertEqual(by_label["LOS"]["sample_count"], 1)
        self.assertEqual(by_label["NLOS"]["sample_count"], 1)
        self.assertAlmostEqual(by_label["LOS"]["uwb_mean_distance_m"], 1.00)
        self.assertAlmostEqual(by_label["NLOS"]["uwb_mean_distance_m"], 2.00)

    def test_measurement_rows_split_same_anchor_by_sample_true_distance(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "measurements.sqlite"
            store = MeasurementStore(db_path)
            session_id = store.create_session({"constellation_label": "bench"})
            self._insert_sample(store, session_id, "7", 1.00, true_distance_m=1.50)
            self._insert_sample(store, session_id, "7", 2.00, true_distance_m=2.50)
            store.close()

            rows = build_measurement_rows(db_path)

        by_truth = {row["true_ground_truth_distance_m"]: row for row in rows}
        self.assertEqual(set(by_truth), {1.50, 2.50})
        self.assertEqual(by_truth[1.50]["sample_count"], 1)
        self.assertEqual(by_truth[2.50]["sample_count"], 1)
        self.assertAlmostEqual(by_truth[1.50]["uwb_mean_distance_m"], 1.00)
        self.assertAlmostEqual(by_truth[2.50]["uwb_mean_distance_m"], 2.00)

    def test_per_anchor_workbook_uses_saved_sample_los_nlos(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = MeasurementStore(Path(tmp) / "measurements.sqlite")
            session_id = store.create_session({"constellation_label": "bench"})
            self._insert_sample(store, session_id, "7", 1.00, los_nlos="LOS")
            self._insert_sample(store, session_id, "7", 2.00, los_nlos="NLOS")
            paths = export_session_per_anchor(
                store.conn,
                session_id,
                Path(tmp) / "exports",
                timestamp="20260619_143015",
            )
            store.close()

            wb = load_workbook(paths[0], read_only=True)
            rows = self._sheet_rows_by_header(wb["Samples"])
            wb.close()

        self.assertIn("_Mixed_20260619_143015.xlsx", paths[0].name)
        self.assertEqual([row["los_nlos"] for row in rows], ["LOS", "NLOS"])

    def test_per_anchor_workbook_uses_saved_sample_true_distance(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = MeasurementStore(Path(tmp) / "measurements.sqlite")
            session_id = store.create_session({"constellation_label": "bench"})
            self._insert_sample(store, session_id, "7", 1.00, true_distance_m=1.50)
            self._insert_sample(store, session_id, "7", 2.00, true_distance_m=2.50)
            paths = export_session_per_anchor(
                store.conn,
                session_id,
                Path(tmp) / "exports",
            )
            store.close()

            wb = load_workbook(paths[0], read_only=True)
            rows = self._sheet_rows_by_header(wb["Samples"])
            wb.close()

        self.assertEqual([row["true_ground_truth_distance_m"] for row in rows], [1.50, 2.50])

    def test_export_session_subset_to_sqlite_contains_only_selected_session(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = MeasurementStore(Path(tmp) / "measurements.sqlite")
            session_a = store.create_session({"constellation_label": "bench_a"})
            session_b = store.create_session({"constellation_label": "bench_b"})
            self._insert_sample(store, session_a, "7", 1.00, los_nlos="LOS")
            self._insert_sample(store, session_b, "8", 2.00, los_nlos="NLOS")
            output = Path(tmp) / "session_a.sqlite"

            store.export_session_subset_to_sqlite(session_a, output)
            store.close()

            con = sqlite3.connect(output)
            con.row_factory = sqlite3.Row
            sessions = con.execute("SELECT id FROM sessions ORDER BY id").fetchall()
            samples = con.execute("SELECT session_id, anchor_id, los_nlos FROM samples").fetchall()
            con.close()

        self.assertEqual([row["id"] for row in sessions], [session_a])
        self.assertEqual(len(samples), 1)
        self.assertEqual(samples[0]["session_id"], session_a)
        self.assertEqual(samples[0]["anchor_id"], "7")
        self.assertEqual(samples[0]["los_nlos"], "LOS")

    def test_measurement_rows_apply_static_range_offset(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "measurements.sqlite"
            store = MeasurementStore(db_path)
            session_id = store.create_session({"constellation_label": "bench"})
            self._insert_sample(store, session_id, "7", 1.25)
            store.close()

            rows = build_measurement_rows(db_path, range_static_offset_m=0.25)

        self.assertEqual(len(rows), 1)
        self.assertAlmostEqual(rows[0]["uwb_mean_distance_m"], 1.0)

    def test_per_anchor_workbook_applies_static_range_offset(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = MeasurementStore(Path(tmp) / "measurements.sqlite")
            session_id = store.create_session({"constellation_label": "bench"})
            self._insert_sample(store, session_id, "7", 1.25)
            paths = export_session_per_anchor(
                store.conn,
                session_id,
                Path(tmp) / "exports",
                range_static_offset_m=0.25,
            )
            store.close()

            wb = load_workbook(paths[0], read_only=True)
            ws = wb["Samples"]
            headers = [cell.value for cell in ws[1]]
            measured = ws.cell(row=2, column=headers.index("measured_uwb_distance_m") + 1).value
            wb.close()

        self.assertAlmostEqual(measured, 1.0)

    def test_base_session_export_splits_cir_real_imag_samples(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = MeasurementStore(Path(tmp) / "measurements.sqlite")
            session_id = store.create_session({"constellation_label": "bench"})
            self._insert_cir_sample(store, session_id, "7")
            path = Path(tmp) / "session.xlsx"
            store.export_session_to_excel(session_id, path)
            store.close()

            wb = load_workbook(path, read_only=True)
            rows = self._sheet_rows_by_header(wb["CIR Samples"])
            wb.close()

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["cir_real"], 1)
        self.assertEqual(rows[0]["cir_imag"], -2)
        self.assertEqual(rows[0]["cir_abs_sample_index"], 640)
        self.assertEqual(rows[0]["cir_byte_offset"], 0)
        self.assertEqual(rows[1]["cir_real"], -1)
        self.assertEqual(rows[1]["cir_imag"], 258)
        self.assertEqual(rows[1]["cir_abs_sample_index"], 641)
        self.assertEqual(rows[1]["cir_byte_offset"], 6)

    def test_per_anchor_export_deduplicates_post_burst_cir_block(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = MeasurementStore(Path(tmp) / "measurements.sqlite")
            session_id = store.create_session({"constellation_label": "bench"})
            self._insert_cir_sample(store, session_id, "7", sample_index=1)
            self._insert_cir_sample(store, session_id, "7", sample_index=2)
            paths = export_session_per_anchor(store.conn, session_id, Path(tmp) / "exports")
            store.close()

            wb = load_workbook(paths[0], read_only=True)
            rows = self._sheet_rows_by_header(wb["CIR Samples"])
            wb.close()

        self.assertEqual(len(rows), 2)
        self.assertEqual({row["cir_block_id"] for row in rows}, {1})
        self.assertEqual([row["cir_window_sample_index"] for row in rows], [0, 1])

    def test_measurement_list_default_filename_includes_export_timestamp(self) -> None:
        self.assertEqual(
            measurement_list_default_filename(
                Path("My Capture.sqlite"),
                timestamp="20260619_143015",
            ),
            "My_Capture_ground_truth_measurement_list_20260619_143015.xlsx",
        )

    def _insert_cir_sample(
        self,
        store: MeasurementStore,
        session_id: str,
        anchor_id: str,
        *,
        sample_index: int = 1,
    ) -> None:
        raw = b"".join(
            [
                (1).to_bytes(3, "little", signed=True),
                (-2).to_bytes(3, "little", signed=True),
                (-1).to_bytes(3, "little", signed=True),
                (258).to_bytes(3, "little", signed=True),
            ]
        )
        store.insert_sample(
            session_id,
            ParsedRecord(
                kind="sample",
                anchor_id=anchor_id,
                clicker_id="C1",
                sample_index=sample_index,
                distance_m=1.25,
                status="ok",
                source="test",
                raw_line=f"S,{anchor_id},{sample_index},1.25",
                event_seq=7,
                burst_id=3,
                diag_source=3,
                cir_start_index=640,
                cir_first_path_index=641,
                cir_raw=raw.hex(),
            ),
        )

    def _insert_sample(
        self,
        store: MeasurementStore,
        session_id: str,
        anchor_id: str,
        distance_m: float,
        *,
        los_nlos: str | None = None,
        true_distance_m: float | None = None,
    ) -> None:
        store.insert_sample(
            session_id,
            ParsedRecord(
                kind="sample",
                anchor_id=anchor_id,
                sample_index=1,
                distance_m=distance_m,
                status="ok",
                source="test",
                raw_line=f"S,{anchor_id},1,{distance_m}",
                los_nlos=los_nlos,
                true_distance_m=true_distance_m,
            ),
        )

    def _anchor_from_filename(self, path: Path) -> str:
        parts = path.name.split("_")
        return parts[parts.index("ANCHOR") + 1]

    def _first_sample_los_nlos(self, path: Path) -> str:
        wb = load_workbook(path, read_only=True)
        ws = wb["Samples"]
        headers = [cell.value for cell in ws[1]]
        col = headers.index("los_nlos") + 1
        value = ws.cell(row=2, column=col).value
        wb.close()
        return str(value)

    def _sheet_rows_by_header(self, ws: object) -> list[dict[str, object]]:
        rows = list(ws.iter_rows(values_only=True))  # type: ignore[attr-defined]
        headers = list(rows[0])
        return [dict(zip(headers, row)) for row in rows[1:]]


if __name__ == "__main__":
    unittest.main()

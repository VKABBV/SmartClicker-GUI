"""SQLite persistence and workbook export for capture sessions."""

from __future__ import annotations

import sqlite3
import uuid
from pathlib import Path
from typing import Any

import openpyxl

from .cir_export import CIR_EXPORT_COLUMNS, build_cir_export_rows
from .common import ParsedRecord, now_local_iso

class MeasurementStore:
    def __init__(self, db_path: Path) -> None:
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

    def _init_schema(self) -> None:
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS sessions (
                id TEXT PRIMARY KEY,
                started_at TEXT NOT NULL,
                stopped_at TEXT,
                port TEXT,
                baud INTEGER,
                tag_id TEXT,
                building_label TEXT,
                constellation_label TEXT,
                condition_los INTEGER NOT NULL DEFAULT 0,
                condition_nlos INTEGER NOT NULL DEFAULT 0,
                ground_truth_m REAL,
                outlier_threshold_m REAL,
                notes TEXT,
                send_start_stop INTEGER NOT NULL DEFAULT 0,
                start_command TEXT,
                stop_command TEXT
            );

            CREATE TABLE IF NOT EXISTS raw_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                raw_line TEXT NOT NULL,
                parsed INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(session_id) REFERENCES sessions(id)
            );

            CREATE TABLE IF NOT EXISTS samples (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                anchor_id TEXT,
                clicker_id TEXT,
                sample_index INTEGER,
                distance_m REAL,
                rx_power_dbm REAL,
                fp_power_dbm REAL,
                cir_power REAL,
                cir_raw TEXT,
                status TEXT,
                error_code TEXT,
                source TEXT,
                raw_line TEXT,
                event_seq INTEGER,
                scheduled_sample_count INTEGER,
                quality INTEGER,
                firmware_timestamp_ms INTEGER,
                phy_config_id INTEGER,
                burst_id INTEGER,
                tlv_json TEXT,
                exchange_stride_us INTEGER,
                burst_duration_ms INTEGER,
                diag_status_flags INTEGER,
                diag_bytes_captured INTEGER,
                diag_bytes_transmitted INTEGER,
                report_fragment_count INTEGER,
                uwb_clock_offset_raw INTEGER,
                uwb_carrier_integrator INTEGER,
                clicker_diag_bytes TEXT,
                cir_first_path_index INTEGER,
                cir_start_index INTEGER,
                diag_source INTEGER,
                los_nlos TEXT,
                true_distance_m REAL,
                FOREIGN KEY(session_id) REFERENCES sessions(id)
            );

            CREATE TABLE IF NOT EXISTS summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                anchor_id TEXT,
                mean_distance_m REAL,
                good_count INTEGER,
                status TEXT,
                source TEXT,
                raw_line TEXT,
                FOREIGN KEY(session_id) REFERENCES sessions(id)
            );

            CREATE TABLE IF NOT EXISTS alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                anchor_id TEXT,
                distance_m REAL,
                ground_truth_m REAL,
                absolute_error_m REAL,
                message TEXT NOT NULL,
                acknowledged INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(session_id) REFERENCES sessions(id)
            );

            CREATE INDEX IF NOT EXISTS idx_samples_session ON samples(session_id);
            CREATE INDEX IF NOT EXISTS idx_samples_anchor ON samples(session_id, anchor_id);
            CREATE INDEX IF NOT EXISTS idx_raw_lines_session ON raw_lines(session_id);
            CREATE INDEX IF NOT EXISTS idx_alerts_session ON alerts(session_id);
            """
        )
        self._migrate_samples_columns()
        self.conn.commit()

    _SAMPLES_COLUMNS: dict[str, str] = {
        "clicker_id": "TEXT",
        "event_seq": "INTEGER",
        "scheduled_sample_count": "INTEGER",
        "quality": "INTEGER",
        "firmware_timestamp_ms": "INTEGER",
        "phy_config_id": "INTEGER",
        "burst_id": "INTEGER",
        "tlv_json": "TEXT",
        "exchange_stride_us": "INTEGER",
        "burst_duration_ms": "INTEGER",
        "diag_status_flags": "INTEGER",
        "diag_bytes_captured": "INTEGER",
        "diag_bytes_transmitted": "INTEGER",
        "report_fragment_count": "INTEGER",
        "uwb_clock_offset_raw": "INTEGER",
        "uwb_carrier_integrator": "INTEGER",
        "clicker_diag_bytes": "TEXT",
        "cir_first_path_index": "INTEGER",
        "cir_start_index": "INTEGER",
        "diag_source": "INTEGER",
        "los_nlos": "TEXT",
        "true_distance_m": "REAL",
    }

    def _migrate_samples_columns(self) -> None:
        existing = {
            row["name"]
            for row in self.conn.execute("PRAGMA table_info(samples)").fetchall()
        }
        for col, col_type in self._SAMPLES_COLUMNS.items():
            if col not in existing:
                self.conn.execute(f"ALTER TABLE samples ADD COLUMN {col} {col_type}")

    def create_session(self, metadata: dict[str, Any]) -> str:
        session_id = str(uuid.uuid4())
        self.conn.execute(
            """
            INSERT INTO sessions (
                id, started_at, port, baud, tag_id, building_label, constellation_label,
                condition_los, condition_nlos, ground_truth_m, outlier_threshold_m, notes,
                send_start_stop, start_command, stop_command
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session_id,
                now_local_iso(),
                metadata.get("port"),
                metadata.get("baud"),
                metadata.get("tag_id"),
                metadata.get("building_label"),
                metadata.get("constellation_label"),
                int(bool(metadata.get("condition_los"))),
                int(bool(metadata.get("condition_nlos"))),
                metadata.get("ground_truth_m"),
                metadata.get("outlier_threshold_m"),
                metadata.get("notes"),
                int(bool(metadata.get("send_start_stop"))),
                metadata.get("start_command"),
                metadata.get("stop_command"),
            ),
        )
        self.conn.commit()
        return session_id

    def finish_session(self, session_id: str) -> None:
        self.conn.execute("UPDATE sessions SET stopped_at = ? WHERE id = ?", (now_local_iso(), session_id))
        self.conn.commit()

    def insert_raw_line(self, session_id: str, raw_line: str, parsed: bool) -> int:
        cur = self.conn.execute(
            "INSERT INTO raw_lines (session_id, timestamp, raw_line, parsed) VALUES (?, ?, ?, ?)",
            (session_id, now_local_iso(), raw_line, int(bool(parsed))),
        )
        self.conn.commit()
        return int(cur.lastrowid)

    def insert_sample(self, session_id: str, record: ParsedRecord) -> int:
        cur = self.conn.execute(
            """
            INSERT INTO samples (
                session_id, timestamp, anchor_id, clicker_id, sample_index, distance_m,
                rx_power_dbm, fp_power_dbm, cir_power, cir_raw, status,
                error_code, source, raw_line, event_seq, scheduled_sample_count,
                quality, firmware_timestamp_ms, phy_config_id, burst_id, tlv_json,
                exchange_stride_us, burst_duration_ms, diag_status_flags,
                diag_bytes_captured, diag_bytes_transmitted, report_fragment_count,
                uwb_clock_offset_raw, uwb_carrier_integrator, clicker_diag_bytes,
                cir_first_path_index, cir_start_index, diag_source, los_nlos,
                true_distance_m
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session_id,
                now_local_iso(),
                record.anchor_id,
                record.clicker_id,
                record.sample_index,
                record.distance_m,
                record.rx_power_dbm,
                record.fp_power_dbm,
                record.cir_power,
                record.cir_raw,
                record.status,
                record.error_code,
                record.source,
                record.raw_line,
                record.event_seq,
                record.scheduled_sample_count,
                record.quality,
                record.firmware_timestamp_ms,
                record.phy_config_id,
                record.burst_id,
                record.tlv_json,
                record.exchange_stride_us,
                record.burst_duration_ms,
                record.diag_status_flags,
                record.diag_bytes_captured,
                record.diag_bytes_transmitted,
                record.report_fragment_count,
                record.uwb_clock_offset_raw,
                record.uwb_carrier_integrator,
                record.clicker_diag_bytes,
                record.cir_first_path_index,
                record.cir_start_index,
                record.diag_source,
                record.los_nlos,
                record.true_distance_m,
            ),
        )
        self.conn.commit()
        return int(cur.lastrowid)

    def merge_diagnostic_into_sample(self, sample_row_id: int, record: ParsedRecord) -> None:
        updates: list[str] = []
        params: list[Any] = []
        for field in (
            "exchange_stride_us", "burst_duration_ms", "diag_status_flags",
            "diag_bytes_captured", "diag_bytes_transmitted", "report_fragment_count",
            "uwb_clock_offset_raw", "uwb_carrier_integrator", "clicker_diag_bytes",
            "cir_raw", "rx_power_dbm", "phy_config_id", "burst_id", "tlv_json",
            "cir_first_path_index", "cir_start_index", "diag_source",
        ):
            value = getattr(record, field, None)
            if value is not None:
                updates.append(f"{field} = ?")
                params.append(value)
        if not updates:
            return
        params.append(sample_row_id)
        self.conn.execute(
            f"UPDATE samples SET {', '.join(updates)} WHERE id = ?",
            params,
        )
        self.conn.commit()

    def update_sample_cir_full(self, sample_row_id: int, cir_full_hex: str) -> None:
        self.conn.execute(
            "UPDATE samples SET cir_raw = ? WHERE id = ?",
            (cir_full_hex, sample_row_id),
        )
        self.conn.commit()

    def insert_summary(self, session_id: str, record: ParsedRecord) -> int:
        cur = self.conn.execute(
            """
            INSERT INTO summaries (
                session_id, timestamp, anchor_id, mean_distance_m, good_count,
                status, source, raw_line
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session_id,
                now_local_iso(),
                record.anchor_id,
                record.mean_distance_m,
                record.good_count,
                record.status,
                record.source,
                record.raw_line,
            ),
        )
        self.conn.commit()
        return int(cur.lastrowid)

    def insert_alert(
        self,
        session_id: str,
        anchor_id: str | None,
        distance_m: float | None,
        ground_truth_m: float | None,
        absolute_error_m: float | None,
        message: str,
    ) -> int:
        cur = self.conn.execute(
            """
            INSERT INTO alerts (
                session_id, timestamp, anchor_id, distance_m, ground_truth_m, absolute_error_m, message
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (session_id, now_local_iso(), anchor_id, distance_m, ground_truth_m, absolute_error_m, message),
        )
        self.conn.commit()
        return int(cur.lastrowid)

    def counts(self, session_id: str) -> dict[str, int]:
        result: dict[str, int] = {}
        for table in ("samples", "summaries", "alerts", "raw_lines"):
            row = self.conn.execute(
                f"SELECT COUNT(*) AS count FROM {table} WHERE session_id = ?",
                (session_id,),
            ).fetchone()
            result[table] = int(row["count"]) if row else 0
        return result

    def session_row(self, session_id: str) -> sqlite3.Row:
        row = self.conn.execute("SELECT * FROM sessions WHERE id = ?", (session_id,)).fetchone()
        if row is None:
            raise ValueError(f"Unknown session id: {session_id}")
        return row

    def list_sessions(self) -> list[sqlite3.Row]:
        return self.conn.execute(
            """
            SELECT
                s.*,
                (
                    SELECT COUNT(*)
                    FROM samples sample
                    WHERE sample.session_id = s.id
                ) AS sample_count,
                (
                    SELECT COUNT(DISTINCT sample.anchor_id)
                    FROM samples sample
                    WHERE sample.session_id = s.id AND sample.anchor_id IS NOT NULL
                ) AS anchor_count
            FROM sessions s
            ORDER BY s.started_at DESC, s.id DESC
            """
        ).fetchall()

    def delete_session(self, session_id: str) -> None:
        self.session_row(session_id)
        for table in (
            "anchor_true_distances",
            "responder_los_nlos_labels",
            "raw_lines",
            "alerts",
            "summaries",
            "samples",
        ):
            if self._table_exists(table):
                self.conn.execute(f"DELETE FROM {table} WHERE session_id = ?", (session_id,))
        self.conn.execute("DELETE FROM sessions WHERE id = ?", (session_id,))
        self.conn.commit()

    def export_session_subset_to_sqlite(self, session_id: str, output_path: Path) -> None:
        self.session_row(session_id)
        output_path = Path(output_path)
        if output_path.expanduser().resolve() == self.db_path.expanduser().resolve():
            raise ValueError("Choose a different file for the session SQLite export.")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        if output_path.exists():
            output_path.unlink()

        tables = [
            table
            for table in (
                "sessions",
                "raw_lines",
                "samples",
                "summaries",
                "alerts",
                "anchor_true_distances",
                "responder_los_nlos_labels",
            )
            if self._table_exists(table)
        ]

        dest = sqlite3.connect(output_path)
        dest.row_factory = sqlite3.Row
        try:
            for table in tables:
                row = self.conn.execute(
                    "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = ?",
                    (table,),
                ).fetchone()
                if row is not None and row["sql"]:
                    dest.execute(row["sql"])

            for table in tables:
                if table == "sessions":
                    rows = self.conn.execute("SELECT * FROM sessions WHERE id = ?", (session_id,)).fetchall()
                else:
                    rows = self.conn.execute(f"SELECT * FROM {table} WHERE session_id = ?", (session_id,)).fetchall()
                self._copy_rows(dest, table, rows)
            dest.commit()
        finally:
            dest.close()

    def _table_exists(self, table_name: str) -> bool:
        row = self.conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
            (table_name,),
        ).fetchone()
        return row is not None

    @staticmethod
    def _copy_rows(dest: sqlite3.Connection, table: str, rows: list[sqlite3.Row]) -> None:
        if not rows:
            return
        columns = list(rows[0].keys())
        placeholders = ", ".join("?" for _ in columns)
        column_sql = ", ".join(columns)
        dest.executemany(
            f"INSERT INTO {table} ({column_sql}) VALUES ({placeholders})",
            ([row[column] for column in columns] for row in rows),
        )

    def export_session_to_excel(self, session_id: str, output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        session_sheet = wb.active
        session_sheet.title = "Session"
        session = self.session_row(session_id)
        session_sheet.append(["field", "value"])
        for key in session.keys():
            session_sheet.append([key, session[key]])
        session_sheet.freeze_panes = "A2"

        sample_rows = self.conn.execute(
            "SELECT * FROM samples WHERE session_id = ? ORDER BY id",
            (session_id,),
        ).fetchall()
        self._write_rows_sheet(wb, "Samples", sample_rows)
        self._write_cir_samples_sheet(wb, sample_rows)
        self._write_table_sheet(wb, "Summaries", "SELECT * FROM summaries WHERE session_id = ? ORDER BY id", session_id)
        self._write_table_sheet(wb, "Alerts", "SELECT * FROM alerts WHERE session_id = ? ORDER BY id", session_id)
        self._write_table_sheet(
            wb,
            "Raw Lines",
            "SELECT id, timestamp, parsed, raw_line FROM raw_lines WHERE session_id = ? ORDER BY id",
            session_id,
        )

        for sheet in wb.worksheets:
            for column_cells in sheet.columns:
                header = column_cells[0].value
                if header is None:
                    continue
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 42)
        wb.save(output_path)

    def _write_table_sheet(self, wb: openpyxl.Workbook, title: str, sql: str, session_id: str) -> None:
        sheet = wb.create_sheet(title)
        rows = self.conn.execute(sql, (session_id,)).fetchall()
        self._write_rows_to_sheet(sheet, rows)

    def _write_rows_sheet(self, wb: openpyxl.Workbook, title: str, rows: list[sqlite3.Row]) -> None:
        sheet = wb.create_sheet(title)
        self._write_rows_to_sheet(sheet, rows)

    def _write_rows_to_sheet(self, sheet: openpyxl.worksheet.worksheet.Worksheet, rows: list[sqlite3.Row]) -> None:
        if not rows:
            sheet.append(["No rows"])
            return
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row[key] for key in headers])
        sheet.freeze_panes = "A2"

    def _write_cir_samples_sheet(self, wb: openpyxl.Workbook, sample_rows: list[sqlite3.Row]) -> None:
        sheet = wb.create_sheet("CIR Samples")
        sheet.append(CIR_EXPORT_COLUMNS)
        for row in build_cir_export_rows(sample_rows):
            sheet.append(row)
        sheet.freeze_panes = "A2"

    def close(self) -> None:
        self.conn.close()

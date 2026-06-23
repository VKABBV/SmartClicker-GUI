#!/usr/bin/env python3
"""Extended UWB capture GUI with measurement workflow controls."""

from __future__ import annotations

from dataclasses import dataclass
import math
import re
import sqlite3
import statistics
import time
from datetime import datetime
from pathlib import Path
from typing import Any
from tkinter import simpledialog

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from . import base_gui as original
from .anchor_geometry import (
    ANCHOR_LAYOUT_ALGORITHM,
    AnchorLayoutResult,
    AnchorPairDistance,
    mirror_layout,
    pair_residuals,
    rotate_layout,
    rotate_layout_to_level,
    solve_anchor_layout,
)
from .common import export_timestamp
from .localization import (
    LOCALIZATION_ALGORITHM,
    LocalizationReading,
    LocalizationResult,
    build_square_simulation,
    solve_position,
)
from .protocol import (
    CommandStatus,
    ProtocolError,
    build_ml_start_anchor_pair_survey_packet,
    command_status_from_packet,
    next_sequence,
    parse_device_id,
)

tk = original.tk
ttk = original.ttk
filedialog = original.filedialog
messagebox = original.messagebox

SOURCE_DIRECT = "direct laser"
SOURCE_PYTHAGOREAN = "pythagorean calculation"
SOURCE_LEGACY = "legacy session laser"
ANCHOR_TRUTH_TABLE = "anchor_true_distances"
RESPONDER_LABEL_TABLE = "responder_los_nlos_labels"
UNKNOWN_LOS_NLOS = "Unknown"
LOS_NLOS_OPTIONS = (UNKNOWN_LOS_NLOS, "LOS", "NLOS")
ILLEGAL_EXCEL_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
LOCALIZATION_VIEW_PAN_FRACTION = 0.15
LOCALIZATION_VIEW_ZOOM_FACTOR = 1.25


@dataclass(frozen=True)
class AnchorPairFailure:
    anchor_a_id: str
    anchor_b_id: str
    status: str
    source: str


def safe_float(value: Any) -> float | None:
    try:
        if value in (None, ""):
            return None
        out = float(value)
        if math.isnan(out) or math.isinf(out):
            return None
        return out
    except (TypeError, ValueError):
        return None


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned.strip("._") or "measurement_list"


def format_meter(value: Any, unknown: str = "unknown") -> str:
    number = safe_float(value)
    if number is None:
        return unknown
    return f"{number:.3f}".rstrip("0").rstrip(".")


def apply_range_offset(value: Any, offset_m: float) -> float | None:
    number = safe_float(value)
    if number is None:
        return None
    return max(number - offset_m, 0.0)


def excel_cell(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_EXCEL_CHARS_RE.sub("", value)
    return value


def excel_row(values: list[Any]) -> list[Any]:
    return [excel_cell(value) for value in values]


def condition_label_from_session(session: sqlite3.Row) -> str:
    label = condition_text(session)
    return label.replace("/", "_") if label else "UNLABELED"


def walk_widgets(widget):
    for child in widget.winfo_children():
        yield child
        yield from walk_widgets(child)


def find_labelframe(root, text: str):
    for widget in walk_widgets(root):
        try:
            if widget.cget("text") == text:
                return widget
        except tk.TclError:
            continue
    return None


def next_constellation_label(current: str) -> str:
    match = re.search(r"^(.*?)(\d+)$", current.strip())
    if not match:
        return f"{current.strip() or 'constellation'}_2"
    prefix, number = match.groups()
    return f"{prefix}{int(number) + 1}"


def measurement_list_default_filename(db_path: Path, timestamp: str | None = None) -> str:
    stamp = timestamp or export_timestamp()
    return f"{safe_filename(db_path.stem)}_ground_truth_measurement_list_{stamp}.xlsx"


def normalize_los_nlos(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text == "LOS":
        return "LOS"
    if text == "NLOS":
        return "NLOS"
    return UNKNOWN_LOS_NLOS


def ensure_responder_label_schema(con: sqlite3.Connection) -> None:
    con.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {RESPONDER_LABEL_TABLE} (
            session_id TEXT NOT NULL,
            anchor_id TEXT NOT NULL,
            ground_truth_los_nlos TEXT NOT NULL DEFAULT 'Unknown',
            updated_at TEXT NOT NULL,
            PRIMARY KEY (session_id, anchor_id),
            FOREIGN KEY(session_id) REFERENCES sessions(id)
        )
        """
    )
    con.commit()


class ExtendedUwbCaptureApp(original.UwbCaptureApp):
    """Original GUI plus targeted measurement workflow extensions."""

    def __init__(self) -> None:
        self.anchor_true_distances: dict[str, dict[str, Any]] = {}
        self.anchor_los_nlos: dict[str, str] = {}
        self.anchor_measured_distances: dict[str, float] = {}
        self.anchor_distance_history: dict[str, list[float]] = {}
        self.detected_anchor_ids: set[str] = set()
        self.localization_rows: dict[str, dict[str, Any]] = {}
        self.localization_row_order: list[str] = []
        self.localization_anchor_positions: dict[str, tuple[str, str]] = {}
        self.localization_result: LocalizationResult | None = None
        self.localization_fullscreen_window: Any | None = None
        self.localization_fullscreen_canvas: Any | None = None
        self.anchor_pair_distances: dict[tuple[str, str], AnchorPairDistance] = {}
        self.anchor_pair_status: dict[tuple[str, str], str] = {}
        self.anchor_pair_failures: dict[tuple[str, str], AnchorPairFailure] = {}
        self.anchor_layout_result: AnchorLayoutResult | None = None
        self.anchor_layout_positions: dict[str, tuple[float, float]] = {}
        self.anchor_layout_drag_start_angle: float | None = None
        self.anchor_layout_drag_positions: dict[str, tuple[float, float]] | None = None
        self.anchor_survey_in_flight = False
        self.anchor_survey_pending_session: int | None = None
        self.anchor_survey_pending_seq: int | None = None
        self.anchor_survey_expected_pair_count: int | None = None
        self.anchor_survey_received_pair_count = 0
        self.anchor_survey_successful_pair_count = 0
        self.live_tracking_active = False
        self.live_tracking_after_id: str | None = None
        self.live_tracking_event_seq: int | None = None
        self.live_tracking_expected_sample_count: int | None = None
        self.live_tracking_received_sample_count = 0
        self.live_tracking_ranges_by_anchor: dict[str, list[float]] = {}
        self.live_tracking_finish_pending = False
        self.live_tracking_finish_after_id: str | None = None
        self._pyth_updating = False
        super().__init__()
        self._install_extensions()

    def _build_variables(self) -> None:
        super()._build_variables()
        self.range_static_offset_var = tk.StringVar(value="0")
        self.localization_view_center_x_var = tk.StringVar()
        self.localization_view_center_y_var = tk.StringVar()
        self.localization_view_scale_var = tk.StringVar()

    def _build_layout(self) -> None:
        root = ttk.Frame(self, padding=10)
        root.pack(fill=tk.BOTH, expand=True)
        root.columnconfigure(0, weight=0)
        root.columnconfigure(1, weight=1)
        root.rowconfigure(0, weight=1)

        left = self._build_scrollable_frame(root, width=480)
        self._scrollable_container.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        right_notebook = ttk.Notebook(root)
        right_notebook.grid(row=0, column=1, sticky="nsew")
        self.capture_tab = ttk.Frame(right_notebook)
        self.capture_tab.columnconfigure(0, weight=1)
        self.capture_tab.rowconfigure(0, weight=1)
        self.capture_tab.rowconfigure(1, weight=1)
        self.localization_tab = ttk.Frame(right_notebook, padding=8)
        self.localization_tab.columnconfigure(0, weight=1)
        self.localization_tab.rowconfigure(1, weight=1)
        self.localization_tab.rowconfigure(2, weight=1)
        self.anchor_geometry_tab = ttk.Frame(right_notebook, padding=8)
        self.anchor_geometry_tab.columnconfigure(0, weight=1)
        self.anchor_geometry_tab.rowconfigure(1, weight=1)
        right_notebook.add(self.capture_tab, text="Capture Data")
        right_notebook.add(self.localization_tab, text="Localization")
        right_notebook.add(self.anchor_geometry_tab, text="Anchor Geometry")
        self.right_notebook = right_notebook

        self._build_connection_panel(left)
        self._build_session_panel(left)
        self._build_storage_panel(left)
        self._build_stats_panel(left)
        self._build_live_panel(self.capture_tab)
        self._build_log_panel(self.capture_tab)
        self._build_localization_panel(self.localization_tab)
        self._build_anchor_geometry_panel(self.anchor_geometry_tab)

    def _build_session_panel(self, parent: Any) -> None:
        frame = ttk.LabelFrame(parent, text="Measurement Session", padding=8)
        frame.pack(fill=tk.X, pady=(0, 8))
        frame.columnconfigure(1, weight=1)

        rows = [
            ("Static range offset m", self.range_static_offset_var),
            ("LOS alert threshold m", self.threshold_var),
            ("Stability alert std m", self.instability_threshold_var),
        ]
        for row_index, (label, variable) in enumerate(rows):
            ttk.Label(frame, text=label).grid(row=row_index, column=0, sticky="w", pady=2)
            ttk.Entry(frame, textvariable=variable).grid(
                row=row_index,
                column=1,
                columnspan=2,
                sticky="ew",
                pady=2,
            )

        self.constellation_status_var = tk.StringVar(
            value=f"Constellation: {self.constellation_var.get().strip() or 'not set'}"
        )

        self._build_serial_control(frame, 3)

        action_row = ttk.Frame(frame)
        action_row.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(10, 0))
        action_row.columnconfigure(0, weight=1)
        action_row.columnconfigure(1, weight=1)
        self.start_button = ttk.Button(
            action_row,
            text="Start Recording",
            style="Accent.TButton",
            command=self.start_capture,
        )
        self.start_button.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        self.stop_button = ttk.Button(action_row, text="Stop Recording", command=self.stop_capture)
        self.stop_button.grid(row=0, column=1, sticky="ew", padx=(5, 0))
        ttk.Label(
            action_row,
            text="Recording saves incoming ML samples to the database. Start it once, then trigger captures below.",
            wraplength=320,
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(4, 0))

        ttk.Label(frame, textvariable=self.constellation_status_var).grid(
            row=5,
            column=0,
            columnspan=3,
            sticky="w",
            pady=(8, 0),
        )
        ttk.Label(frame, textvariable=self.session_status_var, style="Status.TLabel").grid(
            row=6,
            column=0,
            columnspan=3,
            sticky="w",
            pady=(4, 0),
        )

    def _install_extensions(self) -> None:
        self._add_constellation_button()
        self._add_solve_position_button()
        self._add_anchor_true_distance_controls()
        self._add_measurement_list_button()

    def _add_constellation_button(self) -> None:
        action_row = getattr(self, "start_button", None)
        if action_row is None:
            return

        parent = self.start_button.master
        parent.columnconfigure(2, weight=1)
        self.constellation_changed_button = ttk.Button(
            parent,
            text="Constellation",
            command=self.open_constellation_dialog,
        )
        self.constellation_changed_button.grid(
            row=0,
            column=2,
            sticky="ew",
            padx=(5, 0),
        )

    def _add_solve_position_button(self) -> None:
        if getattr(self, "start_button", None) is None:
            return

        parent = self.start_button.master
        parent.columnconfigure(3, weight=1)
        self.solve_position_button = ttk.Button(
            parent,
            text="Solve for Position",
            command=self.solve_latest_position,
        )
        self.solve_position_button.grid(
            row=0,
            column=3,
            sticky="ew",
            padx=(5, 0),
        )

    def _range_static_offset_m(self, *, strict: bool = False) -> float:
        value = safe_float(self.range_static_offset_var.get())
        if value is None:
            if strict:
                raise ValueError("Static range offset must be numeric.")
            return 0.0
        return value

    def _apply_range_offset(self, value: Any, *, strict: bool = False) -> float | None:
        return apply_range_offset(value, self._range_static_offset_m(strict=strict))

    def _record_range_m(self, record: Any, *, strict: bool = False) -> float | None:
        value = record.distance_m if record.distance_m is not None else record.mean_distance_m
        return self._apply_range_offset(value, strict=strict)

    def _build_localization_panel(self, parent: Any) -> None:
        header = ttk.LabelFrame(parent, text="Position Solver", padding=8)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        for column in range(4):
            header.columnconfigure(column, weight=1)

        self.localization_status_var = tk.StringVar(
            value="Capture anchor ranges, enter anchor X/Y positions, then solve."
        )
        self.sim_width_var = tk.StringVar(value="7")
        self.sim_height_var = tk.StringVar(value="7")
        self.live_tracking_interval_var = tk.StringVar(value="2")
        self.live_tracking_status_var = tk.StringVar(value="Live tracking stopped")
        ttk.Button(
            header,
            text="Use Latest Capture",
            command=self.populate_localization_from_latest_capture,
        ).grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ttk.Button(
            header,
            text="Run Square Simulation",
            command=self.run_square_simulation,
        ).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(
            header,
            text="Solve",
            style="Accent.TButton",
            command=self.solve_localization_from_form,
        ).grid(row=0, column=2, sticky="ew", padx=5)
        ttk.Button(
            header,
            text="Clear",
            command=self.clear_localization_inputs,
        ).grid(row=0, column=3, sticky="ew", padx=(5, 0))
        ttk.Label(header, textvariable=self.localization_status_var, style="Status.TLabel").grid(
            row=1,
            column=0,
            columnspan=4,
            sticky="w",
            pady=(8, 0),
        )
        sim_frame = ttk.Frame(header)
        sim_frame.grid(row=2, column=0, columnspan=4, sticky="ew", pady=(8, 0))
        for column in (1, 3):
            sim_frame.columnconfigure(column, weight=1)
        sim_rows = [
            ("Width m", self.sim_width_var),
            ("Height m", self.sim_height_var),
        ]
        for index, (label, variable) in enumerate(sim_rows):
            ttk.Label(sim_frame, text=label).grid(row=0, column=index * 2, sticky="w", padx=(0, 4))
            ttk.Entry(sim_frame, textvariable=variable, width=8).grid(
                row=0,
                column=index * 2 + 1,
                sticky="ew",
                padx=(0, 8),
            )

        live_frame = ttk.Frame(header)
        live_frame.grid(row=3, column=0, columnspan=4, sticky="ew", pady=(8, 0))
        live_frame.columnconfigure(1, weight=1)
        live_frame.columnconfigure(2, weight=1)
        live_frame.columnconfigure(3, weight=1)
        ttk.Label(live_frame, text="Live every s").grid(row=0, column=0, sticky="w", padx=(0, 4))
        ttk.Entry(live_frame, textvariable=self.live_tracking_interval_var, width=8).grid(
            row=0,
            column=1,
            sticky="ew",
            padx=(0, 8),
        )
        self.live_tracking_start_button = ttk.Button(
            live_frame,
            text="Start Live Tracking",
            style="Accent.TButton",
            command=self.start_live_tracking,
        )
        self.live_tracking_start_button.grid(row=0, column=2, sticky="ew", padx=(0, 5))
        self.live_tracking_stop_button = ttk.Button(
            live_frame,
            text="Stop Live Tracking",
            command=self.stop_live_tracking,
        )
        self.live_tracking_stop_button.grid(row=0, column=3, sticky="ew", padx=(5, 0))
        ttk.Label(live_frame, textvariable=self.live_tracking_status_var, style="Status.TLabel").grid(
            row=1,
            column=0,
            columnspan=4,
            sticky="w",
            pady=(6, 0),
        )
        self._set_live_tracking_button_state()

        input_frame = ttk.LabelFrame(parent, text="Anchor Coordinates and Ranges (meters)", padding=8)
        input_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
        input_frame.columnconfigure(2, weight=1)
        input_frame.columnconfigure(3, weight=1)
        input_frame.columnconfigure(4, weight=1)
        self.localization_rows_frame = input_frame

        headings = ("Use", "Anchor", "X m", "Y m", "Range m", "Offset m", "Sigma m")
        for column, heading in enumerate(headings):
            ttk.Label(input_frame, text=heading, style="Status.TLabel").grid(
                row=0,
                column=column,
                sticky="w",
                padx=3,
                pady=(0, 4),
            )

        output_frame = ttk.Frame(parent)
        output_frame.grid(row=2, column=0, sticky="nsew")
        output_frame.columnconfigure(0, weight=1)
        output_frame.columnconfigure(1, weight=1)
        output_frame.rowconfigure(0, weight=1)

        result_frame = ttk.LabelFrame(output_frame, text="Result", padding=8)
        result_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)
        self.localization_result_text = tk.Text(result_frame, height=10, wrap=tk.WORD)
        self.localization_result_text.grid(row=0, column=0, sticky="nsew")
        self.localization_result_text.configure(state=tk.DISABLED)

        plot_frame = ttk.LabelFrame(output_frame, text="Layout Preview", padding=8)
        plot_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        plot_frame.columnconfigure(0, weight=1)
        plot_frame.rowconfigure(1, weight=1)
        plot_toolbar = ttk.Frame(plot_frame)
        plot_toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        for column in range(10):
            plot_toolbar.columnconfigure(column, weight=1)
        ttk.Button(
            plot_toolbar,
            text="-X",
            command=lambda: self.pan_localization_view(-1.0, 0.0),
        ).grid(row=0, column=0, sticky="ew", padx=(0, 4))
        ttk.Button(
            plot_toolbar,
            text="+X",
            command=lambda: self.pan_localization_view(1.0, 0.0),
        ).grid(row=0, column=1, sticky="ew", padx=(0, 4))
        ttk.Button(
            plot_toolbar,
            text="-Y",
            command=lambda: self.pan_localization_view(0.0, -1.0),
        ).grid(row=0, column=2, sticky="ew", padx=(0, 4))
        ttk.Button(
            plot_toolbar,
            text="+Y",
            command=lambda: self.pan_localization_view(0.0, 1.0),
        ).grid(row=0, column=3, sticky="ew", padx=(0, 4))
        ttk.Button(
            plot_toolbar,
            text="Zoom In",
            command=lambda: self.zoom_localization_view(LOCALIZATION_VIEW_ZOOM_FACTOR),
        ).grid(row=0, column=4, sticky="ew", padx=(0, 4))
        ttk.Button(
            plot_toolbar,
            text="Zoom Out",
            command=lambda: self.zoom_localization_view(1.0 / LOCALIZATION_VIEW_ZOOM_FACTOR),
        ).grid(row=0, column=5, sticky="ew", padx=(0, 4))
        ttk.Button(
            plot_toolbar,
            text="Auto",
            command=self.reset_localization_view_controls,
        ).grid(row=0, column=6, sticky="ew", padx=(0, 4))
        ttk.Button(
            plot_toolbar,
            text="Use Solved Layout",
            command=self.populate_localization_from_anchor_layout,
        ).grid(row=0, column=7, columnspan=2, sticky="ew", padx=(0, 4))
        ttk.Button(
            plot_toolbar,
            text="Fullscreen",
            command=self.open_localization_plot_fullscreen,
        ).grid(row=0, column=9, sticky="e")
        self.localization_canvas = tk.Canvas(
            plot_frame,
            height=280,
            background="white",
            highlightthickness=1,
            highlightbackground="#d0d7de",
        )
        self.localization_canvas.grid(row=1, column=0, sticky="nsew")
        self.localization_canvas.bind(
            "<Configure>",
            lambda _event: self._redraw_localization_canvas(self.localization_canvas),
        )
        self._draw_empty_localization_plot()

    def populate_localization_from_anchor_layout(self) -> bool:
        positions = self._current_anchor_layout_positions()
        if not positions:
            self.localization_status_var.set("Solve anchor geometry before loading anchor positions.")
            messagebox.showwarning(
                "No solved anchor layout",
                "Solve anchor geometry before loading anchor positions.",
                parent=self,
            )
            return False

        self._remember_localization_positions()
        for anchor_id, (x_m, y_m) in sorted(positions.items()):
            row = self._ensure_localization_row(anchor_id)
            row["enabled_var"].set(True)
            row["anchor_var"].set(anchor_id)
            row["x_var"].set(format_meter(x_m, ""))
            row["y_var"].set(format_meter(y_m, ""))
            self.localization_anchor_positions[anchor_id] = (
                row["x_var"].get().strip(),
                row["y_var"].get().strip(),
            )

        notebook = self.__dict__.get("right_notebook")
        tab = self.__dict__.get("localization_tab")
        if notebook is not None and tab is not None:
            notebook.select(tab)
        self.localization_status_var.set(
            f"Loaded {len(positions)} anchor coordinate(s) from solved anchor layout."
        )
        self._redraw_localization_views()
        return True

    def _ensure_localization_row(self, anchor_id: str) -> dict[str, Any]:
        anchor_key = anchor_id.strip()
        if not anchor_key:
            anchor_key = f"anchor_{len(self.localization_row_order) + 1}"
        if anchor_key in self.localization_rows:
            return self.localization_rows[anchor_key]

        row_index = len(self.localization_row_order) + 1
        saved_x, saved_y = self.localization_anchor_positions.get(anchor_key, ("", ""))
        enabled_var = tk.BooleanVar(value=True)
        anchor_var = tk.StringVar(value=anchor_key)
        x_var = tk.StringVar(value=saved_x)
        y_var = tk.StringVar(value=saved_y)
        range_var = tk.StringVar(value="")
        offset_var = tk.StringVar(value="0")
        sigma_var = tk.StringVar(value="0.05")

        row = {
            "enabled_var": enabled_var,
            "anchor_var": anchor_var,
            "x_var": x_var,
            "y_var": y_var,
            "range_var": range_var,
            "offset_var": offset_var,
            "sigma_var": sigma_var,
        }
        self.localization_rows[anchor_key] = row
        self.localization_row_order.append(anchor_key)

        ttk.Checkbutton(self.localization_rows_frame, variable=enabled_var).grid(
            row=row_index,
            column=0,
            sticky="w",
            padx=3,
            pady=2,
        )
        ttk.Entry(self.localization_rows_frame, textvariable=anchor_var, width=20).grid(
            row=row_index,
            column=1,
            sticky="ew",
            padx=3,
            pady=2,
        )
        ttk.Entry(self.localization_rows_frame, textvariable=x_var, width=10).grid(
            row=row_index,
            column=2,
            sticky="ew",
            padx=3,
            pady=2,
        )
        ttk.Entry(self.localization_rows_frame, textvariable=y_var, width=10).grid(
            row=row_index,
            column=3,
            sticky="ew",
            padx=3,
            pady=2,
        )
        ttk.Entry(self.localization_rows_frame, textvariable=range_var, width=10).grid(
            row=row_index,
            column=4,
            sticky="ew",
            padx=3,
            pady=2,
        )
        ttk.Entry(self.localization_rows_frame, textvariable=offset_var, width=10).grid(
            row=row_index,
            column=5,
            sticky="ew",
            padx=3,
            pady=2,
        )
        ttk.Entry(self.localization_rows_frame, textvariable=sigma_var, width=10).grid(
            row=row_index,
            column=6,
            sticky="ew",
            padx=3,
            pady=2,
        )
        return row

    def _remember_localization_positions(self) -> None:
        for key in list(self.localization_row_order):
            row = self.localization_rows.get(key)
            if not row:
                continue
            anchor_id = row["anchor_var"].get().strip()
            if not anchor_id:
                continue
            self.localization_anchor_positions[anchor_id] = (
                row["x_var"].get().strip(),
                row["y_var"].get().strip(),
            )

    def _latest_anchor_distances(self) -> dict[str, float]:
        distances: dict[str, float] = {}
        store = getattr(self, "store", None)
        session_id = getattr(self, "current_session_id", None)
        if store is not None and session_id:
            for row in store.conn.execute(
                """
                SELECT anchor_id, distance_m
                FROM samples
                WHERE session_id = ?
                  AND anchor_id IS NOT NULL
                  AND distance_m IS NOT NULL
                ORDER BY id
                """,
                (session_id,),
            ):
                anchor_id = str(row["anchor_id"]).strip()
                if anchor_id:
                    distances[anchor_id] = float(row["distance_m"])
            for row in store.conn.execute(
                """
                SELECT anchor_id, mean_distance_m
                FROM summaries
                WHERE session_id = ?
                  AND anchor_id IS NOT NULL
                  AND mean_distance_m IS NOT NULL
                ORDER BY id
                """,
                (session_id,),
            ):
                anchor_id = str(row["anchor_id"]).strip()
                if anchor_id:
                    distances[anchor_id] = float(row["mean_distance_m"])

        for anchor_id, distance in self.anchor_measured_distances.items():
            distances[str(anchor_id).strip()] = float(distance)
        return {anchor: value for anchor, value in distances.items() if anchor}

    def _update_localization_range(self, anchor_id: str, distance_m: float) -> None:
        if not hasattr(self, "localization_rows_frame"):
            return
        row = self._ensure_localization_row(anchor_id)
        row["range_var"].set(format_meter(distance_m, ""))

    def populate_localization_from_latest_capture(self) -> bool:
        distances = self._latest_anchor_distances()
        if not distances:
            self.localization_status_var.set("No captured anchor ranges are available yet.")
            messagebox.showwarning(
                "No anchor ranges",
                "Capture anchor ranges first, then solve for position.",
                parent=self,
            )
            return False

        self._remember_localization_positions()
        for anchor_id, distance in sorted(distances.items(), key=lambda item: item[0]):
            row = self._ensure_localization_row(anchor_id)
            saved_x, saved_y = self.localization_anchor_positions.get(anchor_id, ("", ""))
            if saved_x and not row["x_var"].get().strip():
                row["x_var"].set(saved_x)
            if saved_y and not row["y_var"].get().strip():
                row["y_var"].set(saved_y)
            row["range_var"].set(format_meter(distance, ""))
            row["enabled_var"].set(True)

        self.localization_status_var.set(
            f"Loaded {len(distances)} latest anchor range(s) from capture data."
        )
        return True

    def run_square_simulation(self) -> None:
        try:
            width_m = self._positive_sim_float(self.sim_width_var, "Simulation width")
            height_m = self._positive_sim_float(self.sim_height_var, "Simulation height")
            scenario = build_square_simulation(
                width_m=width_m,
                height_m=height_m,
            )
        except ValueError as exc:
            self.localization_status_var.set(str(exc))
            self._write_localization_result(str(exc))
            messagebox.showerror("Invalid simulation", str(exc), parent=self)
            return

        for reading in scenario.readings:
            row = self._ensure_localization_row(reading.anchor_id)
            row["enabled_var"].set(True)
            row["anchor_var"].set(reading.anchor_id)
            row["x_var"].set(format_meter(reading.x_m, ""))
            row["y_var"].set(format_meter(reading.y_m, ""))
            row["range_var"].set(format_meter(reading.range_m, ""))
            row["offset_var"].set("0")
            row["sigma_var"].set(format_meter(reading.sigma_m, ""))

        self.localization_status_var.set(
            "Loaded square simulation ranges. The clicker position is solved from ranges only."
        )
        self.solve_localization_from_form()

    def _sim_float(self, variable: Any, label: str) -> float:
        value = safe_float(variable.get())
        if value is None:
            raise ValueError(f"{label} must be numeric.")
        return value

    def _positive_sim_float(self, variable: Any, label: str) -> float:
        value = self._sim_float(variable, label)
        if value <= 0:
            raise ValueError(f"{label} must be greater than 0.")
        return value

    def start_live_tracking(self) -> None:
        if self.live_tracking_active:
            return
        try:
            self._live_tracking_interval_ms()
        except ValueError as exc:
            self.live_tracking_status_var.set(str(exc))
            messagebox.showerror("Invalid live tracking interval", str(exc), parent=self)
            return
        if self.bluetooth_worker is None:
            self.live_tracking_status_var.set("Connect to the clicker before live tracking.")
            messagebox.showwarning(
                "Bluetooth not connected",
                "Connect to the clicker before starting live tracking.",
                parent=self,
            )
            return

        self.live_tracking_active = True
        self.live_tracking_status_var.set("Live tracking active; waiting for ranges.")
        self._set_live_tracking_button_state()
        self._cancel_live_tracking_timer()
        self._live_tracking_tick()

    def stop_live_tracking(self, reason: str = "Live tracking stopped.") -> None:
        if not self.live_tracking_active and self.live_tracking_after_id is None:
            return
        self.live_tracking_active = False
        self._cancel_live_tracking_timer()
        if (
            self.ml_command_in_flight
            and self.ml_pending_mode == original.ML_COLLECTION_MODE_FAST
        ):
            self._abort_ml_command("Live tracking stopped")
        self._reset_live_tracking_batch()
        if hasattr(self, "live_tracking_status_var"):
            self.live_tracking_status_var.set(reason)
        self._set_live_tracking_button_state()

    def _set_live_tracking_button_state(self) -> None:
        if hasattr(self, "live_tracking_start_button"):
            self.live_tracking_start_button.configure(
                state="disabled" if self.live_tracking_active else "normal"
            )
        if hasattr(self, "live_tracking_stop_button"):
            self.live_tracking_stop_button.configure(
                state="normal" if self.live_tracking_active else "disabled"
            )

    def _cancel_live_tracking_timer(self) -> None:
        after_id = self.live_tracking_after_id
        self.live_tracking_after_id = None
        if after_id is None:
            return
        try:
            self.after_cancel(after_id)
        except Exception:
            pass

    def _live_tracking_interval_ms(self) -> int:
        value = safe_float(self.live_tracking_interval_var.get())
        if value is None or value <= 0:
            raise ValueError("Live tracking interval must be greater than 0 seconds.")
        return max(int(value * 1000), 100)

    def _reset_live_tracking_batch(self) -> None:
        self.live_tracking_finish_pending = False
        self.live_tracking_finish_after_id = None
        self.live_tracking_event_seq = None
        self.live_tracking_expected_sample_count = None
        self.live_tracking_received_sample_count = 0
        self.live_tracking_ranges_by_anchor.clear()

    def _request_live_tracking_batch_finish(self) -> None:
        self.live_tracking_finish_pending = True
        if self.live_tracking_finish_after_id is not None:
            return
        try:
            self.live_tracking_finish_after_id = self.after_idle(
                self._finish_pending_live_tracking_batch
            )
        except Exception:
            self.live_tracking_finish_after_id = None

    def _finish_pending_live_tracking_batch(self) -> bool:
        self.live_tracking_finish_after_id = None
        if not self.live_tracking_finish_pending:
            return False
        self.live_tracking_finish_pending = False
        return self._finish_live_tracking_batch(force=True)

    def _record_live_tracking_sample(self, record: Any) -> bool:
        if record.kind not in ("sample", "failure", "summary"):
            return False
        event_seq = record.event_seq
        if self.live_tracking_event_seq is None:
            self.live_tracking_event_seq = event_seq
        if record.scheduled_sample_count is not None:
            self.live_tracking_expected_sample_count = int(record.scheduled_sample_count)
        self.live_tracking_received_sample_count += 1
        distance = record.distance_m if record.distance_m is not None else record.mean_distance_m
        if record.anchor_id and distance is not None:
            anchor_key = str(record.anchor_id).strip()
            if anchor_key:
                self.live_tracking_ranges_by_anchor.setdefault(anchor_key, []).append(float(distance))

        expected = self.live_tracking_expected_sample_count
        if expected is not None:
            self.live_tracking_status_var.set(
                f"Live tracking active; received {self.live_tracking_received_sample_count}/{expected} range rows."
            )
        return False

    def _finish_live_tracking_batch(self, *, force: bool = False) -> bool:
        expected = self.live_tracking_expected_sample_count
        if (
            not force
            and expected is not None
            and self.live_tracking_received_sample_count < expected
        ):
            return False

        ranges_by_anchor = {
            anchor_id: sum(values) / len(values)
            for anchor_id, values in self.live_tracking_ranges_by_anchor.items()
            if values
        }
        received = self.live_tracking_received_sample_count
        self._reset_live_tracking_batch()
        if len(ranges_by_anchor) < 3:
            self.live_tracking_status_var.set(
                f"Live tracking active; received {received} row(s), "
                f"{len(ranges_by_anchor)} usable anchor range(s)."
            )
            return False

        for anchor_key, distance_m in ranges_by_anchor.items():
            self.anchor_measured_distances[anchor_key] = distance_m
            history = self.anchor_distance_history.setdefault(anchor_key, [])
            history.append(distance_m)
            if len(history) > 200:
                del history[: len(history) - 200]
            self._update_localization_range(anchor_key, distance_m)
        return self._try_live_tracking_solve()

    def _schedule_live_tracking_tick(self) -> None:
        if not self.live_tracking_active:
            return
        try:
            interval_ms = self._live_tracking_interval_ms()
        except ValueError as exc:
            self.stop_live_tracking(str(exc))
            return
        self.live_tracking_after_id = self.after(interval_ms, self._live_tracking_tick)

    def _live_tracking_tick(self) -> None:
        self.live_tracking_after_id = None
        if not self.live_tracking_active:
            return
        if self.bluetooth_worker is None:
            self.stop_live_tracking("Live tracking stopped; Bluetooth is disconnected.")
            return

        if self.ml_command_in_flight:
            self.live_tracking_status_var.set("Waiting for the current range request to finish.")
        else:
            sent = self.send_ml_start_collection(collection_mode=original.ML_COLLECTION_MODE_FAST)
            if sent:
                self._reset_live_tracking_batch()
                self.live_tracking_status_var.set("Fast range request sent; waiting for anchor replies.")
            else:
                self.stop_live_tracking("Live tracking stopped; range request was not sent.")
                return
        self._schedule_live_tracking_tick()

    def _try_live_tracking_solve(self) -> bool:
        if not self.live_tracking_active:
            return False
        try:
            solved = self.solve_localization_from_form(
                show_errors=False,
                require_complete_enabled=False,
            )
        except Exception as exc:
            self.live_tracking_status_var.set(f"Live solve failed: {exc}")
            return False
        if not solved:
            self.live_tracking_status_var.set(
                "Live tracking active; enter X/Y for at least three returned anchors."
            )
        return solved

    def solve_latest_position(self) -> None:
        if not self.populate_localization_from_latest_capture():
            return
        if hasattr(self, "right_notebook") and hasattr(self, "localization_tab"):
            self.right_notebook.select(self.localization_tab)
        self.solve_localization_from_form()

    def solve_localization_from_form(
        self,
        *,
        show_errors: bool = True,
        require_complete_enabled: bool = True,
    ) -> bool:
        try:
            readings = self._read_localization_form(require_complete_enabled=require_complete_enabled)
            result = solve_position(readings)
        except ValueError as exc:
            self.localization_status_var.set(str(exc))
            self._write_localization_result(str(exc))
            if show_errors:
                messagebox.showerror("Cannot solve position", str(exc), parent=self)
            return False

        self.localization_result = result
        self._remember_localization_positions()
        self._show_localization_result(result)
        self._draw_localization_plot(result)
        if self.live_tracking_active:
            self.live_tracking_status_var.set(
                f"Live position x={result.x_m:.3f} m, y={result.y_m:.3f} m "
                f"({result.confidence} confidence)."
            )
        return True

    def _read_localization_form(
        self,
        *,
        require_complete_enabled: bool = True,
    ) -> list[LocalizationReading]:
        readings: list[LocalizationReading] = []
        missing: list[str] = []
        static_offset_m = self._range_static_offset_m(strict=True)
        for key in self.localization_row_order:
            row = self.localization_rows.get(key)
            if not row or not row["enabled_var"].get():
                continue
            anchor_id = row["anchor_var"].get().strip() or key
            x_m = safe_float(row["x_var"].get())
            y_m = safe_float(row["y_var"].get())
            range_m = safe_float(row["range_var"].get())
            offset_m = safe_float(row["offset_var"].get()) or 0.0
            total_offset_m = static_offset_m + offset_m
            sigma_m = safe_float(row["sigma_var"].get()) or 0.05
            if x_m is None or y_m is None or range_m is None:
                if require_complete_enabled:
                    missing.append(anchor_id)
                continue
            readings.append(
                LocalizationReading(
                    anchor_id=anchor_id,
                    x_m=x_m,
                    y_m=y_m,
                    range_m=range_m,
                    sigma_m=sigma_m,
                    offset_m=total_offset_m,
                )
            )

        if missing:
            raise ValueError(
                "Enter X, Y, and range values for these enabled anchors: "
                + ", ".join(missing)
            )
        if len(readings) < 3:
            raise ValueError("At least three enabled anchors with X/Y/range values are required.")
        return readings

    def _show_localization_result(self, result: LocalizationResult) -> None:
        lines = [
            f"Algorithm: {LOCALIZATION_ALGORITHM}",
            "Units: meters",
            "",
            f"Estimated position: x={result.x_m:.3f} m, y={result.y_m:.3f} m",
            f"Range fit RMSE: {result.rmse_m:.3f} m",
            f"Radical-axis RMSE: {result.radical_axis_rmse_m:.3f} m",
            f"Confidence: {result.confidence}",
        ]
        if result.common_height_m is not None:
            lines.append(f"Common height component: {result.common_height_m:.3f} m")
        lines.extend(["", "Range residuals:"])
        for anchor_id, residual in sorted(result.range_residuals_m.items()):
            lines.append(
                f"- {anchor_id}: residual={residual:.3f} m"
            )
        lines.extend(["", "Radical-axis line residuals:"])
        for pair_id, residual in sorted(result.residuals_m.items()):
            lines.append(
                f"- {pair_id}: residual={residual:.3f} m"
            )
        if result.warnings:
            lines.append("")
            lines.append("Warnings:")
            lines.extend(f"- {warning}" for warning in result.warnings)
        self._write_localization_result("\n".join(lines))
        self.localization_status_var.set(
            f"Solved position x={result.x_m:.3f} m, y={result.y_m:.3f} m "
            f"({result.confidence} confidence)."
        )

    def _write_localization_result(self, text: str) -> None:
        if not hasattr(self, "localization_result_text"):
            return
        self.localization_result_text.configure(state=tk.NORMAL)
        self.localization_result_text.delete("1.0", tk.END)
        self.localization_result_text.insert("1.0", text)
        self.localization_result_text.configure(state=tk.DISABLED)

    def clear_localization_inputs(self) -> None:
        self.localization_result = None
        for key in self.localization_row_order:
            row = self.localization_rows.get(key)
            if not row:
                continue
            row["range_var"].set("")
            row["enabled_var"].set(True)
        self.localization_status_var.set("Cleared captured ranges. Anchor positions were kept.")
        self._write_localization_result("")
        self._draw_empty_localization_plot()

    def _draw_empty_localization_plot(self) -> None:
        self._redraw_localization_views()

    def _redraw_localization_views(self) -> None:
        self._redraw_localization_canvas(getattr(self, "localization_canvas", None))
        self._redraw_localization_canvas(getattr(self, "localization_fullscreen_canvas", None))

    def _render_empty_localization_plot(self, canvas: Any) -> None:
        canvas.delete("all")
        canvas.create_text(
            18,
            18,
            text="Position plot appears after solving or running the square simulation.",
            anchor="nw",
            fill="#57606a",
        )

    def open_localization_plot_fullscreen(self) -> None:
        window = getattr(self, "localization_fullscreen_window", None)
        if window is not None:
            try:
                if window.winfo_exists():
                    window.lift()
                    window.focus_force()
                    return
            except tk.TclError:
                self.localization_fullscreen_window = None
                self.localization_fullscreen_canvas = None

        window = tk.Toplevel(self)
        window.title("Localization Plot")
        window.configure(background="white")
        self.localization_fullscreen_window = window

        toolbar = ttk.Frame(window, padding=8)
        toolbar.pack(fill=tk.X)
        toolbar.columnconfigure(0, weight=1)
        ttk.Label(toolbar, text="Localization Plot", style="Status.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Button(
            toolbar,
            text="Exit Fullscreen",
            command=self.close_localization_plot_fullscreen,
        ).grid(row=0, column=1, sticky="e")

        canvas = tk.Canvas(
            window,
            background="white",
            highlightthickness=0,
        )
        canvas.pack(fill=tk.BOTH, expand=True)
        self.localization_fullscreen_canvas = canvas
        canvas.bind("<Configure>", lambda _event: self._redraw_localization_canvas(canvas))
        window.bind("<Escape>", lambda _event: self.close_localization_plot_fullscreen())
        window.protocol("WM_DELETE_WINDOW", self.close_localization_plot_fullscreen)
        try:
            window.attributes("-fullscreen", True)
        except tk.TclError:
            try:
                window.state("zoomed")
            except tk.TclError:
                pass
        self._redraw_localization_canvas(canvas)

    def close_localization_plot_fullscreen(self) -> None:
        window = getattr(self, "localization_fullscreen_window", None)
        self.localization_fullscreen_window = None
        self.localization_fullscreen_canvas = None
        if window is None:
            return
        try:
            if window.winfo_exists():
                window.destroy()
        except tk.TclError:
            pass

    def _redraw_localization_canvas(self, canvas: Any) -> None:
        if canvas is None:
            return
        try:
            if not canvas.winfo_exists():
                return
        except tk.TclError:
            return
        if self.localization_result is None:
            self._render_empty_localization_plot(canvas)
        else:
            self._render_localization_plot(canvas, self.localization_result)

    def _draw_localization_plot(self, result: LocalizationResult) -> None:
        self._render_localization_plot(self.localization_canvas, result)
        fullscreen_canvas = getattr(self, "localization_fullscreen_canvas", None)
        if fullscreen_canvas is not None:
            self._redraw_localization_canvas(fullscreen_canvas)

    def _localization_view_number(self, attr_name: str) -> float | None:
        variable = getattr(self, attr_name, None)
        if variable is None:
            return None
        return safe_float(variable.get())

    def _set_localization_view_state(self, center_x_m: float, center_y_m: float, scale: float) -> None:
        self.localization_view_center_x_var.set(f"{center_x_m:.6g}")
        self.localization_view_center_y_var.set(f"{center_y_m:.6g}")
        self.localization_view_scale_var.set(f"{scale:.6g}")

    def _active_localization_plot_size(self) -> tuple[int, int]:
        canvas = self.__dict__.get("localization_canvas")
        if canvas is None:
            return 420, 260
        try:
            return max(canvas.winfo_width(), 420), max(canvas.winfo_height(), 260)
        except Exception:
            return 420, 260

    def _current_localization_plot_view(
        self,
    ) -> tuple[LocalizationResult, float, float, float, float, float, float, float] | None:
        result = self.__dict__.get("localization_result")
        if result is None:
            status_var = self.__dict__.get("localization_status_var")
            if status_var is not None:
                status_var.set("Solve position before changing the plot view.")
            return None
        width, height = self._active_localization_plot_size()
        center_x, center_y, scale, min_x, max_x, min_y, max_y = self._localization_plot_view(
            result,
            width=width,
            height=height,
            margin=32,
        )
        return result, center_x, center_y, scale, min_x, max_x, min_y, max_y

    def pan_localization_view(self, x_direction: float, y_direction: float) -> None:
        current = self._current_localization_plot_view()
        if current is None:
            return
        _result, center_x, center_y, scale, min_x, max_x, min_y, max_y = current
        center_x += (max_x - min_x) * LOCALIZATION_VIEW_PAN_FRACTION * x_direction
        center_y += (max_y - min_y) * LOCALIZATION_VIEW_PAN_FRACTION * y_direction
        self._set_localization_view_state(center_x, center_y, scale)
        self._redraw_localization_views()
        self.localization_status_var.set(
            f"Plot center x={center_x:.2f} m, y={center_y:.2f} m."
        )

    def zoom_localization_view(self, factor: float) -> None:
        current = self._current_localization_plot_view()
        if current is None:
            return
        _result, center_x, center_y, scale, _min_x, _max_x, _min_y, _max_y = current
        scale = max(scale * factor, 1e-9)
        self._set_localization_view_state(center_x, center_y, scale)
        self._redraw_localization_views()
        self.localization_status_var.set(f"Plot scale {scale:.2f} px/m.")

    def _localization_plot_view(
        self,
        result: LocalizationResult,
        width: int,
        height: int,
        margin: int,
    ) -> tuple[float, float, float, float, float, float, float]:
        min_x, max_x, min_y, max_y = self._localization_plot_bounds(result)
        plot_width = max(width - 2 * margin, 1)
        plot_height = max(height - 2 * margin, 1)
        data_width = max(max_x - min_x, 1e-9)
        data_height = max(max_y - min_y, 1e-9)
        auto_scale = min(plot_width / data_width, plot_height / data_height)
        auto_center_x_m = (min_x + max_x) / 2.0
        auto_center_y_m = (min_y + max_y) / 2.0

        manual_scale = self._localization_view_number("localization_view_scale_var")
        scale = manual_scale if manual_scale is not None and manual_scale > 0 else auto_scale
        center_x_m = self._localization_view_number("localization_view_center_x_var")
        center_y_m = self._localization_view_number("localization_view_center_y_var")
        if center_x_m is None:
            center_x_m = auto_center_x_m
        if center_y_m is None:
            center_y_m = auto_center_y_m

        view_width_m = plot_width / max(scale, 1e-9)
        view_height_m = plot_height / max(scale, 1e-9)
        view_min_x = center_x_m - view_width_m / 2.0
        view_max_x = center_x_m + view_width_m / 2.0
        view_min_y = center_y_m - view_height_m / 2.0
        view_max_y = center_y_m + view_height_m / 2.0
        return center_x_m, center_y_m, scale, view_min_x, view_max_x, view_min_y, view_max_y

    def reset_localization_view_controls(self) -> None:
        self.localization_view_center_x_var.set("")
        self.localization_view_center_y_var.set("")
        self.localization_view_scale_var.set("")
        self._redraw_localization_views()
        self.localization_status_var.set("Using automatic localization plot view.")

    def _localization_plot_bounds(self, result: LocalizationResult) -> tuple[float, float, float, float]:
        anchor_points = [(reading.x_m, reading.y_m) for reading in result.processed_readings]
        room_width = safe_float(self.sim_width_var.get()) if hasattr(self, "sim_width_var") else None
        room_height = safe_float(self.sim_height_var.get()) if hasattr(self, "sim_height_var") else None
        if room_width is not None and room_width > 0 and room_height is not None and room_height > 0:
            min_x = 0.0
            max_x = room_width
            min_y = 0.0
            max_y = room_height
        else:
            min_x = min(x for x, _ in anchor_points)
            max_x = max(x for x, _ in anchor_points)
            min_y = min(y for _, y in anchor_points)
            max_y = max(y for _, y in anchor_points)

        for x_m, y_m in anchor_points:
            min_x = min(min_x, x_m)
            max_x = max(max_x, x_m)
            min_y = min(min_y, y_m)
            max_y = max(max_y, y_m)
        if math.isclose(min_x, max_x):
            min_x -= 1.0
            max_x += 1.0
        if math.isclose(min_y, max_y):
            min_y -= 1.0
            max_y += 1.0
        data_width = max(max_x - min_x, 1e-9)
        data_height = max(max_y - min_y, 1e-9)
        pad = max(max(data_width, data_height) * 0.04, 0.15)
        return min_x - pad, max_x + pad, min_y - pad, max_y + pad

    def _render_localization_plot(self, canvas: Any, result: LocalizationResult) -> None:
        if not hasattr(self, "localization_canvas"):
            return
        canvas.delete("all")
        width = max(canvas.winfo_width(), 420)
        height = max(canvas.winfo_height(), 260)
        margin = 32
        (
            center_x_m,
            center_y_m,
            scale,
            view_min_x,
            view_max_x,
            view_min_y,
            view_max_y,
        ) = self._localization_plot_view(result, width, height, margin)
        center_x_px = width / 2.0
        center_y_px = height / 2.0

        def project(x_m: float, y_m: float) -> tuple[float, float]:
            x = center_x_px + (x_m - center_x_m) * scale
            y = center_y_px - (y_m - center_y_m) * scale
            return x, y

        canvas.create_rectangle(
            margin,
            margin,
            width - margin,
            height - margin,
            outline="#d0d7de",
        )
        canvas.create_text(margin, height - margin + 18, text=f"{view_min_x:.1f} m", anchor="n", fill="#57606a")
        canvas.create_text(width - margin, height - margin + 18, text=f"{view_max_x:.1f} m", anchor="n", fill="#57606a")
        canvas.create_text(margin - 8, height - margin, text=f"{view_min_y:.1f} m", anchor="e", fill="#57606a")
        canvas.create_text(margin - 8, margin, text=f"{view_max_y:.1f} m", anchor="e", fill="#57606a")
        canvas.create_text(width / 2, height - 8, text="x position (m)", anchor="s", fill="#57606a")
        canvas.create_text(8, height / 2, text="y position (m)", anchor="w", fill="#57606a", angle=90)

        tag_x, tag_y = project(result.x_m, result.y_m)
        for reading in result.processed_readings:
            anchor_x, anchor_y = project(reading.x_m, reading.y_m)
            radius_px = reading.corrected_range_m * scale
            canvas.create_oval(
                anchor_x - radius_px,
                anchor_y - radius_px,
                anchor_x + radius_px,
                anchor_y + radius_px,
                outline="#8ecae6",
                width=1,
                dash=(4, 3),
            )

        for reading in result.processed_readings:
            anchor_x, anchor_y = project(reading.x_m, reading.y_m)
            canvas.create_line(anchor_x, anchor_y, tag_x, tag_y, fill="#d0d7de")
            canvas.create_oval(
                anchor_x - 6,
                anchor_y - 6,
                anchor_x + 6,
                anchor_y + 6,
                fill="#0969da",
                outline="",
            )
            canvas.create_text(
                anchor_x + 8,
                anchor_y - 8,
                text=f"{reading.anchor_id}\nrange={reading.corrected_range_m:.2f} m",
                anchor="sw",
                fill="#24292f",
            )

        canvas.create_oval(
            tag_x - 8,
            tag_y - 8,
            tag_x + 8,
            tag_y + 8,
            fill="#cf222e",
            outline="",
        )
        canvas.create_text(
            tag_x + 10,
            tag_y,
            text=f"Position\n({result.x_m:.2f}, {result.y_m:.2f})",
            anchor="w",
            fill="#cf222e",
        )

    def _build_anchor_geometry_panel(self, parent: Any) -> None:
        header = ttk.LabelFrame(parent, text="Anchor Pair Survey", padding=8)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        for column in range(6):
            header.columnconfigure(column, weight=1)

        self.anchor_geometry_status_var = tk.StringVar(
            value="Send a survey command or add anchor-to-anchor distances manually."
        )
        self.anchor_survey_discovery_slot_count_var = tk.StringVar(value="8")
        self.anchor_layout_seed_count_var = tk.StringVar(value="24")
        self.anchor_layout_basin_hops_var = tk.StringVar(value="10")
        self.anchor_pair_a_var = tk.StringVar()
        self.anchor_pair_b_var = tk.StringVar()
        self.anchor_pair_distance_var = tk.StringVar()
        self.anchor_pair_sigma_var = tk.StringVar(value="0.05")
        self.anchor_layout_level_a_var = tk.StringVar()
        self.anchor_layout_level_b_var = tk.StringVar()
        self.anchor_layout_rotate_degrees_var = tk.StringVar(value="15")

        ttk.Label(header, text="Discovery slots").grid(row=0, column=0, sticky="w")
        ttk.Entry(header, textvariable=self.anchor_survey_discovery_slot_count_var, width=8).grid(
            row=0,
            column=1,
            sticky="ew",
            padx=(6, 10),
        )
        self.anchor_survey_button = ttk.Button(
            header,
            text="Gather Anchor Distances",
            style="Accent.TButton",
            command=self.send_anchor_pair_survey,
        )
        self.anchor_survey_button.grid(row=0, column=2, columnspan=2, sticky="ew", padx=(0, 5))
        self.anchor_survey_stop_button = ttk.Button(
            header,
            text="Reset Wait",
            command=self.stop_anchor_pair_survey_wait,
        )
        self.anchor_survey_stop_button.grid(row=0, column=4, sticky="ew", padx=5)
        ttk.Button(
            header,
            text="Solve Layout",
            command=self.solve_anchor_geometry,
        ).grid(row=0, column=5, sticky="ew", padx=(5, 0))

        ttk.Label(header, text="Seeds").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(header, textvariable=self.anchor_layout_seed_count_var, width=8).grid(
            row=1,
            column=1,
            sticky="ew",
            padx=(6, 10),
            pady=(8, 0),
        )
        ttk.Label(header, text="Basin hops").grid(row=1, column=2, sticky="w", pady=(8, 0))
        ttk.Entry(header, textvariable=self.anchor_layout_basin_hops_var, width=8).grid(
            row=1,
            column=3,
            sticky="ew",
            padx=(6, 10),
            pady=(8, 0),
        )
        ttk.Button(
            header,
            text="Use In Localization",
            command=self.populate_localization_from_anchor_layout,
        ).grid(row=1, column=4, columnspan=2, sticky="ew", pady=(8, 0))
        ttk.Label(header, textvariable=self.anchor_geometry_status_var, style="Status.TLabel").grid(
            row=2,
            column=0,
            columnspan=6,
            sticky="w",
            pady=(8, 0),
        )

        body = ttk.Frame(parent)
        body.grid(row=1, column=0, sticky="nsew")
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(1, weight=1)

        pair_frame = ttk.LabelFrame(body, text="Anchor Pair Distances", padding=8)
        pair_frame.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=(0, 5))
        pair_frame.columnconfigure(0, weight=1)
        pair_frame.rowconfigure(2, weight=1)

        manual = ttk.Frame(pair_frame)
        manual.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        for column in (1, 3, 5, 7):
            manual.columnconfigure(column, weight=1)
        ttk.Label(manual, text="Anchor A").grid(row=0, column=0, sticky="w")
        self.anchor_pair_a_combo = ttk.Combobox(
            manual,
            textvariable=self.anchor_pair_a_var,
            values=[],
            width=18,
            state="normal",
        )
        self.anchor_pair_a_combo.grid(row=0, column=1, sticky="ew", padx=(4, 8))
        ttk.Label(manual, text="Anchor B").grid(row=0, column=2, sticky="w")
        self.anchor_pair_b_combo = ttk.Combobox(
            manual,
            textvariable=self.anchor_pair_b_var,
            values=[],
            width=18,
            state="normal",
        )
        self.anchor_pair_b_combo.grid(row=0, column=3, sticky="ew", padx=(4, 8))
        ttk.Label(manual, text="Distance m").grid(row=0, column=4, sticky="w")
        ttk.Entry(manual, textvariable=self.anchor_pair_distance_var, width=10).grid(
            row=0,
            column=5,
            sticky="ew",
            padx=(4, 8),
        )
        ttk.Label(manual, text="Sigma m").grid(row=0, column=6, sticky="w")
        ttk.Entry(manual, textvariable=self.anchor_pair_sigma_var, width=8).grid(
            row=0,
            column=7,
            sticky="ew",
            padx=(4, 0),
        )

        buttons = ttk.Frame(pair_frame)
        buttons.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        for column in range(3):
            buttons.columnconfigure(column, weight=1)
        ttk.Button(
            buttons,
            text="Add / Update Pair",
            command=self.add_manual_anchor_pair_distance,
        ).grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ttk.Button(
            buttons,
            text="Delete Selected",
            command=self.delete_selected_anchor_pair_distance,
        ).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(
            buttons,
            text="Clear Pairs",
            command=self.clear_anchor_pair_distances,
        ).grid(row=0, column=2, sticky="ew", padx=(5, 0))

        columns = ("anchor_a", "anchor_b", "distance", "status", "source")
        self.anchor_pair_table = ttk.Treeview(pair_frame, columns=columns, show="headings", height=10)
        headings = {
            "anchor_a": "Anchor A",
            "anchor_b": "Anchor B",
            "distance": "Distance m",
            "status": "Status",
            "source": "Source",
        }
        widths = {
            "anchor_a": 150,
            "anchor_b": 150,
            "distance": 90,
            "status": 90,
            "source": 120,
        }
        for column in columns:
            self.anchor_pair_table.heading(column, text=headings[column])
            self.anchor_pair_table.column(column, width=widths[column], anchor="center")
        pair_scroll = ttk.Scrollbar(pair_frame, orient=tk.VERTICAL, command=self.anchor_pair_table.yview)
        self.anchor_pair_table.configure(yscrollcommand=pair_scroll.set)
        self.anchor_pair_table.grid(row=2, column=0, sticky="nsew")
        pair_scroll.grid(row=2, column=1, sticky="ns")
        self.anchor_pair_table.bind("<<TreeviewSelect>>", self.on_anchor_pair_selected)

        result_frame = ttk.LabelFrame(body, text="Spring Solver", padding=8)
        result_frame.grid(row=2, column=0, sticky="ew", padx=(0, 5), pady=(8, 0))
        result_frame.columnconfigure(0, weight=1)
        self.anchor_layout_result_text = tk.Text(result_frame, height=8, wrap=tk.WORD)
        self.anchor_layout_result_text.grid(row=0, column=0, sticky="ew")
        self.anchor_layout_result_text.configure(state=tk.DISABLED)

        plot_frame = ttk.LabelFrame(body, text="Anchor Layout", padding=8)
        plot_frame.grid(row=0, column=1, rowspan=3, sticky="nsew", padx=(5, 0))
        plot_frame.columnconfigure(0, weight=1)
        plot_frame.rowconfigure(1, weight=1)

        toolbar = ttk.Frame(plot_frame)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        for column in (1, 3, 5):
            toolbar.columnconfigure(column, weight=1)
        ttk.Label(toolbar, text="Level").grid(row=0, column=0, sticky="w")
        self.anchor_layout_level_a_combo = ttk.Combobox(
            toolbar,
            textvariable=self.anchor_layout_level_a_var,
            values=[],
            width=14,
            state="normal",
        )
        self.anchor_layout_level_a_combo.grid(row=0, column=1, sticky="ew", padx=(4, 4))
        self.anchor_layout_level_b_combo = ttk.Combobox(
            toolbar,
            textvariable=self.anchor_layout_level_b_var,
            values=[],
            width=14,
            state="normal",
        )
        self.anchor_layout_level_b_combo.grid(row=0, column=2, sticky="ew", padx=(0, 4))
        ttk.Button(
            toolbar,
            text="Straighten",
            command=self.level_anchor_layout_pair,
        ).grid(row=0, column=3, sticky="ew", padx=(0, 8))
        ttk.Button(toolbar, text="Mirror X", command=lambda: self.mirror_anchor_layout("x")).grid(
            row=0,
            column=4,
            sticky="ew",
            padx=(0, 4),
        )
        ttk.Button(toolbar, text="Mirror Y", command=lambda: self.mirror_anchor_layout("y")).grid(
            row=0,
            column=5,
            sticky="ew",
            padx=(0, 4),
        )

        rotate_toolbar = ttk.Frame(plot_frame)
        rotate_toolbar.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        for column in (1, 3):
            rotate_toolbar.columnconfigure(column, weight=1)
        ttk.Label(rotate_toolbar, text="Rotate deg").grid(row=0, column=0, sticky="w")
        ttk.Entry(
            rotate_toolbar,
            textvariable=self.anchor_layout_rotate_degrees_var,
            width=8,
        ).grid(row=0, column=1, sticky="ew", padx=(4, 8))
        ttk.Button(
            rotate_toolbar,
            text="Apply",
            command=self.rotate_anchor_layout_from_form,
        ).grid(row=0, column=2, sticky="ew", padx=(0, 4))
        ttk.Button(
            rotate_toolbar,
            text="-90",
            command=lambda: self.rotate_anchor_layout(-90.0),
        ).grid(row=0, column=3, sticky="ew", padx=(0, 4))
        ttk.Button(
            rotate_toolbar,
            text="+90",
            command=lambda: self.rotate_anchor_layout(90.0),
        ).grid(row=0, column=4, sticky="ew")

        self.anchor_geometry_canvas = tk.Canvas(
            plot_frame,
            height=360,
            background="white",
            highlightthickness=1,
            highlightbackground="#d0d7de",
        )
        self.anchor_geometry_canvas.grid(row=1, column=0, sticky="nsew")
        self.anchor_geometry_canvas.bind("<Configure>", lambda _event: self._redraw_anchor_geometry_canvas())
        self.anchor_geometry_canvas.bind("<ButtonPress-1>", self._start_anchor_layout_drag_rotate)
        self.anchor_geometry_canvas.bind("<B1-Motion>", self._drag_anchor_layout_rotate)
        self.anchor_geometry_canvas.bind("<ButtonRelease-1>", self._end_anchor_layout_drag_rotate)
        self._set_anchor_survey_state(False)
        self._redraw_anchor_geometry_canvas()

    def _anchor_pair_key(self, anchor_a: str, anchor_b: str) -> tuple[str, str]:
        return tuple(sorted((anchor_a.strip(), anchor_b.strip())))

    def _anchor_pair_anchor_ids(self) -> list[str]:
        anchor_ids = set(self.detected_anchor_ids)
        for pair in self.anchor_pair_distances.values():
            anchor_ids.add(pair.anchor_a_id)
            anchor_ids.add(pair.anchor_b_id)
        for failure in self.anchor_pair_failures.values():
            anchor_ids.add(failure.anchor_a_id)
            anchor_ids.add(failure.anchor_b_id)
        for anchor_id in self.anchor_layout_positions:
            anchor_ids.add(anchor_id)
        return sorted(anchor_id for anchor_id in anchor_ids if anchor_id)

    def _update_anchor_pair_combos(self) -> None:
        values = self._anchor_pair_anchor_ids()
        for name in (
            "anchor_pair_a_combo",
            "anchor_pair_b_combo",
            "anchor_layout_level_a_combo",
            "anchor_layout_level_b_combo",
        ):
            combo = getattr(self, name, None)
            if combo is not None:
                combo.configure(values=values)
        if len(values) >= 2:
            if not self.anchor_layout_level_a_var.get().strip():
                self.anchor_layout_level_a_var.set(values[0])
            if not self.anchor_layout_level_b_var.get().strip():
                self.anchor_layout_level_b_var.set(values[1])

    def add_manual_anchor_pair_distance(self) -> None:
        anchor_a = self.anchor_pair_a_var.get().strip()
        anchor_b = self.anchor_pair_b_var.get().strip()
        distance = safe_float(self.anchor_pair_distance_var.get())
        sigma = safe_float(self.anchor_pair_sigma_var.get()) or 0.05
        if not anchor_a or not anchor_b or anchor_a == anchor_b:
            messagebox.showerror("Invalid pair", "Enter two different anchor IDs.", parent=self)
            return
        if distance is None or distance <= 0:
            messagebox.showerror("Invalid distance", "Distance must be numeric and greater than 0.", parent=self)
            return
        if sigma <= 0:
            messagebox.showerror("Invalid sigma", "Sigma must be greater than 0.", parent=self)
            return
        self._save_anchor_pair_distance(anchor_a, anchor_b, distance, sigma, "manual", "ok")
        self.anchor_geometry_status_var.set(
            f"Saved {anchor_a}-{anchor_b} = {distance:.3f} m."
        )

    def _save_anchor_pair_distance(
        self,
        anchor_a: str,
        anchor_b: str,
        distance_m: float,
        sigma_m: float,
        source: str,
        status: str,
    ) -> None:
        key = self._anchor_pair_key(anchor_a, anchor_b)
        anchor_a_key, anchor_b_key = key
        pair = AnchorPairDistance(
            anchor_a_id=anchor_a_key,
            anchor_b_id=anchor_b_key,
            distance_m=distance_m,
            sigma_m=sigma_m,
            source=source,
        )
        self.anchor_pair_distances[key] = pair
        self.anchor_pair_status[key] = status
        self.anchor_pair_failures.pop(key, None)
        self._register_detected_anchor(anchor_a_key)
        self._register_detected_anchor(anchor_b_key)
        self._refresh_anchor_pair_table()
        self._update_anchor_pair_combos()

    def _remember_anchor_pair_failure(
        self,
        anchor_a: str,
        anchor_b: str,
        status: str,
        source: str,
    ) -> None:
        key = self._anchor_pair_key(anchor_a, anchor_b)
        anchor_a_key, anchor_b_key = key
        if key in self.anchor_pair_distances:
            return
        self.anchor_pair_failures[key] = AnchorPairFailure(
            anchor_a_id=anchor_a_key,
            anchor_b_id=anchor_b_key,
            status=status,
            source=source,
        )
        self._refresh_anchor_pair_table()

    def delete_selected_anchor_pair_distance(self) -> None:
        selected = self.anchor_pair_table.selection()
        if not selected:
            return
        for item in selected:
            values = self.anchor_pair_table.item(item, "values")
            if len(values) < 2:
                continue
            key = self._anchor_pair_key(str(values[0]), str(values[1]))
            self.anchor_pair_distances.pop(key, None)
            self.anchor_pair_status.pop(key, None)
            self.anchor_pair_failures.pop(key, None)
        self.anchor_layout_result = None
        self.anchor_layout_positions = {}
        self._refresh_anchor_pair_table()
        self._update_anchor_pair_combos()
        self._write_anchor_layout_result("")
        self._redraw_anchor_geometry_canvas()

    def clear_anchor_pair_distances(self) -> None:
        self.anchor_pair_distances.clear()
        self.anchor_pair_status.clear()
        self.anchor_pair_failures.clear()
        self.anchor_layout_result = None
        self.anchor_layout_positions = {}
        self._refresh_anchor_pair_table()
        self._update_anchor_pair_combos()
        self._write_anchor_layout_result("")
        self.anchor_geometry_status_var.set("Cleared anchor pair distances.")
        self._redraw_anchor_geometry_canvas()

    def _refresh_anchor_pair_table(self) -> None:
        if not hasattr(self, "anchor_pair_table"):
            return
        for item in self.anchor_pair_table.get_children():
            self.anchor_pair_table.delete(item)
        keys = set(self.anchor_pair_distances)
        keys.update(self.anchor_pair_failures)
        for key in sorted(keys):
            pair = self.anchor_pair_distances.get(key)
            if pair is None:
                failure = self.anchor_pair_failures[key]
                values = (
                    failure.anchor_a_id,
                    failure.anchor_b_id,
                    "",
                    failure.status,
                    failure.source,
                )
            else:
                values = (
                    pair.anchor_a_id,
                    pair.anchor_b_id,
                    format_meter(pair.distance_m, ""),
                    self.anchor_pair_status.get(key, "ok"),
                    pair.source,
                )
            self.anchor_pair_table.insert(
                "",
                tk.END,
                values=values,
            )

    def on_anchor_pair_selected(self, _event: Any = None) -> None:
        selected = self.anchor_pair_table.selection()
        if not selected:
            return
        values = self.anchor_pair_table.item(selected[0], "values")
        if len(values) < 3:
            return
        self.anchor_pair_a_var.set(str(values[0]))
        self.anchor_pair_b_var.set(str(values[1]))
        self.anchor_pair_distance_var.set(str(values[2]))
        self.anchor_layout_level_a_var.set(str(values[0]))
        self.anchor_layout_level_b_var.set(str(values[1]))

    def send_anchor_pair_survey(self) -> bool:
        if self.bluetooth_worker is None:
            self.anchor_geometry_status_var.set("Connect to the clicker before gathering anchor distances.")
            self.log_raw("# Did not send anchor pair survey; Bluetooth is not connected.")
            return False
        if self.ml_command_in_flight:
            self.anchor_geometry_status_var.set("Wait for the ML collection command to finish first.")
            self.log_raw("# Did not send anchor pair survey; ML collection is in flight.")
            return False
        if self.anchor_survey_in_flight:
            self.anchor_geometry_status_var.set("Anchor pair survey is already waiting for results.")
            return False
        try:
            discovery_slots = self._anchor_survey_discovery_slot_count()
            session_id = self._ml_session_id()
            source_id = parse_device_id(self.host_id_var.get(), default=0)
            if source_id == 0:
                raise ProtocolError("Host ID must be a stable non-zero device ID.")
            destination_id = parse_device_id(self.clicker_id_var.get(), default=0)
            self.protocol_sequence = next_sequence(self.protocol_sequence)
            packet = build_ml_start_anchor_pair_survey_packet(
                source_id=source_id,
                destination_id=destination_id,
                session_id=session_id,
                sequence=self.protocol_sequence,
                discovery_slot_count=discovery_slots,
            )
        except ProtocolError as exc:
            messagebox.showerror("Invalid anchor survey command", str(exc), parent=self)
            return False

        sent = self.bluetooth_worker.send_packet(packet)
        if not sent:
            self.anchor_geometry_status_var.set("Bluetooth is not ready for anchor pair survey.")
            self.log_raw("# Did not send anchor pair survey; Bluetooth is not ready.")
            return False

        self.anchor_survey_in_flight = True
        self.anchor_survey_pending_session = session_id
        self.anchor_survey_pending_seq = self.protocol_sequence
        self.anchor_survey_expected_pair_count = None
        self.anchor_survey_received_pair_count = 0
        self.anchor_survey_successful_pair_count = 0
        self._set_anchor_survey_state(True)
        self.anchor_geometry_status_var.set(
            "Anchor pair survey sent; waiting for pair rows and final command result."
        )
        self.log_raw(
            f"# Queued ML_START_ANCHOR_PAIR_SURVEY to {self.clicker_id_var.get().strip() or 'broadcast'} "
            f"(session={session_id} seq={self.protocol_sequence})"
        )
        return True

    def _anchor_survey_discovery_slot_count(self) -> int | None:
        text = self.anchor_survey_discovery_slot_count_var.get().strip()
        if not text:
            return None
        value = original.safe_int(text)
        if value is None or not (2 <= value <= 8):
            raise ProtocolError("Anchor pair survey discovery slots must be from 2 to 8.")
        return value

    def _set_anchor_survey_state(self, in_flight: bool) -> None:
        if hasattr(self, "anchor_survey_button"):
            self.anchor_survey_button.configure(state="disabled" if in_flight else "normal")
        if hasattr(self, "anchor_survey_stop_button"):
            self.anchor_survey_stop_button.configure(state="normal" if in_flight else "disabled")

    def stop_anchor_pair_survey_wait(self) -> None:
        if not self.anchor_survey_in_flight:
            return
        self._finish_anchor_pair_survey("Anchor pair survey wait reset.")

    def _anchor_survey_progress_text(self) -> str:
        received = self.anchor_survey_received_pair_count
        expected = self.anchor_survey_expected_pair_count
        successful = self.anchor_survey_successful_pair_count
        distance_text = f"{successful} usable distance(s)"
        if expected is None:
            return f"{received} pair row(s) received, {distance_text}"
        return f"{received}/{expected} pair row(s) received, {distance_text}"

    def _finish_anchor_pair_survey(self, reason: str | None = None) -> None:
        progress = self._anchor_survey_progress_text()
        self.anchor_survey_in_flight = False
        self.anchor_survey_pending_session = None
        self.anchor_survey_pending_seq = None
        self.anchor_survey_expected_pair_count = None
        self.anchor_survey_received_pair_count = 0
        successful_pair_count = self.anchor_survey_successful_pair_count
        self.anchor_survey_successful_pair_count = 0
        self._set_anchor_survey_state(False)
        message = reason or f"Anchor pair survey ended; {progress}."
        self.anchor_geometry_status_var.set(message)
        if successful_pair_count > 0 and self.anchor_pair_distances:
            self.solve_anchor_geometry(show_errors=False)

    def _handle_anchor_survey_command_result(self, packet: Any) -> None:
        matches = (
            self.anchor_survey_pending_session == packet.session_id
            and self.anchor_survey_pending_seq == packet.sequence
        )
        if not matches:
            return
        status = command_status_from_packet(packet)
        if status is None or status == CommandStatus.COMMAND_OK:
            self._finish_anchor_pair_survey(
                f"Anchor pair survey complete; {self._anchor_survey_progress_text()}."
            )
            return

        status_name = original._command_status_name(status)
        if status == CommandStatus.COMMAND_TIMEOUT:
            self._finish_anchor_pair_survey(
                f"Anchor pair survey timed out; {self._anchor_survey_progress_text()}."
            )
        else:
            self._finish_anchor_pair_survey(f"Anchor pair survey result: {status_name}.")

    def _handle_anchor_pair_record(self, record: Any) -> None:
        anchor_a = str(record.anchor_id or "").strip()
        anchor_b = str(getattr(record, "peer_anchor_id", "") or "").strip()
        if not anchor_a or not anchor_b:
            return
        self._register_detected_anchor(anchor_a)
        self._register_detected_anchor(anchor_b)
        if self.anchor_survey_in_flight:
            self.anchor_survey_received_pair_count += 1
            if record.scheduled_sample_count is not None:
                self.anchor_survey_expected_pair_count = int(record.scheduled_sample_count)
        source = "survey"
        if record.event_seq is not None:
            source = f"survey event {record.event_seq}"
        if record.kind == "survey_pair_failure" or record.distance_m is None:
            status = record.status or record.error_code or "unknown"
            self._remember_anchor_pair_failure(anchor_a, anchor_b, status, source)
            self.anchor_geometry_status_var.set(
                f"Anchor pair {anchor_a}-{anchor_b} failed: "
                f"{status} "
                f"({self._anchor_survey_progress_text()})."
            )
            self.log_raw(
                f"# Anchor pair survey failure {anchor_a}-{anchor_b}: "
                f"{status}"
            )
            self._update_anchor_pair_combos()
            return
        distance_m = self._apply_range_offset(record.distance_m)
        if distance_m is None:
            return
        self._save_anchor_pair_distance(
            anchor_a,
            anchor_b,
            distance_m,
            0.05,
            source,
            record.status or "ok",
        )
        if self.anchor_survey_in_flight:
            self.anchor_survey_successful_pair_count += 1
        self.anchor_geometry_status_var.set(
            f"Received pair {anchor_a}-{anchor_b}: {distance_m:.3f} m "
            f"({self._anchor_survey_progress_text()})."
        )
        self._redraw_anchor_geometry_canvas()

    def _anchor_layout_seed_count(self) -> int:
        value = original.safe_int(self.anchor_layout_seed_count_var.get())
        if value is None or value <= 0:
            raise ValueError("Seed count must be a positive integer.")
        return min(value, 128)

    def _anchor_layout_basin_hops(self) -> int:
        value = original.safe_int(self.anchor_layout_basin_hops_var.get())
        if value is None or value < 0:
            raise ValueError("Basin hops must be 0 or greater.")
        return min(value, 128)

    def solve_anchor_geometry(self, *, show_errors: bool = True) -> bool:
        if not self.anchor_pair_distances:
            self.anchor_geometry_status_var.set("No anchor-to-anchor distances are available.")
            if show_errors:
                messagebox.showwarning(
                    "No anchor pairs",
                    "Gather or add anchor-to-anchor distances before solving.",
                    parent=self,
                )
            return False
        try:
            result = solve_anchor_layout(
                list(self.anchor_pair_distances.values()),
                seed_count=self._anchor_layout_seed_count(),
                basin_hops=self._anchor_layout_basin_hops(),
            )
        except ValueError as exc:
            self.anchor_geometry_status_var.set(str(exc))
            self._write_anchor_layout_result(str(exc))
            if show_errors:
                messagebox.showerror("Cannot solve anchor geometry", str(exc), parent=self)
            return False

        self.anchor_layout_result = result
        self.anchor_layout_positions = dict(result.positions_m)
        self._update_anchor_pair_combos()
        self._show_anchor_layout_result(result)
        self._redraw_anchor_geometry_canvas()
        return True

    def _show_anchor_layout_result(self, result: AnchorLayoutResult) -> None:
        lines = [
            f"Algorithm: {ANCHOR_LAYOUT_ALGORITHM}",
            "Units: meters",
            "",
            f"Anchors: {len(result.positions_m)}",
            f"Pairs: {len(result.processed_pairs)}",
            f"Spring energy: {result.energy:.6g}",
            f"Pair RMSE: {result.rmse_m:.4f} m",
            f"Max residual: {result.max_residual_m:.4f} m",
            f"Seeds tried: {result.seed_count}",
            f"Accepted basin hops: {result.basin_hop_count}",
            "",
            "Anchor coordinates:",
        ]
        for anchor_id, (x_m, y_m) in sorted(self.anchor_layout_positions.items()):
            lines.append(f"- {anchor_id}: x={x_m:.3f}, y={y_m:.3f}")
        residuals = pair_residuals(self.anchor_layout_positions, result.processed_pairs)
        lines.extend(["", "Pair residuals:"])
        for pair_id, residual in sorted(residuals.items()):
            lines.append(f"- {pair_id}: {residual:.4f} m")
        if result.warnings:
            lines.append("")
            lines.append("Warnings:")
            lines.extend(f"- {warning}" for warning in result.warnings)
        self._write_anchor_layout_result("\n".join(lines))
        self.anchor_geometry_status_var.set(
            f"Solved {len(result.positions_m)} anchors with RMSE {result.rmse_m:.3f} m."
        )

    def _write_anchor_layout_result(self, text: str) -> None:
        if not hasattr(self, "anchor_layout_result_text"):
            return
        self.anchor_layout_result_text.configure(state=tk.NORMAL)
        self.anchor_layout_result_text.delete("1.0", tk.END)
        self.anchor_layout_result_text.insert("1.0", text)
        self.anchor_layout_result_text.configure(state=tk.DISABLED)

    def _current_anchor_layout_positions(self) -> dict[str, tuple[float, float]]:
        return dict(self.anchor_layout_positions)

    def _apply_anchor_layout_positions(
        self,
        positions: dict[str, tuple[float, float]],
        status: str,
    ) -> None:
        self.anchor_layout_positions = positions
        if self.anchor_layout_result is not None:
            self._show_anchor_layout_result(self.anchor_layout_result)
        self.anchor_geometry_status_var.set(status)
        self._redraw_anchor_geometry_canvas()

    def mirror_anchor_layout(self, axis: str) -> None:
        positions = self._current_anchor_layout_positions()
        if not positions:
            self.anchor_geometry_status_var.set("Solve anchor geometry before mirroring.")
            return
        transformed = mirror_layout(positions, axis)
        self._apply_anchor_layout_positions(transformed, f"Mirrored layout across {axis.upper()} axis.")

    def rotate_anchor_layout(self, degrees: float) -> None:
        positions = self._current_anchor_layout_positions()
        if not positions:
            self.anchor_geometry_status_var.set("Solve anchor geometry before rotating.")
            return
        transformed = rotate_layout(positions, degrees)
        self._apply_anchor_layout_positions(transformed, f"Rotated layout by {degrees:.1f} degrees.")

    def rotate_anchor_layout_from_form(self) -> None:
        degrees = safe_float(self.anchor_layout_rotate_degrees_var.get())
        if degrees is None:
            messagebox.showerror("Invalid rotation", "Rotation must be numeric degrees.", parent=self)
            return
        self.rotate_anchor_layout(degrees)

    def level_anchor_layout_pair(self) -> None:
        positions = self._current_anchor_layout_positions()
        if not positions:
            self.anchor_geometry_status_var.set("Solve anchor geometry before straightening.")
            return
        anchor_a = self.anchor_layout_level_a_var.get().strip()
        anchor_b = self.anchor_layout_level_b_var.get().strip()
        if not anchor_a or not anchor_b or anchor_a == anchor_b:
            messagebox.showerror("Invalid anchors", "Select two different anchors to straighten.", parent=self)
            return
        try:
            transformed = rotate_layout_to_level(positions, anchor_a, anchor_b)
        except ValueError as exc:
            messagebox.showerror("Cannot straighten layout", str(exc), parent=self)
            return
        self._apply_anchor_layout_positions(
            transformed,
            f"Straightened {anchor_a}-{anchor_b}; both anchors are on the same Y.",
        )

    def _redraw_anchor_geometry_canvas(self) -> None:
        canvas = getattr(self, "anchor_geometry_canvas", None)
        if canvas is None:
            return
        try:
            if not canvas.winfo_exists():
                return
        except tk.TclError:
            return
        positions = self._current_anchor_layout_positions()
        if not positions:
            self._render_empty_anchor_geometry_plot(canvas)
        else:
            self._render_anchor_geometry_plot(canvas, positions)

    def _render_empty_anchor_geometry_plot(self, canvas: Any) -> None:
        canvas.delete("all")
        canvas.create_text(
            18,
            18,
            text="Anchor geometry appears after solving pair distances.",
            anchor="nw",
            fill="#57606a",
        )

    def _render_anchor_geometry_plot(
        self,
        canvas: Any,
        positions: dict[str, tuple[float, float]],
    ) -> None:
        canvas.delete("all")
        width = max(canvas.winfo_width(), 420)
        height = max(canvas.winfo_height(), 320)
        xs = [x for x, _y in positions.values()]
        ys = [y for _x, y in positions.values()]
        min_x = min(xs)
        max_x = max(xs)
        min_y = min(ys)
        max_y = max(ys)
        if math.isclose(min_x, max_x):
            min_x -= 1.0
            max_x += 1.0
        if math.isclose(min_y, max_y):
            min_y -= 1.0
            max_y += 1.0
        data_width = max_x - min_x
        data_height = max_y - min_y
        pad_x = max(data_width * 0.12, 0.35)
        pad_y = max(data_height * 0.12, 0.35)
        min_x -= pad_x
        max_x += pad_x
        min_y -= pad_y
        max_y += pad_y

        margin = 36
        plot_width = max(width - 2 * margin, 1)
        plot_height = max(height - 2 * margin, 1)
        scale = min(plot_width / max(max_x - min_x, 1e-9), plot_height / max(max_y - min_y, 1e-9))
        center_x_m = (min_x + max_x) / 2.0
        center_y_m = (min_y + max_y) / 2.0
        center_x_px = width / 2.0
        center_y_px = height / 2.0

        def project(x_m: float, y_m: float) -> tuple[float, float]:
            return (
                center_x_px + (x_m - center_x_m) * scale,
                center_y_px - (y_m - center_y_m) * scale,
            )

        canvas.create_rectangle(
            margin,
            margin,
            width - margin,
            height - margin,
            outline="#d0d7de",
        )
        canvas.create_text(width / 2, height - 8, text="x (m)", anchor="s", fill="#57606a")
        canvas.create_text(8, height / 2, text="y (m)", anchor="w", fill="#57606a", angle=90)

        pairs = (
            self.anchor_layout_result.processed_pairs
            if self.anchor_layout_result is not None
            else tuple(self.anchor_pair_distances.values())
        )
        residuals = pair_residuals(positions, pairs)
        for pair in pairs:
            if pair.anchor_a_id not in positions or pair.anchor_b_id not in positions:
                continue
            ax, ay = project(*positions[pair.anchor_a_id])
            bx, by = project(*positions[pair.anchor_b_id])
            pair_id = f"{pair.anchor_a_id}-{pair.anchor_b_id}"
            residual = residuals.get(pair_id, 0.0)
            color = "#cf222e" if abs(residual) > 0.10 else "#8c959f"
            canvas.create_line(ax, ay, bx, by, fill=color, width=2)
            canvas.create_text(
                (ax + bx) / 2.0,
                (ay + by) / 2.0,
                text=f"{pair.distance_m:.2f} m",
                anchor="s",
                fill=color,
            )

        for anchor_id, (x_m, y_m) in sorted(positions.items()):
            x_px, y_px = project(x_m, y_m)
            canvas.create_oval(
                x_px - 7,
                y_px - 7,
                x_px + 7,
                y_px + 7,
                fill="#0969da",
                outline="",
            )
            canvas.create_text(
                x_px + 9,
                y_px - 9,
                text=f"{anchor_id}\n({x_m:.2f}, {y_m:.2f})",
                anchor="sw",
                fill="#24292f",
            )

    def _start_anchor_layout_drag_rotate(self, event: Any) -> None:
        positions = self._current_anchor_layout_positions()
        if not positions:
            return
        canvas = self.anchor_geometry_canvas
        cx = max(canvas.winfo_width(), 1) / 2.0
        cy = max(canvas.winfo_height(), 1) / 2.0
        self.anchor_layout_drag_start_angle = math.atan2(event.y - cy, event.x - cx)
        self.anchor_layout_drag_positions = positions

    def _drag_anchor_layout_rotate(self, event: Any) -> None:
        if self.anchor_layout_drag_start_angle is None or self.anchor_layout_drag_positions is None:
            return
        canvas = self.anchor_geometry_canvas
        cx = max(canvas.winfo_width(), 1) / 2.0
        cy = max(canvas.winfo_height(), 1) / 2.0
        current_angle = math.atan2(event.y - cy, event.x - cx)
        degrees = math.degrees(current_angle - self.anchor_layout_drag_start_angle)
        self.anchor_layout_positions = rotate_layout(self.anchor_layout_drag_positions, degrees)
        self.anchor_geometry_status_var.set(f"Rotating layout by {degrees:.1f} degrees.")
        self._redraw_anchor_geometry_canvas()

    def _end_anchor_layout_drag_rotate(self, _event: Any) -> None:
        if self.anchor_layout_drag_start_angle is None:
            return
        self.anchor_layout_drag_start_angle = None
        self.anchor_layout_drag_positions = None
        if self.anchor_layout_result is not None:
            self._show_anchor_layout_result(self.anchor_layout_result)

    def _add_anchor_true_distance_controls(self) -> None:
        session_frame = find_labelframe(self, "Measurement Session")
        if session_frame is None:
            return

        self.anchor_id_var = tk.StringVar()
        self.anchor_true_distance_var = tk.StringVar()
        self.anchor_los_nlos_var = tk.StringVar(value=UNKNOWN_LOS_NLOS)
        self.pyth_side_a_var = tk.StringVar()
        self.pyth_side_b_var = tk.StringVar()
        self.pyth_result_var = tk.StringVar(value="")

        rows = [
            int(child.grid_info().get("row", 0))
            for child in session_frame.grid_slaves()
            if child.grid_info()
        ]
        row_index = (max(rows) + 1) if rows else 0
        session_frame.columnconfigure(2, weight=1)

        container = ttk.LabelFrame(
            session_frame,
            text="Per-responder distance and LOS/NLOS",
            padding=8,
        )
        container.grid(
            row=row_index,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=(8, 0),
        )
        container.columnconfigure(1, weight=1)
        container.columnconfigure(3, weight=1)

        ttk.Label(container, text="Responder ID").grid(row=0, column=0, sticky="w")
        self.anchor_id_combo = ttk.Combobox(
            container,
            textvariable=self.anchor_id_var,
            values=[],
            width=16,
            state="normal",
        )
        self.anchor_id_combo.grid(row=0, column=1, sticky="ew", padx=(6, 12))

        ttk.Label(container, text="True distance m").grid(row=0, column=2, sticky="w")
        ttk.Entry(container, textvariable=self.anchor_true_distance_var, width=14).grid(
            row=0,
            column=3,
            sticky="ew",
            padx=(6, 12),
        )
        ttk.Button(
            container,
            text="Save direct distance",
            command=self.save_direct_true_distance,
        ).grid(row=0, column=4, sticky="ew")

        ttk.Label(container, text="LOS/NLOS").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.anchor_los_nlos_combo = ttk.Combobox(
            container,
            textvariable=self.anchor_los_nlos_var,
            values=LOS_NLOS_OPTIONS,
            state="readonly",
            width=16,
        )
        self.anchor_los_nlos_combo.grid(row=1, column=1, sticky="ew", padx=(6, 12), pady=(8, 0))
        ttk.Button(
            container,
            text="Save LOS/NLOS for responder",
            command=self.save_anchor_los_nlos,
        ).grid(row=1, column=2, columnspan=3, sticky="ew", pady=(8, 0))

        helper = ttk.Frame(container)
        helper.grid(row=2, column=0, columnspan=5, sticky="ew", pady=(8, 0))
        for col in (1, 3, 5):
            helper.columnconfigure(col, weight=1)

        ttk.Label(helper, text="Side A m").grid(row=0, column=0, sticky="w")
        ttk.Entry(helper, textvariable=self.pyth_side_a_var, width=10).grid(
            row=0,
            column=1,
            sticky="ew",
            padx=(6, 12),
        )
        ttk.Label(helper, text="Side B m").grid(row=0, column=2, sticky="w")
        ttk.Entry(helper, textvariable=self.pyth_side_b_var, width=10).grid(
            row=0,
            column=3,
            sticky="ew",
            padx=(6, 12),
        )
        ttk.Label(helper, text="Result m").grid(row=0, column=4, sticky="w")
        ttk.Entry(helper, textvariable=self.pyth_result_var, width=10, state="readonly").grid(
            row=0,
            column=5,
            sticky="ew",
            padx=(6, 0),
        )

        ttk.Button(helper, text="Calculate", command=self.calculate_pythagorean_distance).grid(
            row=2,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(8, 0),
        )
        ttk.Button(
            helper,
            text="Use calculated distance for selected anchor",
            command=self.use_pythagorean_distance_for_anchor,
        ).grid(row=2, column=2, columnspan=4, sticky="ew", padx=(8, 0), pady=(8, 0))

        self.anchor_truth_table = ttk.Treeview(
            container,
            columns=("anchor_id", "measured_m", "std_cm", "true_distance_m", "diff_m", "los_nlos"),
            show="headings",
            height=6,
        )
        self.anchor_truth_table.heading("anchor_id", text="Responder ID")
        self.anchor_truth_table.heading("measured_m", text="Measured m")
        self.anchor_truth_table.heading("std_cm", text="Std cm")
        self.anchor_truth_table.heading("true_distance_m", text="True distance m")
        self.anchor_truth_table.heading("diff_m", text="Diff m")
        self.anchor_truth_table.heading("los_nlos", text="LOS/NLOS")
        self.anchor_truth_table.column("anchor_id", width=100, anchor="center")
        self.anchor_truth_table.column("measured_m", width=80, anchor="center")
        self.anchor_truth_table.column("std_cm", width=70, anchor="center")
        self.anchor_truth_table.column("true_distance_m", width=100, anchor="center")
        self.anchor_truth_table.column("diff_m", width=70, anchor="center")
        self.anchor_truth_table.column("los_nlos", width=80, anchor="center")
        self.anchor_truth_table.grid(row=3, column=0, columnspan=5, sticky="ew", pady=(8, 0))
        self.anchor_truth_table.tag_configure("mismatch", background="#f8d7da")
        self.anchor_truth_table.bind("<<TreeviewSelect>>", self.on_anchor_truth_selected)

        self.pyth_side_a_var.trace_add("write", self._on_pythagorean_sides_changed)
        self.pyth_side_b_var.trace_add("write", self._on_pythagorean_sides_changed)

    def _add_measurement_list_button(self) -> None:
        storage = find_labelframe(self, "Storage")
        if storage is None:
            return

        self.export_measurement_list_button = ttk.Button(
            storage,
            text="Export Measurement List",
            command=self.export_measurement_list,
        )
        self.export_measurement_list_button.grid(
            row=3,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=(8, 0),
        )

    def _selected_anchor_id(self) -> str | None:
        anchor_id = self.anchor_id_var.get().strip()
        if not anchor_id:
            messagebox.showerror(
                "Missing responder ID",
                "Select or enter a responder ID before saving.",
                parent=self,
            )
            return None
        return anchor_id

    def _positive_float_from_var(self, variable: Any, label: str) -> float | None:
        value = safe_float(variable.get())
        if value is None or value <= 0:
            messagebox.showerror(
                "Invalid value",
                f"{label} must be numeric and greater than 0.",
                parent=self,
            )
            return None
        return value

    def _on_pythagorean_sides_changed(self, *_args: Any) -> None:
        if self._pyth_updating:
            return
        self._update_pythagorean_from_sides()

    def _update_pythagorean_from_sides(self) -> float | None:
        side_a = safe_float(self.pyth_side_a_var.get())
        side_b = safe_float(self.pyth_side_b_var.get())
        if side_a is None or side_a <= 0 or side_b is None or side_b <= 0:
            self._pyth_updating = True
            try:
                self.pyth_result_var.set("")
            finally:
                self._pyth_updating = False
            return None

        result = math.sqrt(side_a**2 + side_b**2)
        self._pyth_updating = True
        try:
            self.pyth_result_var.set(format_meter(result))
        finally:
            self._pyth_updating = False
        return result

    def save_direct_true_distance(self) -> None:
        anchor_id = self._selected_anchor_id()
        if anchor_id is None:
            return
        distance = self._positive_float_from_var(
            self.anchor_true_distance_var,
            "Direct true distance",
        )
        if distance is None:
            return
        self._save_anchor_true_distance(
            anchor_id,
            distance,
            SOURCE_DIRECT,
            None,
            None,
        )

    def save_anchor_los_nlos(self) -> None:
        anchor_id = self._selected_anchor_id()
        if anchor_id is None:
            return
        condition = normalize_los_nlos(self.anchor_los_nlos_var.get())
        self.anchor_los_nlos[anchor_id] = condition
        self._register_detected_anchor(anchor_id)
        self._refresh_anchor_truth_table()
        self._persist_anchor_los_nlos(anchor_id)
        self.log_raw(f"# LOS/NLOS saved for responder {anchor_id}: {condition}")

    def calculate_pythagorean_distance(self) -> float | None:
        side_a = self._positive_float_from_var(self.pyth_side_a_var, "Pythagorean Side A")
        side_b = self._positive_float_from_var(self.pyth_side_b_var, "Pythagorean Side B")
        if side_a is None or side_b is None:
            return None
        result = math.sqrt(side_a**2 + side_b**2)
        self._pyth_updating = True
        try:
            self.pyth_result_var.set(format_meter(result))
        finally:
            self._pyth_updating = False
        return result

    def use_pythagorean_distance_for_anchor(self) -> None:
        anchor_id = self._selected_anchor_id()
        if anchor_id is None:
            return
        side_a = self._positive_float_from_var(self.pyth_side_a_var, "Pythagorean Side A")
        side_b = self._positive_float_from_var(self.pyth_side_b_var, "Pythagorean Side B")
        if side_a is None or side_b is None:
            return
        result = math.sqrt(side_a**2 + side_b**2)
        self.pyth_result_var.set(format_meter(result))
        self.anchor_true_distance_var.set(format_meter(result))
        self._save_anchor_true_distance(
            anchor_id,
            result,
            SOURCE_PYTHAGOREAN,
            side_a,
            side_b,
        )

    def _save_anchor_true_distance(
        self,
        anchor_id: str,
        distance: float,
        source: str,
        side_a: float | None,
        side_b: float | None,
    ) -> None:
        self.anchor_true_distances[anchor_id] = {
            "anchor_id": anchor_id,
            "true_distance_m": distance,
            "distance_source": source,
            "pythagorean_side_a_m": side_a,
            "pythagorean_side_b_m": side_b,
        }
        self._register_detected_anchor(anchor_id)
        self._refresh_anchor_truth_table()
        self._persist_anchor_true_distance(anchor_id)
        self.log_raw(
            f"# True distance saved for anchor {anchor_id}: {format_meter(distance)} m ({source})"
        )

    def on_anchor_truth_selected(self, _event: Any = None) -> None:
        selected = self.anchor_truth_table.selection()
        if not selected:
            return
        values = self.anchor_truth_table.item(selected[0], "values")
        if not values:
            return
        anchor_id = str(values[0])
        self.anchor_id_var.set(anchor_id)
        self.anchor_los_nlos_var.set(self._anchor_los_nlos_for(anchor_id))
        data = self.anchor_true_distances.get(anchor_id)
        if not data:
            self.anchor_true_distance_var.set("")
            self.pyth_side_a_var.set("")
            self.pyth_side_b_var.set("")
            self.pyth_result_var.set("")
            return
        self.anchor_true_distance_var.set(format_meter(data["true_distance_m"]))
        self.pyth_side_a_var.set(
            "" if data.get("pythagorean_side_a_m") is None else format_meter(data["pythagorean_side_a_m"])
        )
        self.pyth_side_b_var.set(
            "" if data.get("pythagorean_side_b_m") is None else format_meter(data["pythagorean_side_b_m"])
        )
        if data.get("distance_source") == SOURCE_PYTHAGOREAN:
            self.pyth_result_var.set(format_meter(data["true_distance_m"]))
        else:
            self.pyth_result_var.set("")

    def _refresh_anchor_truth_table(self) -> None:
        if not hasattr(self, "anchor_truth_table"):
            return
        for item in self.anchor_truth_table.get_children():
            self.anchor_truth_table.delete(item)
        threshold = safe_float(self.threshold_var.get()) or 0.0
        anchor_ids = sorted(
            set(self.detected_anchor_ids) | set(self.anchor_true_distances) | set(self.anchor_los_nlos),
            key=str,
        )
        for anchor_id in anchor_ids:
            data = self.anchor_true_distances.get(anchor_id)
            true_distance = data["true_distance_m"] if data else None
            measured = self._apply_range_offset(self.anchor_measured_distances.get(anchor_id))
            raw_history = self.anchor_distance_history.get(anchor_id, [])
            history = [
                corrected
                for value in raw_history
                for corrected in [self._apply_range_offset(value)]
                if corrected is not None
            ]
            std_cm = (statistics.stdev(history) * 100) if len(history) > 1 else None
            diff = abs(measured - true_distance) if measured is not None and true_distance is not None else None
            tags = ()
            if diff is not None and diff > threshold and true_distance is not None:
                tags = ("mismatch",)
            self.anchor_truth_table.insert(
                "",
                tk.END,
                values=(
                    anchor_id,
                    format_meter(measured, "") if measured is not None else "",
                    format_meter(std_cm, "") if std_cm is not None else "",
                    format_meter(true_distance, "") if true_distance is not None else "",
                    format_meter(diff, "") if diff is not None else "",
                    self._anchor_los_nlos_for(anchor_id),
                ),
                tags=tags,
            )

    def _register_detected_anchor(self, anchor_id: Any) -> None:
        if anchor_id in (None, ""):
            return
        anchor_text = str(anchor_id).strip()
        if not anchor_text:
            return
        was_new = anchor_text not in self.detected_anchor_ids
        self.detected_anchor_ids.add(anchor_text)
        if hasattr(self, "anchor_id_combo"):
            self.anchor_id_combo.configure(values=sorted(self.detected_anchor_ids, key=str))
        if was_new:
            if hasattr(self, "localization_rows_frame"):
                self._ensure_localization_row(anchor_text)
            self._refresh_anchor_truth_table()

    def mark_constellation_changed(self) -> None:
        self.open_constellation_dialog()

    def open_constellation_dialog(self) -> None:
        current = self.constellation_var.get().strip()
        window = tk.Toplevel(self)
        window.title("Constellation")
        window.transient(self)
        window.grab_set()
        window.columnconfigure(0, weight=1)

        notebook = ttk.Notebook(window)
        notebook.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        tab = ttk.Frame(notebook, padding=10)
        tab.columnconfigure(1, weight=1)
        notebook.add(tab, text="Constellation label")

        mode_var = tk.StringVar(value="same" if current else "new")
        label_var = tk.StringVar(value=current or "constellation_1")

        ttk.Label(tab, text="Is this still the same constellation?").grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(0, 8),
        )
        ttk.Radiobutton(tab, text="Yes, same constellation", variable=mode_var, value="same").grid(
            row=1,
            column=0,
            columnspan=2,
            sticky="w",
        )
        ttk.Radiobutton(tab, text="No, new constellation", variable=mode_var, value="new").grid(
            row=2,
            column=0,
            columnspan=2,
            sticky="w",
        )
        ttk.Radiobutton(
            tab,
            text="New label, same distances (only LOS/NLOS changes)",
            variable=mode_var,
            value="same_distances",
        ).grid(
            row=3,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(0, 8),
        )
        ttk.Label(tab, text="Constellation label").grid(row=4, column=0, sticky="w", pady=2)
        label_entry = ttk.Entry(tab, textvariable=label_var, width=34)
        label_entry.grid(row=4, column=1, sticky="ew", pady=2)
        label_entry.focus_set()

        button_row = ttk.Frame(tab)
        button_row.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        button_row.columnconfigure(0, weight=1)
        ttk.Button(
            button_row,
            text="Save",
            command=lambda: self._apply_constellation_choice(mode_var.get(), label_var.get(), window),
        ).grid(row=0, column=0, sticky="e", padx=(0, 6))
        ttk.Button(button_row, text="Cancel", command=window.destroy).grid(row=0, column=1, sticky="e")

    def _apply_constellation_choice(self, mode: str, label: str, window: Any) -> None:
        current = self.constellation_var.get().strip()
        label = label.strip()
        if mode == "same" and not label:
            label = current
        if not label:
            messagebox.showerror(
                "Missing constellation label",
                "Enter a constellation label before saving.",
                parent=window,
            )
            return

        if mode == "same":
            self.constellation_var.set(label)
            self._set_constellation_status(label)
            self.log_raw(f"# Constellation confirmed unchanged: {label}")
            window.destroy()
            return

        was_capturing = bool(getattr(self, "capture_active", False))
        if was_capturing:
            if mode == "same_distances":
                question = (
                    "Stop the current session and start a new one with this label? "
                    "True distances are kept; only LOS/NLOS and measured data reset."
                )
            else:
                question = (
                    "Stop the current logging session and start a new one "
                    "with this constellation label?"
                )
            proceed = messagebox.askyesno(
                "New constellation",
                question,
                parent=window,
            )
            if not proceed:
                return
            self.stop_capture()

        if mode == "same_distances":
            self.anchor_measured_distances.clear()
            self.anchor_distance_history.clear()
            self.anchor_los_nlos.clear()
            self._refresh_anchor_truth_table()
            self.log_raw(
                f"# New constellation label: {label} (true distances kept, "
                "LOS/NLOS and measured data reset)"
            )
        else:
            self.log_raw(f"# New constellation label: {label}")

        self.constellation_var.set(label)
        self._set_constellation_status(label)
        window.destroy()

        if was_capturing:
            self.start_capture()

    def _set_constellation_status(self, label: str) -> None:
        if hasattr(self, "constellation_status_var"):
            self.constellation_status_var.set(f"Constellation: {label or 'not set'}")

    def _ensure_anchor_truth_schema(self) -> None:
        store = getattr(self, "store", None)
        if store is None:
            return
        store.conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {ANCHOR_TRUTH_TABLE} (
                session_id TEXT NOT NULL,
                anchor_id TEXT NOT NULL,
                true_distance_m REAL NOT NULL,
                distance_source TEXT NOT NULL,
                pythagorean_side_a_m REAL,
                pythagorean_side_b_m REAL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (session_id, anchor_id),
                FOREIGN KEY(session_id) REFERENCES sessions(id)
            )
            """
        )
        store.conn.commit()

    def _persist_anchor_true_distance(self, anchor_id: str) -> None:
        store = getattr(self, "store", None)
        session_id = getattr(self, "current_session_id", None)
        if store is None or not session_id:
            return
        data = self.anchor_true_distances.get(anchor_id)
        if not data:
            return
        self._ensure_anchor_truth_schema()
        store.conn.execute(
            f"""
            INSERT OR REPLACE INTO {ANCHOR_TRUTH_TABLE} (
                session_id, anchor_id, true_distance_m, distance_source,
                pythagorean_side_a_m, pythagorean_side_b_m, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session_id,
                anchor_id,
                data["true_distance_m"],
                data["distance_source"],
                data.get("pythagorean_side_a_m"),
                data.get("pythagorean_side_b_m"),
                original.now_local_iso(),
            ),
        )
        store.conn.commit()

    def _persist_all_anchor_true_distances(self) -> None:
        for anchor_id in list(self.anchor_true_distances):
            self._persist_anchor_true_distance(anchor_id)

    def _ensure_responder_label_schema(self) -> None:
        store = getattr(self, "store", None)
        if store is None:
            return
        ensure_responder_label_schema(store.conn)

    def _persist_anchor_los_nlos(self, anchor_id: str) -> None:
        store = getattr(self, "store", None)
        session_id = getattr(self, "current_session_id", None)
        if store is None or not session_id:
            return
        self._ensure_responder_label_schema()
        store.conn.execute(
            f"""
            INSERT OR REPLACE INTO {RESPONDER_LABEL_TABLE} (
                session_id, anchor_id, ground_truth_los_nlos, updated_at
            )
            VALUES (?, ?, ?, ?)
            """,
            (
                session_id,
                anchor_id,
                self._anchor_los_nlos_for(anchor_id),
                original.now_local_iso(),
            ),
        )
        store.conn.commit()

    def _persist_all_anchor_los_nlos(self) -> None:
        for anchor_id in list(self.anchor_los_nlos):
            self._persist_anchor_los_nlos(anchor_id)

    def _anchor_true_distance_for(self, anchor_id: Any) -> dict[str, Any] | None:
        if anchor_id not in (None, ""):
            data = self.anchor_true_distances.get(str(anchor_id).strip())
            if data:
                return data
        return None

    def _anchor_los_nlos_for(self, anchor_id: Any) -> str:
        if anchor_id in (None, ""):
            return UNKNOWN_LOS_NLOS
        return normalize_los_nlos(self.anchor_los_nlos.get(str(anchor_id).strip(), UNKNOWN_LOS_NLOS))

    def start_capture(self) -> None:
        super().start_capture()
        if getattr(self, "store", None) is None or not getattr(self, "current_session_id", None):
            return
        self._ensure_anchor_truth_schema()
        self._ensure_responder_label_schema()
        self._persist_all_anchor_true_distances()
        self._persist_all_anchor_los_nlos()

    def disconnect_transport(self) -> None:
        self.stop_live_tracking("Live tracking stopped; Bluetooth is disconnected.")
        if self.anchor_survey_in_flight:
            self._finish_anchor_pair_survey("Anchor pair survey stopped; Bluetooth is disconnected.")
        super().disconnect_transport()

    def _handle_command_result(self, packet: Any) -> None:
        live_fast_result = (
            self.live_tracking_active
            and self.ml_pending_session == packet.session_id
            and self.ml_pending_seq == packet.sequence
            and self.ml_pending_mode == original.ML_COLLECTION_MODE_FAST
        )
        super()._handle_command_result(packet)
        if live_fast_result:
            self._request_live_tracking_batch_finish()
        self._handle_anchor_survey_command_result(packet)

    def handle_serial_line(self, line: str) -> None:
        super().handle_serial_line(line)

    def handle_records(self, records: list[Any]) -> None:
        normal_records: list[Any] = []
        for record in records:
            if str(getattr(record, "kind", "")).startswith("survey_pair"):
                self._handle_anchor_pair_record(record)
            else:
                normal_records.append(record)
        if not normal_records:
            return

        localization_updated = False
        for record in normal_records:
            if record.anchor_id:
                anchor_key = str(record.anchor_id).strip()
                self._register_detected_anchor(record.anchor_id)
                if self.live_tracking_active and record.kind in ("sample", "failure", "summary"):
                    self._record_live_tracking_sample(record)
                    continue
                if record.distance_m is not None:
                    self.anchor_measured_distances[anchor_key] = float(record.distance_m)
                    history = self.anchor_distance_history.setdefault(anchor_key, [])
                    history.append(float(record.distance_m))
                    if len(history) > 200:
                        del history[: len(history) - 200]
                    self._update_localization_range(anchor_key, float(record.distance_m))
                    localization_updated = True
        if self.live_tracking_finish_pending:
            self._finish_pending_live_tracking_batch()
        if localization_updated and self.live_tracking_active:
            self._try_live_tracking_solve()
        self._refresh_anchor_truth_table()
        super().handle_records(normal_records)

    def check_los_alert(self, record: Any) -> float | None:
        distance = self._record_range_m(record)
        if distance is None:
            return None

        truth = self._anchor_true_distance_for(record.anchor_id)
        if not truth:
            return None

        true_distance = truth["true_distance_m"]
        absolute_error = abs(float(distance) - float(true_distance))
        threshold = safe_float(self.threshold_var.get()) or 0.0

        if (
            getattr(self, "store", None) is not None
            and getattr(self, "current_session_id", None)
            and self._anchor_los_nlos_for(record.anchor_id) == "LOS"
            and absolute_error > threshold
        ):
            anchor = record.anchor_id or "unknown"
            message = (
                f"LOS measurement from anchor {anchor} is {absolute_error:.3f} m away "
                f"from true distance ({float(distance):.3f} m measured vs "
                f"{float(true_distance):.3f} m true)."
            )
            self.store.insert_alert(
                self.current_session_id,
                anchor,
                float(distance),
                float(true_distance),
                absolute_error,
                message,
            )
            self.log_alert(message)

            current_time = time.monotonic()
            last_prompt = self.last_alert_prompt.get(anchor, 0.0)
            if current_time - last_prompt > original.ALERT_PROMPT_COOLDOWN_SECONDS:
                self.last_alert_prompt[anchor] = current_time
                messagebox.showwarning("LOS measurement warning", message, parent=self)
        return absolute_error

    def check_instability_alert(self, record: Any) -> float | None:
        if record.kind == "summary" or record.distance_m is None:
            return None
        threshold = safe_float(self.instability_threshold_var.get())
        if threshold is None or threshold <= 0:
            return None
        distance = self._apply_range_offset(record.distance_m)
        if distance is None:
            return None
        anchor = record.anchor_id or "unknown"
        values = self.anchor_distance_windows.setdefault(anchor, [])
        values.append(distance)
        if len(values) > original.INSTABILITY_WINDOW_SIZE:
            del values[0 : len(values) - original.INSTABILITY_WINDOW_SIZE]
        if len(values) < min(5, original.INSTABILITY_WINDOW_SIZE):
            return None
        mean = sum(values) / len(values)
        std = math.sqrt(sum((value - mean) ** 2 for value in values) / len(values))
        if std > threshold and self.store is not None and self.current_session_id is not None:
            now = time.monotonic()
            last = self.last_instability_prompt.get(anchor, 0.0)
            if now - last > original.ALERT_PROMPT_COOLDOWN_SECONDS:
                self.last_instability_prompt[anchor] = now
                message = f"Anchor {anchor} distance is unstable: std {std:.3f} m over {len(values)} samples."
                self.store.insert_alert(self.current_session_id, anchor, distance, None, std, message)
                self.log_alert(message)
        return std

    def add_record_to_table(self, record: Any, alert_error: float | None = None) -> None:
        display_time = datetime.now().strftime("%H:%M:%S")
        distance = self._record_range_m(record)
        truth = self._anchor_true_distance_for(record.anchor_id)
        error = alert_error
        if error is None and truth and distance is not None:
            error = abs(float(distance) - float(truth["true_distance_m"]))

        row = (
            display_time,
            record.anchor_id or "",
            "" if record.sample_index is None else str(record.sample_index),
            "" if distance is None else f"{float(distance):.3f}",
            "" if error is None else f"{float(error):.3f}",
            "" if record.rx_power_dbm is None else f"{float(record.rx_power_dbm):.2f}",
            "" if record.fp_power_dbm is None else f"{float(record.fp_power_dbm):.2f}",
            "" if record.cir_power is None else f"{float(record.cir_power):.2f}",
            record.status or "",
            record.source or "",
        )

        tags: tuple[str, ...] = ()
        threshold = safe_float(self.threshold_var.get())
        if error is not None and threshold is not None and error > threshold:
            tags = ("alert",)
        if record.kind == "failure":
            tags = ("failure",)
        elif record.kind == "summary":
            tags = ("summary",)

        self.sample_table.insert("", tk.END, values=row, tags=tags)
        self.table_rows += 1
        if self.table_rows > original.MAX_TABLE_ROWS:
            first = self.sample_table.get_children()[0]
            self.sample_table.delete(first)
            self.table_rows -= 1
        self.sample_table.yview_moveto(1.0)

    def export_current_session(self) -> None:
        if getattr(self, "store", None) is None or not getattr(self, "current_session_id", None):
            messagebox.showwarning(
                "No session",
                "Start a capture session before exporting.",
                parent=self,
            )
            return

        if getattr(self, "capture_active", False):
            proceed = messagebox.askyesno(
                "Session still running",
                "The session is still capturing. Export a snapshot of the data collected so far?",
                parent=self,
            )
            if not proceed:
                return

        self._ensure_anchor_truth_schema()
        self._ensure_responder_label_schema()
        self._persist_all_anchor_true_distances()
        self._persist_all_anchor_los_nlos()

        output_dir = filedialog.askdirectory(
            initialdir=self.output_dir_var.get(),
            title="Choose folder for per-anchor Excel exports",
            parent=self,
        )
        if not output_dir:
            return

        try:
            paths = export_session_per_anchor(
                self.store.conn,
                self.current_session_id,
                Path(output_dir),
                range_static_offset_m=self._range_static_offset_m(),
            )
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc), parent=self)
            return

        if not paths:
            messagebox.showwarning(
                "No measurements",
                "The current session does not contain anchor measurements yet.",
                parent=self,
            )
            return

        messagebox.showinfo(
            "Export complete",
            "Wrote:\n" + "\n".join(str(path) for path in paths),
            parent=self,
        )

    def export_measurement_list(self) -> None:
        db_path = Path(self.db_path_var.get()).expanduser()
        if not db_path.exists():
            messagebox.showwarning(
                "No measurements",
                "No SQLite capture database exists yet. Start logging first.",
                parent=self,
            )
            return

        rows = build_measurement_rows(db_path, range_static_offset_m=self._range_static_offset_m())
        if not rows:
            messagebox.showwarning(
                "No measurements",
                "The current capture database does not contain measurement rows yet.",
                parent=self,
            )
            return

        default_name = measurement_list_default_filename(db_path)
        output_path = filedialog.asksaveasfilename(
            initialdir=self.output_dir_var.get(),
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            parent=self,
        )
        if not output_path:
            return

        try:
            write_measurement_workbook(rows, Path(output_path))
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc), parent=self)
            return

        messagebox.showinfo("Export complete", f"Wrote {output_path}", parent=self)


def table_exists(con: sqlite3.Connection, table_name: str) -> bool:
    row = con.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def fetch_anchor_truths(con: sqlite3.Connection, session_id: str) -> dict[str, dict[str, Any]]:
    if not table_exists(con, ANCHOR_TRUTH_TABLE):
        return {}
    rows = con.execute(
        f"""
        SELECT anchor_id, true_distance_m, distance_source,
               pythagorean_side_a_m, pythagorean_side_b_m
        FROM {ANCHOR_TRUTH_TABLE}
        WHERE session_id = ?
        """,
        (session_id,),
    ).fetchall()
    return {
        str(row["anchor_id"]): {
            "anchor_id": str(row["anchor_id"]),
            "true_distance_m": row["true_distance_m"],
            "distance_source": row["distance_source"],
            "pythagorean_side_a_m": row["pythagorean_side_a_m"],
            "pythagorean_side_b_m": row["pythagorean_side_b_m"],
        }
        for row in rows
    }


def fetch_anchor_los_nlos(con: sqlite3.Connection, session_id: str) -> dict[str, str]:
    if not table_exists(con, RESPONDER_LABEL_TABLE):
        return {}
    rows = con.execute(
        f"""
        SELECT anchor_id, ground_truth_los_nlos
        FROM {RESPONDER_LABEL_TABLE}
        WHERE session_id = ?
        """,
        (session_id,),
    ).fetchall()
    return {
        str(row["anchor_id"]): normalize_los_nlos(row["ground_truth_los_nlos"])
        for row in rows
    }


def session_anchor_ids(con: sqlite3.Connection, session_id: str) -> list[str]:
    rows = con.execute(
        """
        SELECT anchor_id FROM samples WHERE session_id = ? AND anchor_id IS NOT NULL
        UNION
        SELECT anchor_id FROM summaries WHERE session_id = ? AND anchor_id IS NOT NULL
        ORDER BY anchor_id
        """,
        (session_id, session_id),
    ).fetchall()
    return [str(row["anchor_id"]) for row in rows if row["anchor_id"] not in (None, "")]


def measured_mean_for_anchor(
    samples: list[sqlite3.Row],
    summaries: list[sqlite3.Row],
    *,
    range_static_offset_m: float = 0.0,
) -> float | None:
    values = [
        corrected
        for row in samples
        for corrected in [apply_range_offset(row["distance_m"], range_static_offset_m)]
        if corrected is not None
    ]
    if values:
        return statistics.fmean(values)
    summary_values = [
        corrected
        for row in summaries
        for corrected in [apply_range_offset(row["mean_distance_m"], range_static_offset_m)]
        if corrected is not None
    ]
    if summary_values:
        return summary_values[-1]
    return None


def export_session_per_anchor(
    con: sqlite3.Connection,
    session_id: str,
    output_dir: Path,
    *,
    timestamp: str | None = None,
    range_static_offset_m: float = 0.0,
) -> list[Path]:
    session = con.execute("SELECT * FROM sessions WHERE id = ?", (session_id,)).fetchone()
    if session is None:
        raise ValueError(f"Unknown session id: {session_id}")

    truths = fetch_anchor_truths(con, session_id)
    los_nlos_by_anchor = fetch_anchor_los_nlos(con, session_id)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = timestamp or export_timestamp()
    constellation = safe_filename(session["constellation_label"] or "constellation")

    paths: list[Path] = []
    for anchor_id in session_anchor_ids(con, session_id):
        samples = con.execute(
            "SELECT * FROM samples WHERE session_id = ? AND anchor_id = ? ORDER BY id",
            (session_id, anchor_id),
        ).fetchall()
        summaries = con.execute(
            "SELECT * FROM summaries WHERE session_id = ? AND anchor_id = ? ORDER BY id",
            (session_id, anchor_id),
        ).fetchall()
        alerts = con.execute(
            "SELECT * FROM alerts WHERE session_id = ? AND anchor_id = ? ORDER BY id",
            (session_id, anchor_id),
        ).fetchall()
        raw_lines = con.execute(
            "SELECT id, timestamp, parsed, raw_line FROM raw_lines WHERE session_id = ? ORDER BY id",
            (session_id,),
        ).fetchall()

        truth = truths.get(anchor_id)
        los_nlos = los_nlos_by_anchor.get(anchor_id, UNKNOWN_LOS_NLOS)
        mean_measured = measured_mean_for_anchor(
            samples,
            summaries,
            range_static_offset_m=range_static_offset_m,
        )

        filename = (
            f"CONSTELLATION_{constellation}_"
            f"ANCHOR_{safe_filename(anchor_id)}_"
            f"TRUE_{format_meter(truth['true_distance_m'] if truth else None)}m_"
            f"MEASURED_{format_meter(mean_measured)}m_"
            f"{safe_filename(los_nlos)}_{timestamp}.xlsx"
        )
        path = output_dir / filename
        write_anchor_session_workbook(
            session,
            anchor_id,
            truth,
            los_nlos,
            samples,
            summaries,
            alerts,
            raw_lines,
            mean_measured,
            path,
            range_static_offset_m=range_static_offset_m,
        )
        paths.append(path)
    return paths


def write_anchor_session_workbook(
    session: sqlite3.Row,
    anchor_id: str,
    truth: dict[str, Any] | None,
    los_nlos: str,
    samples: list[sqlite3.Row],
    summaries: list[sqlite3.Row],
    alerts: list[sqlite3.Row],
    raw_lines: list[sqlite3.Row],
    mean_measured: float | None,
    path: Path,
    *,
    range_static_offset_m: float = 0.0,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Samples"

    true_distance = truth["true_distance_m"] if truth else None
    true_source = truth["distance_source"] if truth else ""
    side_a = truth.get("pythagorean_side_a_m") if truth else None
    side_b = truth.get("pythagorean_side_b_m") if truth else None
    los_nlos = normalize_los_nlos(los_nlos)

    columns = [
        "timestamp",
        "anchor_id",
        "clicker_id",
        "sample_index",
        "event_seq",
        "scheduled_sample_count",
        "measured_uwb_distance_m",
        "true_ground_truth_distance_m",
        "true_distance_source",
        "pythagorean_side_a_m",
        "pythagorean_side_b_m",
        "los_nlos",
        "rx_dbm",
        "fp_dbm",
        "cir_power",
        "quality",
        "firmware_timestamp_ms",
        "phy_config_id",
        "burst_id",
        "alert_metric_m",
        "status",
        "source",
        "cir_raw",
        "tlv_json",
        "raw_line",
    ]
    ws.append(columns)

    for row in samples:
        measured = apply_range_offset(row["distance_m"], range_static_offset_m)
        alert_metric = (
            abs(float(measured) - float(true_distance))
            if measured is not None and true_distance is not None
            else None
        )
        ws.append(
            excel_row([
                row["timestamp"],
                row["anchor_id"],
                row["clicker_id"],
                row["sample_index"],
                row["event_seq"],
                row["scheduled_sample_count"],
                measured,
                true_distance,
                true_source,
                side_a,
                side_b,
                los_nlos,
                row["rx_power_dbm"],
                row["fp_power_dbm"],
                row["cir_power"],
                row["quality"],
                row["firmware_timestamp_ms"],
                row["phy_config_id"],
                row["burst_id"],
                alert_metric,
                row["status"],
                row["source"],
                row["cir_raw"],
                row["tlv_json"],
                row["raw_line"],
            ])
        )

    add_rows_sheet(
        wb,
        "Summaries",
        [
            "timestamp",
            "anchor_id",
            "mean_measured_uwb_distance_m",
            "true_ground_truth_distance_m",
            "true_distance_source",
            "pythagorean_side_a_m",
            "pythagorean_side_b_m",
            "los_nlos",
            "alert_metric_m",
            "good_count",
            "status",
            "source",
            "raw_line",
        ],
        [
            [
                row["timestamp"],
                row["anchor_id"],
                apply_range_offset(row["mean_distance_m"], range_static_offset_m),
                true_distance,
                true_source,
                side_a,
                side_b,
                los_nlos,
                abs(float(apply_range_offset(row["mean_distance_m"], range_static_offset_m)) - float(true_distance))
                if row["mean_distance_m"] is not None and true_distance is not None
                else None,
                row["good_count"],
                row["status"],
                row["source"],
                row["raw_line"],
            ]
            for row in summaries
        ],
    )

    add_rows_sheet(
        wb,
        "Anchor True Distance",
        ["anchor_id", "true_distance_m", "distance_source", "pythagorean_side_a_m", "pythagorean_side_b_m"],
        [[anchor_id, true_distance, true_source, side_a, side_b]],
    )
    add_rows_sheet(
        wb,
        "Session",
        ["field", "value"],
        [
            ["session_id", session["id"]],
            ["started_at", session["started_at"]],
            ["stopped_at", session["stopped_at"]],
            ["constellation_label", session["constellation_label"]],
            ["session_condition_hint", condition_text(session)],
            ["los_nlos", los_nlos],
            ["anchor_id", anchor_id],
            ["mean_measured_uwb_distance_m", mean_measured],
            ["true_ground_truth_distance_m", true_distance],
            ["true_distance_source", true_source],
        ],
    )
    add_rows_sheet(wb, "Alerts", list(alerts[0].keys()) if alerts else ["No rows"], alerts)
    add_rows_sheet(wb, "Raw Lines", list(raw_lines[0].keys()) if raw_lines else ["No rows"], raw_lines)

    style_workbook(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def add_rows_sheet(
    wb: openpyxl.Workbook,
    title: str,
    columns: list[str],
    rows: list[Any],
) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    if not rows:
        return
    for row in rows:
        if isinstance(row, sqlite3.Row):
            ws.append(excel_row([row[col] for col in columns]))
        else:
            ws.append(excel_row(list(row)))


def style_workbook(wb: openpyxl.Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            header = column_cells[0].value
            if not header:
                continue
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 42)


def build_measurement_rows(
    db_path: Path,
    *,
    range_static_offset_m: float = 0.0,
) -> list[dict[str, Any]]:
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    try:
        sessions = con.execute(
            """
            SELECT *
            FROM sessions
            ORDER BY started_at, id
            """
        ).fetchall()

        sample_rows = con.execute(
            """
            SELECT session_id, anchor_id, distance_m
            FROM samples
            WHERE distance_m IS NOT NULL
            """
        ).fetchall()

        summary_rows = con.execute(
            """
            SELECT session_id, anchor_id, mean_distance_m, good_count, status, source
            FROM summaries
            WHERE mean_distance_m IS NOT NULL
            ORDER BY timestamp
            """
        ).fetchall()

        truth_rows = (
            con.execute(
                f"""
                SELECT session_id, anchor_id, true_distance_m, distance_source,
                       pythagorean_side_a_m, pythagorean_side_b_m
                FROM {ANCHOR_TRUTH_TABLE}
                """
            ).fetchall()
            if table_exists(con, ANCHOR_TRUTH_TABLE)
            else []
        )

        label_rows = (
            con.execute(
                f"""
                SELECT session_id, anchor_id, ground_truth_los_nlos
                FROM {RESPONDER_LABEL_TABLE}
                """
            ).fetchall()
            if table_exists(con, RESPONDER_LABEL_TABLE)
            else []
        )
    finally:
        con.close()

    samples_by_key: dict[tuple[str, str], list[float]] = {}
    for row in sample_rows:
        anchor_id = row["anchor_id"] or ""
        key = (row["session_id"], anchor_id)
        corrected = apply_range_offset(row["distance_m"], range_static_offset_m)
        if corrected is not None:
            samples_by_key.setdefault(key, []).append(corrected)

    latest_summary_by_key: dict[tuple[str, str], sqlite3.Row] = {}
    for row in summary_rows:
        anchor_id = row["anchor_id"] or ""
        latest_summary_by_key[(row["session_id"], anchor_id)] = row

    truth_by_key: dict[tuple[str, str], sqlite3.Row] = {}
    for row in truth_rows:
        anchor_id = row["anchor_id"] or ""
        truth_by_key[(row["session_id"], anchor_id)] = row

    label_by_key: dict[tuple[str, str], str] = {}
    for row in label_rows:
        anchor_id = row["anchor_id"] or ""
        label_by_key[(row["session_id"], anchor_id)] = normalize_los_nlos(row["ground_truth_los_nlos"])

    session_by_id = {row["id"]: row for row in sessions}
    keys = sorted(
        set(samples_by_key) | set(latest_summary_by_key),
        key=lambda item: (session_by_id[item[0]]["started_at"], item[1]),
    )

    rows: list[dict[str, Any]] = []
    for session_id, anchor_id in keys:
        session = session_by_id[session_id]
        values = samples_by_key.get((session_id, anchor_id), [])
        latest_summary = latest_summary_by_key.get((session_id, anchor_id))
        truth = truth_by_key.get((session_id, anchor_id))
        mean_distance = (
            apply_range_offset(latest_summary["mean_distance_m"], range_static_offset_m)
            if latest_summary is not None
            else statistics.fmean(values)
            if values
            else None
        )
        rows.append(
            {
                "session_id": session_id,
                "started_at": session["started_at"],
                "stopped_at": session["stopped_at"],
                "constellation_label": session["constellation_label"],
                "condition_hint": condition_text(session),
                "tag_id": session["tag_id"] or "",
                "anchor_id": anchor_id,
                "uwb_mean_distance_m": mean_distance,
                "sample_count": len(values),
                "sample_std_m": statistics.stdev(values) if len(values) > 1 else None,
                "true_ground_truth_distance_m": truth["true_distance_m"] if truth is not None else None,
                "true_distance_source": truth["distance_source"] if truth is not None else "",
                "pythagorean_side_a_m": truth["pythagorean_side_a_m"] if truth is not None else None,
                "pythagorean_side_b_m": truth["pythagorean_side_b_m"] if truth is not None else None,
                "session_ground_truth_m": session["ground_truth_m"],
                "outlier_threshold_m": session["outlier_threshold_m"],
                "ground_truth_los_nlos": label_by_key.get((session_id, anchor_id), UNKNOWN_LOS_NLOS),
                "status": latest_summary["status"] if latest_summary is not None else "",
                "source": latest_summary["source"] if latest_summary is not None else "",
            }
        )
    return rows


def condition_text(session: sqlite3.Row) -> str:
    bits = []
    if session["condition_los"]:
        bits.append("LOS")
    if session["condition_nlos"]:
        bits.append("NLOS")
    return "/".join(bits)


def write_measurement_workbook(rows: list[dict[str, Any]], path: Path) -> None:
    columns = [
        "session_id",
        "started_at",
        "stopped_at",
        "constellation_label",
        "condition_hint",
        "tag_id",
        "anchor_id",
        "true_ground_truth_distance_m",
        "true_distance_source",
        "pythagorean_side_a_m",
        "pythagorean_side_b_m",
        "uwb_mean_distance_m",
        "range_minus_true_m",
        "ground_truth_los_nlos",
        "session_ground_truth_m",
        "outlier_threshold_m",
        "sample_count",
        "sample_std_m",
        "status",
        "source",
        "notes",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Measurements"
    ws.append(columns)

    for row in rows:
        ws.append([row.get(col, "") if row.get(col) is not None else "" for col in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    computed_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    col_index = {name: idx + 1 for idx, name in enumerate(columns)}
    editable_cols = [
        "tag_id",
        "ground_truth_los_nlos",
        "notes",
    ]
    computed_cols = [
        "range_minus_true_m",
    ]

    for row_idx in range(2, len(rows) + 2):
        true_distance = ws.cell(row_idx, col_index["true_ground_truth_distance_m"]).coordinate
        uwb = ws.cell(row_idx, col_index["uwb_mean_distance_m"]).coordinate
        ws.cell(row_idx, col_index["range_minus_true_m"]).value = (
            f'=IF(OR({true_distance}="",ISBLANK({uwb})),"",{uwb}-{true_distance})'
        )

    for col_name in editable_cols:
        col = col_index[col_name]
        for row_idx in range(2, max(len(rows) + 2, 200)):
            ws.cell(row_idx, col).fill = editable_fill

    for col_name in computed_cols:
        col = col_index[col_name]
        for row_idx in range(2, max(len(rows) + 2, 200)):
            ws.cell(row_idx, col).fill = computed_fill

    validation = DataValidation(type="list", formula1='"LOS,NLOS,Unknown"', allow_blank=True)
    ws.add_data_validation(validation)
    gt_col_letter = openpyxl.utils.get_column_letter(col_index["ground_truth_los_nlos"])
    validation.add(f"{gt_col_letter}2:{gt_col_letter}1000")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col_name in enumerate(columns, start=1):
        width = min(max(len(col_name) + 2, 12), 34)
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    info = wb.create_sheet("Instructions")
    info.append(["Field", "How to use it"])
    info.append(["Yellow cells", "Review or adjust after the capture session: tag ID, LOS/NLOS ground truth, and notes."])
    info.append(["true_ground_truth_distance_m", "Per-anchor true tag-to-anchor distance saved in the Measurement Session controls."])
    info.append(["true_distance_source", "direct laser or pythagorean calculation."])
    info.append(["range_minus_true_m", "UWB mean distance minus the saved per-anchor true distance."])
    info.append(["Constellation", "Use the GUI's Constellation Changed button whenever the anchor layout changes."])
    info.column_dimensions["A"].width = 28
    info.column_dimensions["B"].width = 120
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    app = ExtendedUwbCaptureApp()
    app.mainloop()


if __name__ == "__main__":
    main()
